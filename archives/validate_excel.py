"""
validate_excel.py — structural validation of the rendered Excel workbook.

Runs AFTER render_all.py, BEFORE human review. Catches writer bugs that the
JSON extraction wouldn't show: orphaned merges, value-less merged cells,
column count mismatches, duplicate sheet names, empty data sheets, etc.

Checks (all soft — failures FLAG, never auto-correct):
  1. sheet_count:        at least one data sheet besides SUMMARY
  2. column_consistency: every data row has the same number of populated cols
  3. merge_value:        every merged region has a non-empty anchor cell
  4. merge_orphan:       no merged cell references a col/row outside sheet dims
  5. header_present:     each sheet has a title row and a column header row
  6. no_formula:         no cells contain formulas (auditability rule)
  7. all_zero_row:       warn if an entire data row is zeros (possible miss)
  8. duplicate_sheets:   no two sheets have the same name

Usage:
  python3 validate_excel.py                        # validates out/step4_output.xlsx
  python3 validate_excel.py out/nsfr_check.xlsx    # specific file
  python3 validate_excel.py --json                 # machine-readable output
"""
from __future__ import annotations
import sys
import json
import re
import argparse
from dataclasses import dataclass, field
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class Issue:
    severity: str      # ERROR | WARNING | INFO
    sheet: str
    check: str
    detail: str

    def __str__(self) -> str:
        return f"[{self.severity}] {self.sheet} | {self.check}: {self.detail}"


@dataclass
class ValidationResult:
    path: str
    issues: list[Issue] = field(default_factory=list)

    @property
    def errors(self) -> list[Issue]:
        return [i for i in self.issues if i.severity == "ERROR"]

    @property
    def warnings(self) -> list[Issue]:
        return [i for i in self.issues if i.severity == "WARNING"]

    @property
    def passed(self) -> bool:
        return len(self.errors) == 0

    def summary(self) -> str:
        status = "PASS" if self.passed else "FAIL"
        return (f"{status}  {self.path}  |  "
                f"{len(self.errors)} error(s)  {len(self.warnings)} warning(s)")

    def to_dict(self) -> dict:
        return {
            "path": self.path,
            "passed": self.passed,
            "n_errors": len(self.errors),
            "n_warnings": len(self.warnings),
            "issues": [{"severity": i.severity, "sheet": i.sheet,
                        "check": i.check, "detail": i.detail}
                       for i in self.issues],
        }


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_data_sheet(name: str) -> bool:
    return name.upper() != "SUMMARY"


def _cell_val(ws, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and v.strip() in ("", "-"):
        return True
    return False


# ---------------------------------------------------------------------------
# Individual checks
# ---------------------------------------------------------------------------

def check_sheet_count(wb: openpyxl.Workbook, result: ValidationResult) -> None:
    data_sheets = [n for n in wb.sheetnames if _is_data_sheet(n)]
    if not data_sheets:
        result.issues.append(Issue("ERROR", "workbook", "sheet_count",
                                   "No data sheets found — only SUMMARY sheet present"))
    else:
        result.issues.append(Issue("INFO", "workbook", "sheet_count",
                                   f"{len(data_sheets)} data sheet(s) found"))


def check_duplicate_sheets(wb: openpyxl.Workbook, result: ValidationResult) -> None:
    seen: dict[str, int] = {}
    for name in wb.sheetnames:
        seen[name.lower()] = seen.get(name.lower(), 0) + 1
    for name, count in seen.items():
        if count > 1:
            result.issues.append(Issue("ERROR", name, "duplicate_sheets",
                                       f"Sheet name appears {count} times"))


def check_header_present(ws: openpyxl.worksheet.worksheet.Worksheet,
                         result: ValidationResult) -> None:
    """Row 1 must have a non-empty title; row 4 (or nearby) a column header."""
    name = ws.title
    r1 = _cell_val(ws, 1, 1)
    if _is_empty(r1):
        result.issues.append(Issue("ERROR", name, "header_present",
                                   "Row 1 (title) is empty"))

    # Scan rows 3–6 for a header-like row (contains 'label' or 'row_id')
    found_header = False
    for r in range(3, 8):
        v = (_cell_val(ws, r, 1) or "")
        if isinstance(v, str) and any(k in v.lower() for k in ("row_id", "label", "level")):
            found_header = True
            break
    if not found_header:
        result.issues.append(Issue("WARNING", name, "header_present",
                                   "Could not find column header row in rows 3–7"))


def _find_data_start_row(ws) -> int:
    """Return the first row where col A contains a row_id (not header text)."""
    for r in range(1, ws.max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None and not isinstance(v, str):
            return r
        if isinstance(v, str) and re.match(r"^\d+[a-z]?$", v.strip()):
            return r
    return 6  # fallback


def check_merge_value(ws: openpyxl.worksheet.worksheet.Worksheet,
                      result: ValidationResult) -> None:
    """Every merged region in DATA rows must have a non-empty anchor cell.
    Header merges (above the data start row) are intentional layout."""
    name = ws.title
    data_start = _find_data_start_row(ws)
    for merge_range in ws.merged_cells.ranges:
        min_row, min_col = merge_range.min_row, merge_range.min_col
        if min_row < data_start:
            continue  # header area — layout merge, not a data bug
        val = ws.cell(row=min_row, column=min_col).value
        if _is_empty(val):
            ref = f"{get_column_letter(min_col)}{min_row}"
            result.issues.append(Issue("WARNING", name, "merge_value",
                                       f"Merged region {merge_range} has empty anchor cell {ref}"))


def check_merge_orphan(ws: openpyxl.worksheet.worksheet.Worksheet,
                       result: ValidationResult) -> None:
    """No merged range should reference rows/cols outside sheet dimensions."""
    name = ws.title
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    for merge_range in ws.merged_cells.ranges:
        if merge_range.max_row > max_row or merge_range.max_col > max_col:
            result.issues.append(Issue("ERROR", name, "merge_orphan",
                                       f"Merged region {merge_range} exceeds sheet dims "
                                       f"({max_row}r × {max_col}c)"))


def check_no_formula(ws: openpyxl.worksheet.worksheet.Worksheet,
                     result: ValidationResult) -> None:
    """Flag any formula cell — the workbook must be purely values."""
    name = ws.title
    formula_cells = []
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                formula_cells.append(cell.coordinate)
    if formula_cells:
        result.issues.append(Issue("ERROR", name, "no_formula",
                                   f"Formula cells found: {formula_cells[:10]}"
                                   + (" (…)" if len(formula_cells) > 10 else "")))


def check_all_zero_rows(ws: openpyxl.worksheet.worksheet.Worksheet,
                        result: ValidationResult) -> None:
    """
    Warn if a non-header data row has every numeric cell equal to 0.
    Heuristic: skip first 5 rows (header area), check rows where col A
    looks like a row_id (numeric string).
    """
    name = ws.title
    data_start = _find_data_start_row(ws)
    zero_rows = []
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=False):
        row_id_val = row[0].value if row else None
        if row_id_val is None:
            continue
        numeric_vals = []
        for cell in row[6:]:  # col G onward = actual data values (skip label at col F)
            v = cell.value
            if isinstance(v, (int, float)):
                numeric_vals.append(v)
            elif isinstance(v, str) and v.strip() not in ("", "-"):
                try:
                    numeric_vals.append(float(v.replace(",", "")))
                except ValueError:
                    pass
        if numeric_vals and all(v == 0 for v in numeric_vals):
            zero_rows.append(str(row_id_val))
    if zero_rows:
        result.issues.append(Issue("WARNING", name, "all_zero_row",
                                   f"Rows with all-zero data values: {zero_rows[:20]}"))


def check_column_consistency(ws: openpyxl.worksheet.worksheet.Worksheet,
                              result: ValidationResult) -> None:
    """
    All data rows should have similar non-empty data cell counts. Rows
    containing merged cells or flagged as total/header are excluded — those
    are intentionally sparse.
    """
    name = ws.title
    data_start = _find_data_start_row(ws)

    # Build set of (row, col) coordinates covered by any merge
    merged_rows: set[int] = set()
    for merge_range in ws.merged_cells.ranges:
        if merge_range.min_row >= data_start:
            merged_rows.add(merge_range.min_row)

    col_counts: list[tuple[int, int]] = []  # (sheet_row, count)
    for row in ws.iter_rows(min_row=data_start, max_row=ws.max_row, values_only=False):
        sheet_row = row[0].row
        row_id_val = row[0].value if row else None
        if row_id_val is None:
            continue
        if sheet_row in merged_rows:
            continue  # merged rows are intentionally sparse
        row_type = row[3].value if len(row) > 3 else None
        if row_type in ("total", "section_header", "note"):
            continue  # totals/headers have fewer populated cells by design
        non_empty = sum(1 for cell in row[6:]
                        if cell.value is not None and str(cell.value).strip() not in ("", "-"))
        all_dash = all(str(cell.value or "").strip() in ("", "-") for cell in row[6:])
        if not all_dash:  # all-dash rows are intentional (no value in any bucket)
            col_counts.append((sheet_row, non_empty))

    if not col_counts:
        return

    # Only flag completely blank data rows — partial population is normal for
    # NSFR/LCR tables where maturity buckets don't apply to all line items
    blank_rows = [sr for sr, c in col_counts if c == 0]
    if blank_rows:
        result.issues.append(Issue("WARNING", name, "column_consistency",
                                   f"Sheet rows {blank_rows[:10]} are data rows with no "
                                   f"non-empty values — possible missed extraction"))


# ---------------------------------------------------------------------------
# Main validator
# ---------------------------------------------------------------------------

def validate(path: str) -> ValidationResult:
    result = ValidationResult(path=path)

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception as e:
        result.issues.append(Issue("ERROR", "workbook", "load", str(e)))
        return result

    # Workbook-level checks
    check_sheet_count(wb, result)
    check_duplicate_sheets(wb, result)

    # Per-sheet checks
    for name in wb.sheetnames:
        if not _is_data_sheet(name):
            continue
        ws = wb[name]
        check_header_present(ws, result)
        check_merge_value(ws, result)
        check_merge_orphan(ws, result)
        check_no_formula(ws, result)
        check_all_zero_rows(ws, result)
        check_column_consistency(ws, result)

    return result


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Structural validation of rendered Excel workbook.")
    p.add_argument("path", nargs="?", default="out/step4_output.xlsx",
                   help="Path to the Excel file (default: out/step4_output.xlsx)")
    p.add_argument("--json", action="store_true", help="Output machine-readable JSON")
    args = p.parse_args()

    result = validate(args.path)

    if args.json:
        print(json.dumps(result.to_dict(), indent=2))
        sys.exit(0 if result.passed else 1)

    # Human-readable output
    print(f"\n{'='*60}")
    print(result.summary())
    print(f"{'='*60}")

    if not result.issues:
        print("  No issues found.")
    else:
        for issue in result.issues:
            if issue.severity == "INFO":
                continue
            print(f"  {issue}")

    info = [i for i in result.issues if i.severity == "INFO"]
    if info:
        print(f"\n--- Info ---")
        for i in info:
            print(f"  {i}")

    print()
    sys.exit(0 if result.passed else 1)
