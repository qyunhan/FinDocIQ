"""
compare_excel.py — values-only accuracy comparator. ZERO API calls.

Scores a PRODUCED workbook (from extract_to_excel.py) against a GOLDEN workbook
on the thing that actually matters — the cell VALUES — while ignoring everything
that legitimately differs between the two: tab names, sheet order, banner/title
rows, formatting, and the hierarchy metadata columns.

How it stays format-agnostic:
  * Each data cell is keyed by (normalised row label, normalised column header),
    NOT by spreadsheet position. So reordered rows, extra banner rows, or a
    different column layout don't register as errors.
  * Column headers that are dates are canonicalised ("31 Dec 25" == "31 Dec 2025").
  * Values are compared numerically (commas stripped, (1,234)=-1234, "17.0%"=17.0,
    "-"/blank/"NM" treated as no-value), with a small tolerance for rounding.
  * Tabs are matched between the two books by title-word overlap (so golden
    "A2 - Key Prudential Metrics" matches produced "A.4 - Overview of Key Prudential").

Output: per-tab and overall value accuracy + a list of real value mismatches /
missing / extra cells. A low score here means WRONG VALUES, not format drift.

Usage:
  python compare_excel.py --produced out/dbs_sections.xlsx --golden "/path/DBS_4Q25_Pillar3.xlsx"
  python compare_excel.py --produced out/sections.xlsx --golden golden.xlsx --out out/score.json
  python compare_excel.py ... --tol 0.51        # numeric match tolerance (abs)
"""
from __future__ import annotations
import os, sys, re, json, argparse
import openpyxl

META_KEYS = {"uniquerowid", "hierarchylevel", "parentrowid", "rowid", "level", "parent"}
MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"], 1)}

def _key(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", str(s or "").lower())

def _norm_label(s) -> str:
    s = str(s or "").lower()
    s = re.sub(r"\(\d+\)|\d+/|\bof which:?\b", " ", s)   # footnote marks / "of which"
    return re.sub(r"[^a-z0-9]+", " ", s).strip()

def _norm_header(s) -> str:
    """Canonicalise a column header; dates -> 'YYYY-MM'/'YYYY-MM-DD' so formats match."""
    t = str(s or "").strip()
    m = re.search(r"(\d{1,2})\s+([A-Za-z]{3,})\s+(\d{2,4})", t)        # 31 Dec 2025 / 31 Dec 25
    if m:
        d, mon, y = int(m.group(1)), MONTHS.get(m.group(2)[:3].lower()), int(m.group(3))
        if mon:
            y = y + 2000 if y < 100 else y
            return f"{y:04d}-{mon:02d}-{d:02d}"
    m = re.search(r"([A-Za-z]{3,})\s+(\d{4})", t)                       # Dec 2025
    if m and MONTHS.get(m.group(1)[:3].lower()):
        return f"{int(m.group(2)):04d}-{MONTHS[m.group(1)[:3].lower()]:02d}"
    return re.sub(r"[^a-z0-9]+", " ", t.lower()).strip()

_NOVAL = {"", "-", "–", "—", "n.m.", "nm", "na", "n/a", "#", "none"}
def _norm_value(v):
    """-> float for numbers/percentages, None for dashes/blanks/NM, else lowercased str."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if s.lower() in _NOVAL:
        return None
    pct = s.endswith("%")
    t = s.rstrip("%").replace(",", "").strip()
    neg = t.startswith("(") and t.endswith(")")
    core = t[1:-1] if neg else t
    try:
        f = float(core)
        return -f if neg else f
    except ValueError:
        return s.lower()

def _is_header_row(cells) -> bool:
    keys = {_key(c) for c in cells if c not in (None, "")}
    return len(keys & META_KEYS) >= 2

def parse_workbook(path: str) -> dict:
    """sheet name -> {(row_label, col_header): value}.  Merges all tables on a sheet."""
    wb = openpyxl.load_workbook(path, data_only=True)
    out = {}
    for ws in wb.worksheets:
        if ws.title.lower() in ("contents", "table of contents", "toc"):
            continue
        rows = [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
                for r in range(1, ws.max_row + 1)]
        cells, label_col, headers = {}, None, {}
        for row in rows:
            if _is_header_row(row):
                meta_idx = [i for i, c in enumerate(row) if _key(c) in META_KEYS]
                label_col = (max(meta_idx) + 1) if meta_idx else 0
                headers = {i: _norm_header(row[i]) for i in range(label_col + 1, len(row))
                           if row[i] not in (None, "")}
                continue
            if label_col is None or label_col >= len(row):
                continue
            lbl = row[label_col]
            if lbl in (None, ""):
                continue
            rl = _norm_label(lbl)
            for ci, hdr in headers.items():
                if ci < len(row):
                    val = _norm_value(row[ci])
                    if val is not None:
                        cells[(rl, hdr)] = val
        if cells:
            out[ws.title] = cells
    return out

def _title_overlap(a: str, b: str) -> float:
    wa = set(re.findall(r"[a-z0-9]+", a.lower())) - {"a", "b", "c", "the", "of", "and"}
    wb = set(re.findall(r"[a-z0-9]+", b.lower())) - {"a", "b", "c", "the", "of", "and"}
    return len(wa & wb) / (min(len(wa), len(wb)) or 1)

def match_tabs(produced: dict, golden: dict) -> list[tuple]:
    pairs, used = [], set()
    for g in golden:
        best, score = None, 0.0
        for p in produced:
            if p in used:
                continue
            s = _title_overlap(p, g)
            if s > score:
                best, score = p, s
        if best and score >= 0.34:
            pairs.append((g, best, round(score, 2)))
            used.add(best)
        else:
            pairs.append((g, None, 0.0))
    return pairs

def values_equal(a, b, tol: float) -> bool:
    if isinstance(a, float) and isinstance(b, float):
        return abs(a - b) <= tol or (a != 0 and abs(a - b) / abs(a) <= 0.001)
    return a == b

def main():
    ap = argparse.ArgumentParser(description="Values-only Excel accuracy comparator (zero API)")
    ap.add_argument("--produced", required=True)
    ap.add_argument("--golden", required=True)
    ap.add_argument("--out", default="out/score_report.json")
    ap.add_argument("--tol", type=float, default=0.51, help="abs tolerance for numeric match")
    ap.add_argument("--show", type=int, default=12, help="max mismatches to print per tab")
    args = ap.parse_args()
    for f in (args.produced, args.golden):
        if not os.path.exists(f):
            sys.exit(f"not found: {f}")

    produced = parse_workbook(args.produced)
    golden   = parse_workbook(args.golden)
    pairs    = match_tabs(produced, golden)

    report, tot_g = {"tabs": []}, 0
    tot_match = tot_mis = tot_missing = 0
    for g, p, score in pairs:
        gcells = golden[g]
        pcells = produced.get(p, {}) if p else {}
        match = mismatch = missing = 0
        examples = []
        for key, gv in gcells.items():
            if key in pcells:
                if values_equal(gv, pcells[key], args.tol):
                    match += 1
                else:
                    mismatch += 1
                    if len(examples) < args.show:
                        examples.append({"row": key[0], "col": key[1],
                                         "golden": gv, "produced": pcells[key]})
            else:
                missing += 1
                if len(examples) < args.show:
                    examples.append({"row": key[0], "col": key[1], "golden": gv, "produced": "MISSING"})
        extra = len(set(pcells) - set(gcells))
        n = len(gcells)
        tot_g += n; tot_match += match; tot_mis += mismatch; tot_missing += missing
        report["tabs"].append({
            "golden_tab": g, "produced_tab": p, "match_score": score,
            "golden_values": n, "matched": match, "mismatched": mismatch,
            "missing": missing, "extra_in_produced": extra,
            "accuracy": round(match / n, 4) if n else None, "examples": examples,
        })

    report["overall"] = {
        "golden_value_cells": tot_g, "matched": tot_match,
        "mismatched": tot_mis, "missing": tot_missing,
        "accuracy": round(tot_match / tot_g, 4) if tot_g else None,
    }
    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    json.dump(report, open(args.out, "w"), indent=2)

    print(f"\nVALUE ACCURACY: {report['overall']['accuracy']}  "
          f"({tot_match}/{tot_g} golden value cells matched; {tot_mis} wrong, {tot_missing} missing)")
    print(f"{'golden tab':<34} {'produced tab':<30} {'acc':>6} {'match/tot':>10}")
    for t in report["tabs"]:
        print(f"  {t['golden_tab'][:32]:<32} {str(t['produced_tab'])[:28]:<28} "
              f"{str(t['accuracy']):>6} {t['matched']}/{t['golden_values']:<6}")
        for ex in t["examples"][:args.show]:
            print(f"        ⚠ [{ex['row'][:30]} | {ex['col']}]  golden={ex['golden']}  produced={ex['produced']}")
    print(f"\n→ full report: {args.out}")

if __name__ == "__main__":
    main()
