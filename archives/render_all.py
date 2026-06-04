"""
render_all.py — render all extracted tables into one Excel workbook.

One sheet per table. First sheet is a summary index.
Flagged cells (CONFLICT, ILLEGIBLE, suspicious patterns) are highlighted.
Tables with issues are marked in the summary.

Usage:
  python3 render_all.py
  python3 render_all.py --extracted out/step3_extracted --manifest out/step2_table_map.csv --out out/step4_output.xlsx
"""
from __future__ import annotations
import os
import sys
import csv
import json
import re
import argparse
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from extract_tables import _border_cell_spans

PDF_PATH    = None   # set at runtime from CLI arg or auto-detected

# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------
C_NAVY       = "1F3864"   # header background
C_WHITE      = "FFFFFF"
C_BLUE_LIGHT = "BDD7EE"   # group header / leaf header
C_GREY       = "D9D9D9"   # total rows
C_YELLOW     = "FFEB9C"   # needs_review warning in summary
C_RED_LIGHT  = "FFC7CE"   # CONFLICT cell
C_ORANGE     = "FFCC99"   # ILLEGIBLE cell
C_GREEN      = "C6EFCE"   # reconcile=true in summary
C_RED        = "FFC7CE"   # reconcile=false in summary

THIN  = Side(style="thin",   color="CCCCCC")
THICK = Side(style="medium", color="AAAAAA")
BORDER      = Border(left=THIN,  right=THIN,  top=THIN,  bottom=THIN)
BORDER_THICK= Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

INDENT = {0: 0, 1: 0, 2: 2, 3: 4, 4: 6}

# ---------------------------------------------------------------------------
# Pattern detection — flags suspicious cell values
# ---------------------------------------------------------------------------
PLACEHOLDER_RE = re.compile(
    r"^(tbd|xxx+|test|illegible|missing|unknown|placeholder)$", re.I
)
# Note: "NA" and "N/A" are intentionally excluded — they are valid printed values
# in Basel capital instrument disclosure tables (MAS template uses "NA" for Not Applicable).
ALL_ZERO_THRESHOLD = 0.80   # if >80% of a row's numeric cells are 0 → suspicious


def _flag_cell(value: Any) -> str | None:
    """Return a flag string if the value is suspicious, else None."""
    if value is None:
        return None
    s = str(value).strip()
    if s.upper() == "ILLEGIBLE":
        return "ILLEGIBLE"
    if s.upper().startswith("CONFLICT:"):
        return "CONFLICT"
    if PLACEHOLDER_RE.match(s):
        return "PLACEHOLDER"
    return None


def _row_all_zero(cells: dict, col_ids: list[str]) -> bool:
    """True if all numeric cells in the row are zero (suspicious blank-row extraction)."""
    nums = []
    for cid in col_ids:
        v = cells.get(cid)
        if v is None or v == "-" or v == "":
            continue
        try:
            nums.append(float(str(v).replace(",", "")))
        except ValueError:
            pass
    return len(nums) >= 3 and all(n == 0.0 for n in nums)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------
def _hdr_cell(ws, row, col, value, fill_hex, font_color="FFFFFF", bold=True, wrap=False):
    c = ws.cell(row, col, value)
    c.fill   = PatternFill("solid", fgColor=fill_hex)
    c.font   = Font(color=font_color, bold=bold, size=10)
    c.border = BORDER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    return c


def _data_cell(ws, row, col, value, fill_hex=None, bold=False, align="right"):
    c = ws.cell(row, col, value)
    if fill_hex:
        c.fill = PatternFill("solid", fgColor=fill_hex)
    c.font   = Font(bold=bold, size=10)
    c.border = BORDER
    c.alignment = Alignment(horizontal=align, vertical="center")
    return c


# ---------------------------------------------------------------------------
# Per-table sheet renderer
# ---------------------------------------------------------------------------
def render_table_sheet(wb: openpyxl.Workbook, data: dict, meta: dict) -> dict:
    """
    Write one sheet for the table. Returns a summary record dict.
    """
    tid      = data.get("table_id", "unknown")
    rep_date = data.get("reporting_date", "")
    columns  = data.get("columns", [])
    rows     = data.get("rows", [])
    chk      = data.get("self_check", {})

    sheet_name = tid[:31]
    ws = wb.create_sheet(title=sheet_name)

    # ---- Row 1: title -------------------------------------------------------
    title_str = f"{meta.get('section_title','')}"
    if meta.get("table_title") and meta["table_title"] != meta.get("section_title"):
        title_str += f"  |  {meta['table_title']}"
    title_str += f"  |  {rep_date}"
    ws.cell(1, 1, title_str).font = Font(bold=True, size=12)

    # ---- Row 2: self-check summary ------------------------------------------
    reconcile_val = chk.get("totals_reconcile", "?")
    reconcile_str = str(reconcile_val).lower()
    ws.cell(2, 1, (
        f"self_check: {chk.get('n_rows')} rows × {chk.get('n_cols')} cols  |  "
        f"totals_reconcile={reconcile_str}  |  "
        f"layout={meta.get('layout','')}  |  "
        f"notes: {chk.get('notes') or ''}"
    )).font = Font(italic=True, size=9, color="666666")

    # ---- Row 3: blank spacer ------------------------------------------------

    # ---- Rows 4+5: column headers -------------------------------------------
    META_COLS = ["row_id", "level", "parent", "row_type", "footnotes", "label"]
    HDR_ROW   = 4
    data_col_start = len(META_COLS) + 1

    # Meta headers (row 4, spanning into row 5 via merge)
    for ci, h in enumerate(META_COLS, start=1):
        _hdr_cell(ws, HDR_ROW, ci, h, C_NAVY)
        _hdr_cell(ws, HDR_ROW + 1, ci, "", C_NAVY)
        ws.merge_cells(start_row=HDR_ROW, start_column=ci,
                       end_row=HDR_ROW + 1, end_column=ci)

    # Group headers (row 4) + leaf headers (row 5)
    # Uses col_span/row_span from docling when available, otherwise falls back
    # to group-change detection for the merge boundaries.
    prev_group     = None
    group_start_ci = data_col_start

    for ci_offset, col in enumerate(columns):
        ci       = data_col_start + ci_offset
        group    = col.get("group") or ""
        leaf     = col.get("leaf", col["col_id"])
        col_span = int(col.get("col_span") or 1)
        # Cap row_span to the header area (HDR_ROW through HDR_ROW+1) — never bleed into data rows
        row_span = min(int(col.get("row_span") or 1), 2)

        # Leaf header — row 5 (or row 4 if it spans both rows)
        leaf_row = HDR_ROW + 1
        _hdr_cell(ws, leaf_row, ci,
                  f'{col["col_id"]}\n{leaf}',
                  C_BLUE_LIGHT, font_color="000000", bold=True, wrap=True)
        # Merge leaf cell across columns if col_span > 1
        if col_span > 1:
            ws.merge_cells(start_row=leaf_row, start_column=ci,
                           end_row=leaf_row,   end_column=ci + col_span - 1)
        # Merge leaf cell up into group row if row_span > 1 (header spans both header rows)
        if row_span > 1:
            ws.merge_cells(start_row=HDR_ROW, start_column=ci,
                           end_row=HDR_ROW + row_span - 1, end_column=ci + col_span - 1)

        # Group header — row 4
        if group != prev_group:
            if prev_group is not None and ci - 1 >= group_start_ci:
                ws.merge_cells(start_row=HDR_ROW, start_column=group_start_ci,
                               end_row=HDR_ROW,   end_column=ci - 1)
            _hdr_cell(ws, HDR_ROW, ci, group, C_NAVY)
            group_start_ci = ci
            prev_group = group

    # Merge last group span
    last_ci = data_col_start + len(columns) - 1
    if last_ci >= group_start_ci:
        ws.merge_cells(start_row=HDR_ROW, start_column=group_start_ci,
                       end_row=HDR_ROW,   end_column=last_ci)

    # ---- Data rows ----------------------------------------------------------
    DATA_ROW_START = HDR_ROW + 2
    col_ids = [c["col_id"] for c in columns]

    flag_count      = 0
    all_zero_count  = 0
    rowspan_covered: set[tuple[int, int]] = set()  # (sheet_row, col_index) covered by row_span merges

    for ri, row in enumerate(rows):
        dr       = DATA_ROW_START + ri
        level    = row.get("hierarchy_level", 1)
        indent   = " " * INDENT.get(level, 0)
        is_total = str(row.get("row_type", "")).lower() == "total"
        is_hdr   = str(row.get("row_type", "")).lower() in ("section_header", "sub_header")

        row_fill = C_GREY if is_total else (C_BLUE_LIGHT if is_hdr else None)
        row_bold = is_total or is_hdr

        row_span = int(row.get("row_span") or 1)
        meta_vals = [
            row.get("row_id", ""),
            level,
            row.get("parent_row_id", ""),
            row.get("row_type", ""),
            ", ".join(row.get("footnote_marks", [])),
            indent + row.get("label", ""),
        ]
        for ci, val in enumerate(meta_vals, start=1):
            if (dr, ci) in rowspan_covered:
                continue  # cell is interior of a downward row_span merge — do not write
            is_label = ci == len(META_COLS)
            align = "left" if is_label else "center"
            cell = _data_cell(ws, dr, ci, val, fill_hex=row_fill, bold=row_bold, align=align)
            if is_label:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            # Merge label cell downward if row_span > 1 (e.g. a section header spanning sub-rows)
            if row_span > 1:
                try:
                    ws.merge_cells(start_row=dr, start_column=ci,
                                   end_row=dr + row_span - 1, end_column=ci)
                    for skip in range(1, row_span):
                        rowspan_covered.add((dr + skip, ci))
                except Exception:
                    pass  # skip if already merged from a previous overlap

        cells      = row.get("cells", {})
        cell_spans = row.get("cell_spans", {})

        # All-zero row check
        if _row_all_zero(cells, col_ids):
            all_zero_count += 1

        # Build set of col_ids that are covered by a merge (absent from cells after _derive_cell_spans)
        covered: set[str] = set()
        for cid, span in cell_spans.items():
            for skip in range(1, span):
                idx = col_ids.index(cid) + skip
                if idx < len(col_ids):
                    covered.add(col_ids[idx])

        for ci_offset, cid in enumerate(col_ids):
            ci  = data_col_start + ci_offset

            # Skip covered cells — horizontal merge or downward row_span merge
            if cid in covered or (dr, ci) in rowspan_covered:
                continue

            val = cells.get(cid, "")

            flag = _flag_cell(val)
            if flag == "CONFLICT":
                cell_fill = C_RED_LIGHT
                flag_count += 1
            elif flag in ("ILLEGIBLE", "PLACEHOLDER"):
                cell_fill = C_ORANGE
                flag_count += 1
            else:
                cell_fill = row_fill

            _data_cell(ws, dr, ci, val, fill_hex=cell_fill, bold=row_bold)

            # Merge data cell horizontally if cell_spans says so
            span = int(cell_spans.get(cid, 1))
            if span > 1:
                try:
                    ws.merge_cells(start_row=dr, start_column=ci,
                                   end_row=dr,   end_column=ci + span - 1)
                except Exception:
                    pass

    # ---- Column widths ------------------------------------------------------
    ws.column_dimensions[get_column_letter(1)].width = 8    # row_id
    ws.column_dimensions[get_column_letter(2)].width = 6    # level
    ws.column_dimensions[get_column_letter(3)].width = 8    # parent
    ws.column_dimensions[get_column_letter(4)].width = 12   # row_type
    ws.column_dimensions[get_column_letter(5)].width = 10   # footnotes
    ws.column_dimensions[get_column_letter(6)].width = 45   # label
    for ci_offset in range(len(columns)):
        ws.column_dimensions[get_column_letter(data_col_start + ci_offset)].width = 14

    ws.row_dimensions[HDR_ROW + 1].height = 40
    ws.freeze_panes = ws.cell(DATA_ROW_START, data_col_start)

    # ---- Compute confidence score -------------------------------------------
    reconciles = str(reconcile_val).lower() in ("true", "1", "yes", "not_applicable")
    confidence = 1.0
    if not reconciles:
        confidence -= 0.30
    if flag_count > 0:
        confidence -= min(0.40, flag_count * 0.10)
    if all_zero_count > 0:
        confidence -= min(0.20, all_zero_count * 0.05)
    confidence = max(0.0, round(confidence, 2))

    needs_review = confidence < 0.90 or flag_count > 0 or not reconciles

    return {
        "table_id":        tid,
        "section_id":      meta.get("section_id", ""),
        "section_title":   meta.get("section_title", ""),
        "table_title":     meta.get("table_title", ""),
        "reporting_date":  rep_date,
        "layout":          meta.get("layout", ""),
        "pages":           meta.get("pages", ""),
        "n_rows":          chk.get("n_rows", "?"),
        "n_cols":          chk.get("n_cols", "?"),
        "totals_reconcile": reconcile_str,
        "flag_count":      flag_count,
        "all_zero_rows":   all_zero_count,
        "confidence":      confidence,
        "needs_review":    needs_review,
        "notes":           chk.get("notes") or "",
    }


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------
SUMMARY_COLS = [
    ("table_id",        12, "left"),
    ("section_id",       9, "center"),
    ("table_title",     28, "left"),
    ("reporting_date",  14, "center"),
    ("layout",          14, "center"),
    ("pages",            8, "center"),
    ("n_rows",           7, "center"),
    ("n_cols",           7, "center"),
    ("totals_reconcile",14, "center"),
    ("flag_count",       9, "center"),
    ("all_zero_rows",   11, "center"),
    ("confidence",      11, "center"),
    ("needs_review",    12, "center"),
    ("notes",           40, "left"),
]


def render_summary_sheet(wb: openpyxl.Workbook, summary_rows: list[dict]):
    ws = wb.create_sheet(title="SUMMARY", index=0)

    ws.cell(1, 1, "FinDocIQ — Extraction Summary").font = Font(bold=True, size=13)
    ws.cell(2, 1,
        f"{len(summary_rows)} tables extracted  |  "
        f"{sum(1 for r in summary_rows if r['needs_review'])} need review  |  "
        f"avg confidence: {sum(r['confidence'] for r in summary_rows)/max(len(summary_rows),1):.2f}"
    ).font = Font(italic=True, size=10, color="555555")

    HDR = 4
    for ci, (col_key, width, _) in enumerate(SUMMARY_COLS, start=1):
        _hdr_cell(ws, HDR, ci, col_key, C_NAVY)
        ws.column_dimensions[get_column_letter(ci)].width = width

    for ri, rec in enumerate(summary_rows):
        dr = HDR + 1 + ri
        for ci, (col_key, _, align) in enumerate(SUMMARY_COLS, start=1):
            val = rec.get(col_key, "")

            # Row background
            if rec["needs_review"]:
                row_fill = C_YELLOW
            elif rec["totals_reconcile"] == "false":
                row_fill = C_RED_LIGHT
            else:
                row_fill = None

            c = ws.cell(dr, ci, val)
            if row_fill:
                c.fill = PatternFill("solid", fgColor=row_fill)
            c.border = BORDER
            c.alignment = Alignment(horizontal=align, vertical="center")
            c.font = Font(size=10)

            # Hyperlink to the table sheet
            if col_key == "table_id":
                sheet_ref = f"'{val[:31]}'!A1"
                c.hyperlink = f"#{sheet_ref}"
                c.font = Font(size=10, color="0563C1", underline="single")

            # Colour confidence cell
            if col_key == "confidence":
                if isinstance(val, float):
                    if val >= 0.95:
                        c.fill = PatternFill("solid", fgColor=C_GREEN)
                    elif val >= 0.80:
                        c.fill = PatternFill("solid", fgColor=C_YELLOW)
                    else:
                        c.fill = PatternFill("solid", fgColor=C_RED_LIGHT)

            # Colour reconcile cell
            if col_key == "totals_reconcile":
                if str(val).lower() == "true":
                    c.fill = PatternFill("solid", fgColor=C_GREEN)
                elif str(val).lower() == "false":
                    c.fill = PatternFill("solid", fgColor=C_RED_LIGHT)

    ws.freeze_panes = ws.cell(HDR + 1, 1)
    ws.row_dimensions[HDR].height = 20


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def render_all(extracted_dir: str, manifest_path: str, out_path: str, pdf_path: str | None = None):
    # Load manifest for metadata
    meta_by_id: dict[str, dict] = {}
    if os.path.exists(manifest_path):
        with open(manifest_path) as f:
            for row in csv.DictReader(f):
                meta_by_id[row["table_id"]] = row

    # Collect and sort JSON files
    json_files = sorted(
        f for f in os.listdir(extracted_dir) if f.endswith(".json")
    )
    if not json_files:
        sys.exit(f"No JSON files found in {extracted_dir}")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default empty sheet

    summary_rows: list[dict] = []
    total = len(json_files)

    for i, fname in enumerate(json_files, start=1):
        tid = fname.replace(".json", "")
        path = os.path.join(extracted_dir, fname)
        print(f"  [{i:>3}/{total}] {tid}", end="")

        try:
            with open(path) as f:
                data = json.load(f)

            # Inject cell_spans + corrected cell values from PDF vertical borders.
            if pdf_path and os.path.exists(pdf_path):
                meta_row  = meta_by_id.get(tid, {})
                page_nums = [int(p) for p in meta_row.get("pages", "").split("+") if p]
                if page_nums:
                    border_info = _border_cell_spans(pdf_path, page_nums, data)
                    for row in data.get("rows", []):
                        rid = str(row.get("row_id", ""))
                        if rid in border_info:
                            info = border_info[rid]
                            row["cell_spans"] = info["spans"]
                            # Merge corrected anchor values back; leave uncorrected cols alone
                            for cid, val in info["cells"].items():
                                row["cells"][cid] = val
                            # Remove covered cols from cells so renderer skips them
                            col_ids_all = [c["col_id"] for c in data.get("columns", [])]
                            for cid, span in info["spans"].items():
                                start = col_ids_all.index(cid)
                                for skip in range(1, span):
                                    covered = col_ids_all[start + skip]
                                    row["cells"].pop(covered, None)

            meta = meta_by_id.get(tid, {})
            rec  = render_table_sheet(wb, data, meta)
            summary_rows.append(rec)
            flag_str = f"  ⚠️  {rec['flag_count']} flags" if rec["flag_count"] else ""
            review_str = "  👁 needs review" if rec["needs_review"] else ""
            print(f"  conf={rec['confidence']:.2f}{flag_str}{review_str}")
        except Exception as e:
            print(f"  ❌ ERROR: {e}")
            summary_rows.append({
                "table_id": tid, "section_id": "", "section_title": "",
                "table_title": "", "reporting_date": "", "layout": "",
                "pages": "", "n_rows": "?", "n_cols": "?",
                "totals_reconcile": "error", "flag_count": 0,
                "all_zero_rows": 0, "confidence": 0.0,
                "needs_review": True, "notes": str(e),
            })

    render_summary_sheet(wb, summary_rows)

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)

    needs_review = sum(1 for r in summary_rows if r["needs_review"])
    avg_conf     = sum(r["confidence"] for r in summary_rows) / max(len(summary_rows), 1)
    print(f"\n✅ Saved → {out_path}")
    print(f"   {len(summary_rows)} tables  |  {needs_review} need review  |  avg confidence {avg_conf:.2f}")


if __name__ == "__main__":
    p = argparse.ArgumentParser(description="Render all extracted tables into one Excel workbook.")
    p.add_argument("--extracted", default="out/step3_extracted",    help="Directory of extracted JSON files")
    p.add_argument("--manifest",  default="out/step2_table_map.csv", help="Table manifest CSV")
    p.add_argument("--out",       default="out/step4_output.xlsx",   help="Output Excel path")
    p.add_argument("--pdf",       default=None,                      help="Source PDF for border-based merge detection")
    p.add_argument("--table",     default=None,                      help="Render only this table ID (e.g. t085_p84_85)")
    args = p.parse_args()

    # Auto-detect PDF if not given: prefer DBS_4Q25 pattern, else first *.pdf
    pdf = args.pdf
    if not pdf:
        import glob
        pdfs = sorted(glob.glob("*.pdf"))
        # Prefer the most recently modified PDF
        if pdfs:
            pdf = max(pdfs, key=lambda p: os.path.getmtime(p))
            print(f"  📄 Auto-detected PDF: {pdf}")

    extracted_dir = args.extracted
    # If --table given, create a temp dir with only that file
    if args.table:
        import tempfile, shutil
        src = os.path.join(extracted_dir, f"{args.table}.json")
        if not os.path.exists(src):
            print(f"ERROR: {src} not found")
            sys.exit(1)
        tmp = tempfile.mkdtemp()
        shutil.copy(src, tmp)
        extracted_dir = tmp

    render_all(extracted_dir, args.manifest, args.out, pdf_path=pdf)
