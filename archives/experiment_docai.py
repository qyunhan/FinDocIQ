"""
experiment_docai.py — Google Document AI Layout Parser experiment.

Tests two OCBC sections against our current Gemini extraction:
  - Section 24: NSFR (pages 96-98) — tests merged cell detection
  - Section 18.4: F-IRBA CCR (pages 75-78) — tests multi-table spanning pages

Output: out/docai_experiment/
  - <section>_docai.xlsx      — Document AI extracted tables
  - <section>_docai_raw.json  — raw Document AI response (for debugging)
  - <section>_comparison.txt  — side-by-side diff vs current Gemini output

Usage:
  python experiment_docai.py
"""
from __future__ import annotations
import os, io, json, sys
from google.cloud import documentai
import pypdfium2 as pdfium
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
os.environ.setdefault("GOOGLE_CLOUD_PROJECT", "data-extraction-497009")

PROJECT_ID   = "data-extraction-497009"
LOCATION     = "us"
PROCESSOR_ID = "ac3dce4152b9b719"
PROCESSOR_NAME = f"projects/{PROJECT_ID}/locations/{LOCATION}/processors/{PROCESSOR_ID}"

PDF_PATH  = "OCBC_4Q25_Pillar 3.pdf"
OUT_DIR   = "out/docai_experiment"

SECTIONS = {
    "24_NSFR":   {"pages": list(range(96, 99)),  "title": "Net Stable Funding Ratio"},
    "18_4_FIRBA": {"pages": list(range(75, 79)),  "title": "F-IRBA - CCR Exposures by Portfolio and PD Range"},
}

BRAND   = "CC0000"   # OCBC red
HDR     = "1F3864"   # header navy
WHITE   = "FFFFFF"

# ---------------------------------------------------------------------------
def cut_pdf(pdf_path: str, pages_1based: list[int]) -> bytes:
    src  = pdfium.PdfDocument(pdf_path)
    dest = pdfium.PdfDocument.new()
    dest.import_pages(src, [p - 1 for p in pages_1based])
    buf = io.BytesIO()
    dest.save(buf)
    return buf.getvalue()

def process_with_docai(pdf_bytes: bytes) -> dict:
    """Call Document AI and return the full response as a plain dict (documentLayout format)."""
    import google.auth, google.auth.transport.requests
    from google.protobuf import json_format

    creds, _ = google.auth.default(scopes=["https://www.googleapis.com/auth/cloud-platform"])
    creds.refresh(google.auth.transport.requests.Request())

    client = documentai.DocumentProcessorServiceClient(
        credentials=creds,
        client_options={"api_endpoint": f"{LOCATION}-documentai.googleapis.com"}
    )
    raw_doc = documentai.RawDocument(content=pdf_bytes, mime_type="application/pdf")
    request = documentai.ProcessRequest(name=PROCESSOR_NAME, raw_document=raw_doc)
    result  = client.process_document(request=request)
    return json_format.MessageToDict(result._pb)

# ---------------------------------------------------------------------------
def _block_text(block: dict) -> str:
    """Recursively collect all text from a block and its children."""
    parts = []
    tb = block.get("textBlock", {})
    if tb.get("text"):
        parts.append(tb["text"])
    for child in tb.get("blocks", []):
        parts.append(_block_text(child))
    for child in block.get("blocks", []):
        parts.append(_block_text(child))
    return " ".join(p for p in parts if p).strip()

def _parse_cell(cell_dict: dict) -> dict:
    text_parts = []
    for blk in cell_dict.get("blocks", []):
        t = _block_text(blk)
        if t:
            text_parts.append(t)
    return {
        "text": " ".join(text_parts).strip(),
        "row_span": cell_dict.get("rowSpan", 1),
        "col_span": cell_dict.get("colSpan", 1),
    }

def _collect_tables(blocks: list, page_start: int = 1) -> list[dict]:
    """Walk the nested block tree and yield every tableBlock as a flat dict."""
    tables = []
    for blk in blocks:
        tb = blk.get("textBlock", {})
        if "tableBlock" in blk:
            tbl = blk["tableBlock"]
            page = blk.get("pageSpan", {}).get("pageStart", page_start)
            rows = []
            for hr in tbl.get("headerRows", []):
                cells = [dict(**_parse_cell(c), is_header=True) for c in hr.get("cells", [])]
                rows.append({"cells": cells, "is_header": True})
            for br in tbl.get("bodyRows", []):
                cells = [dict(**_parse_cell(c), is_header=False) for c in br.get("cells", [])]
                rows.append({"cells": cells, "is_header": False})
            tables.append({
                "page": page,
                "title": "",
                "rows": rows,
                "merged_cells_detected": any(
                    c["col_span"] > 1 or c["row_span"] > 1
                    for r in rows for c in r["cells"]
                ),
            })
        # recurse into textBlock children
        if tb.get("blocks"):
            # capture the nearest heading as title for the next tableBlock
            heading_text = tb.get("text", "")
            child_tables = _collect_tables(tb["blocks"], page_start)
            for ct in child_tables:
                if not ct["title"] and heading_text:
                    ct["title"] = heading_text
            tables.extend(child_tables)
        if blk.get("blocks"):
            tables.extend(_collect_tables(blk["blocks"], page_start))
    return tables

def extract_tables(doc_dict: dict) -> list[dict]:
    """Extract tables from the documentLayout.blocks structure returned by Layout Parser."""
    blocks = doc_dict.get("document", {}).get("documentLayout", {}).get("blocks", [])
    return _collect_tables(blocks)

# ---------------------------------------------------------------------------
def write_excel(tables: list[dict], section_id: str, title: str) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for ti, tbl in enumerate(tables):
        sheet_name = f"{section_id}_t{ti+1}"[:31]
        ws = wb.create_sheet(title=sheet_name)

        # title row
        ws.cell(1, 1, f"{title} — Table {ti+1} (p{tbl['page']})").font = Font(bold=True, size=11)
        if tbl["merged_cells_detected"]:
            ws.cell(2, 1, "⚠ Merged cells detected and expanded").font = Font(italic=True, color="CC0000")
        start_row = 3

        # expand merged cells into a flat grid first
        grid = _expand_to_grid(tbl["rows"])

        for ri, grid_row in enumerate(grid):
            for ci, cell_val in enumerate(grid_row):
                is_hdr = ri < _header_row_count(tbl["rows"])
                c = ws.cell(start_row + ri, ci + 1, cell_val)
                if is_hdr:
                    c.font    = Font(bold=True, color=WHITE)
                    c.fill    = PatternFill("solid", fgColor=HDR)
                    c.alignment = Alignment(horizontal="center", wrap_text=True)
                else:
                    c.alignment = Alignment(wrap_text=True)

        # column widths
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=8)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 30)

    path = os.path.join(OUT_DIR, f"{section_id}_docai.xlsx")
    wb.save(path)
    return path

def _header_row_count(rows: list[dict]) -> int:
    return sum(1 for r in rows if r["is_header"])

def _expand_to_grid(rows: list[dict]) -> list[list[str]]:
    """Expand col_span/row_span into a simple 2D grid of strings.
    Merged cells: the value goes in the top-left cell, '' fills the span."""
    if not rows:
        return []

    # figure out total columns
    max_cols = 0
    for r in rows:
        total = sum(c["col_span"] for c in r["cells"])
        max_cols = max(max_cols, total)

    n_rows = len(rows)
    grid = [[""] * max_cols for _ in range(n_rows)]
    occupied = [[False] * max_cols for _ in range(n_rows)]

    for ri, row in enumerate(rows):
        ci_grid = 0
        for cell in row["cells"]:
            # skip already-occupied cells (from a row_span above)
            while ci_grid < max_cols and occupied[ri][ci_grid]:
                ci_grid += 1
            if ci_grid >= max_cols:
                break

            text = cell["text"]
            rs   = cell["row_span"]
            cs   = cell["col_span"]

            # mark the span as occupied and fill value in top-left
            for dr in range(rs):
                for dc in range(cs):
                    rr, cc = ri + dr, ci_grid + dc
                    if rr < n_rows and cc < max_cols:
                        occupied[rr][cc] = True
                        grid[rr][cc] = text if (dr == 0 and dc == 0) else f"↑ (merged)"

            ci_grid += cs

    return grid

# ---------------------------------------------------------------------------
def save_raw(doc_dict: dict, section_id: str):
    """Save the raw Document AI response dict as JSON for inspection."""
    path = os.path.join(OUT_DIR, f"{section_id}_docai_raw.json")
    with open(path, "w") as f:
        json.dump(doc_dict, f, indent=2)
    return path

def print_summary(tables: list[dict], section_id: str):
    print(f"\n{'='*60}")
    print(f"  {section_id}")
    print(f"{'='*60}")
    print(f"  Tables found: {len(tables)}")
    for i, t in enumerate(tables):
        n_hdr  = sum(1 for r in t["rows"] if r["is_header"])
        n_body = sum(1 for r in t["rows"] if not r["is_header"])
        merged = t["merged_cells_detected"]
        print(f"  Table {i+1} (page {t['page']}): {n_hdr} header rows, {n_body} body rows  "
              f"{'⚠ MERGED CELLS' if merged else 'no merges'}")
        # show header cells
        for hr in t["rows"]:
            if not hr["is_header"]:
                break
            cells_str = " | ".join(
                f"{c['text'][:20]}{'[cs'+str(c['col_span'])+']' if c['col_span']>1 else ''}"
                for c in hr["cells"]
            )
            print(f"    HDR: {cells_str}")

# ---------------------------------------------------------------------------
def main():
    os.makedirs(OUT_DIR, exist_ok=True)

    for section_id, info in SECTIONS.items():
        pages = info["pages"]
        title = info["title"]
        print(f"\n→ Processing {section_id}: pages {pages[0]}–{pages[-1]}")

        # cut PDF
        pdf_bytes = cut_pdf(PDF_PATH, pages)
        print(f"  PDF slice: {len(pdf_bytes):,} bytes")

        # send to Document AI
        print(f"  Sending to Document AI Layout Parser...")
        doc_dict = process_with_docai(pdf_bytes)
        doc_node = doc_dict.get("document", {})
        n_blocks = len(doc_node.get("documentLayout", {}).get("blocks", []))
        print(f"  ✓ Response received  ({n_blocks} top-level layout blocks)")

        # extract tables
        tables = extract_tables(doc_dict)

        # print summary
        print_summary(tables, section_id)

        # save raw JSON
        raw_path = save_raw(doc_dict, section_id)
        print(f"  Raw JSON → {raw_path}")

        # write Excel
        if tables:
            xlsx_path = write_excel(tables, section_id, title)
            print(f"  Excel   → {xlsx_path}")
        else:
            print(f"  ⚠ No tables extracted — check raw JSON for what Document AI returned")

    print(f"\n✓ Done. Outputs in {OUT_DIR}/")

if __name__ == "__main__":
    main()
