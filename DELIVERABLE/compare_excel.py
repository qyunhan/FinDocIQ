"""
compare_excel.py — cell-by-cell diff of two Excel workbooks.

Usage:
  python compare_excel.py outputs/ocbc.xlsx outputs/ocbc_baseline.xlsx
  python compare_excel.py outputs/ocbc.xlsx outputs/ocbc_baseline.xlsx --sheets-only
  python compare_excel.py outputs/ocbc.xlsx outputs/ocbc_baseline.xlsx --ignore-cost

Exits 0 if identical, 1 if differences found.
"""
from __future__ import annotations
import argparse, sys
import openpyxl
from openpyxl.utils import get_column_letter


def cell_val(cell) -> str:
    v = cell.value
    if v is None:
        return ""
    return str(v).strip()


def compare(path_a: str, path_b: str,
            sheets_only: bool = False,
            ignore_sheets: set[str] | None = None) -> bool:
    """Returns True if identical (within the compared scope)."""
    wa = openpyxl.load_workbook(path_a, data_only=True)
    wb = openpyxl.load_workbook(path_b, data_only=True)

    ignore_sheets = ignore_sheets or set()
    sheets_a = [s for s in wa.sheetnames if s not in ignore_sheets]
    sheets_b = [s for s in wb.sheetnames if s not in ignore_sheets]

    ok = True

    # --- sheet-level diff ---
    only_a = [s for s in sheets_a if s not in sheets_b]
    only_b = [s for s in sheets_b if s not in sheets_a]
    if only_a:
        print(f"  ❌ sheets only in A: {only_a}")
        ok = False
    if only_b:
        print(f"  ❌ sheets only in B: {only_b}")
        ok = False

    order_a = [s for s in sheets_a if s in sheets_b]
    order_b = [s for s in sheets_b if s in sheets_a]
    if order_a != order_b:
        print(f"  ⚠  sheet order differs")
        print(f"     A: {order_a}")
        print(f"     B: {order_b}")
        ok = False

    if sheets_only:
        return ok

    # --- cell-level diff (shared sheets only) ---
    common = [s for s in sheets_a if s in sheets_b]
    cell_diffs = 0
    MAX_DIFFS = 50   # cap output so it doesn't flood

    for sname in common:
        wsa = wa[sname]
        wsb = wb[sname]
        max_row = max(wsa.max_row or 0, wsb.max_row or 0)
        max_col = max(wsa.max_column or 0, wsb.max_column or 0)

        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                va = cell_val(wsa.cell(r, c))
                vb = cell_val(wsb.cell(r, c))
                if va != vb:
                    if cell_diffs < MAX_DIFFS:
                        col_ltr = get_column_letter(c)
                        print(f"  ❌ [{sname}] {col_ltr}{r}:  A={repr(va)}  B={repr(vb)}")
                    cell_diffs += 1
                    ok = False

    if cell_diffs > MAX_DIFFS:
        print(f"  ... and {cell_diffs - MAX_DIFFS} more cell difference(s) (truncated)")

    return ok


def main():
    ap = argparse.ArgumentParser(description="Cell-by-cell Excel diff")
    ap.add_argument("file_a")
    ap.add_argument("file_b")
    ap.add_argument("--sheets-only", action="store_true",
                    help="only compare sheet names and order, not cell values")
    ap.add_argument("--ignore-cost", action="store_true",
                    help="skip the 'Cost' sheet (changes every run due to timestamps)")
    args = ap.parse_args()

    ignore = {"Cost"} if args.ignore_cost else set()

    print(f"Comparing:")
    print(f"  A: {args.file_a}")
    print(f"  B: {args.file_b}")
    if ignore:
        print(f"  Ignoring sheets: {sorted(ignore)}")
    print()

    wa = openpyxl.load_workbook(args.file_a, data_only=True)
    wb_  = openpyxl.load_workbook(args.file_b, data_only=True)
    print(f"  A: {len(wa.sheetnames)} sheets")
    print(f"  B: {len(wb_.sheetnames)} sheets")
    print()

    identical = compare(args.file_a, args.file_b,
                        sheets_only=args.sheets_only,
                        ignore_sheets=ignore)

    if identical:
        print("✅  Identical")
        sys.exit(0)
    else:
        print("\n❌  Differences found")
        sys.exit(1)


if __name__ == "__main__":
    main()
