"""
SLIDE_Extract.py — CFO presentation slide deck → Excel workbook.
Hybrid routing architecture:
  Pass 0: classify_slide  → element_types.json (cheap micro-call)
  Route:
    Visual slides (any chart type) → single-pass: image + JSON schema in one call
    Text slides (text_table/npa)   → multi-pass: Pass 1 describe → Pass 2 transcribe
  Validate + correct once if waterfall/label errors found
  Pass 3: render_to_excel → one worksheet per slide

Usage:
  export GEMINI_API_KEY=...
  python3 SLIDE_Extract.py DBS4Q25_CFO_presentation.pdf
  python3 SLIDE_Extract.py DBS4Q25_CFO_presentation.pdf --slide 5
  python3 SLIDE_Extract.py DBS4Q25_CFO_presentation.pdf --start-slide 3
  python3 SLIDE_Extract.py DBS4Q25_CFO_presentation.pdf --dry-run
  python3 SLIDE_Extract.py DBS4Q25_CFO_presentation.pdf --force
"""
from __future__ import annotations
import os, sys, json, io, re, argparse, time, datetime
from pathlib import Path
from typing import Any
import pypdfium2 as pdfium
try:
    import pdfplumber
    _HAS_PDFPLUMBER = True
except ImportError:
    _HAS_PDFPLUMBER = False
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel, field_validator
from google import genai
from google.genai import types

# ---------------------------------------------------------------------------
MODEL      = "gemini-2.5-flash"
IMAGE_SCALE = 3.0

BANKS = {
    "DBS":  {"institution": "DBS Group Holdings Ltd",
             "brand": "CC0000", "match": r"\bDBS\b"},
    "OCBC": {"institution": "Oversea-Chinese Banking Corporation Limited",
             "brand": "CC0000", "match": r"OCBC|Oversea[- ]?Chinese"},
    "UOB":  {"institution": "United Overseas Bank Limited",
             "brand": "1B6EC2", "match": r"\bUOB\b|United Overseas"},
}

# Gemini 2.5 Flash pricing
INPUT_PRICE_PER_M  = 0.30
OUTPUT_PRICE_PER_M = 2.50

NAVY      = "1F3864"
DARK_GREY = "404040"
MID_GREY  = "595959"
WHITE     = "FFFFFF"
YELLOW    = "FFFF00"
TOTAL_BG  = "F2F2F2"
NUM_FMT   = '#,##0;(#,##0);"-"'

BRAND_COLOURS = {
    "DBS":  "CC0000",
    "OCBC": "CC0000",
    "UOB":  "1B6EC2",
}

# Gemini sometimes uses synonyms; normalise before DataPoint construction.
ROW_TYPE_ALIASES: dict[str, str] = {
    "header":         "total",
    "section_header": "total",
    "sub_header":     "sub",
    "subtotal":       "total",
    "opening":        "start",
    "closing":        "end",
    "component":      "bridge",
    "commentary":     "note",
    "footnote":       "note",
}

_run_usage: dict = {"calls": 0, "prompt": 0, "output": 0, "cost": 0.0}

CONTRACTS_PATH = Path(__file__).with_name("chart_contracts.json")

# Chart types that benefit from single-pass (image stays present during schema fill)
VISUAL_TYPES = {
    "waterfall", "stacked_bar", "stacked_bar_with_overlay",
    "trend_line", "kpi_grid", "pie", "donut_dual_ring",
}

# ===========================================================================
# PASS 0 — CLASSIFY
# ===========================================================================
KNOWN_TYPES = {
    "text_table", "waterfall", "stacked_bar", "stacked_bar_with_overlay",
    "trend_line", "kpi_grid", "pie", "donut_dual_ring",
    "npa_movement_table", "none"
}

CLASSIFY_PROMPT = """Look at this slide and list every distinct visual data element type present.

Return ONLY a JSON array of strings using these type names:
  "text_table"               - a printed table with rows and columns
  "waterfall"                - a bridge/waterfall chart showing running total deltas
  "stacked_bar"              - bars made of stacked coloured segments, no overlay line
  "stacked_bar_with_overlay" - stacked bars PLUS a trend line (%, bps) on the same axis
  "trend_line"               - line chart showing values over time, no bars
  "kpi_grid"                 - individual KPI metric boxes or callout figures
  "pie"                      - pie or donut chart (single ring)
  "donut_dual_ring"          - two concentric donut rings representing two time periods
  "npa_movement_table"       - NPA roll-forward table (opening + flows = closing)
  "none"                     - no data elements (title / agenda / closing slide)

If you see something that does not fit any of the above, invent a short
snake_case name for it and include it in the array.

Examples:
  ["text_table", "waterfall"]
  ["stacked_bar_with_overlay", "kpi_grid"]
  ["none"]
  ["text_table", "bullet_bridge"]   <- invented type
"""


def classify_slide(client, img_bytes: bytes) -> list[str]:
    """Pass 0: micro-call to identify element types. Returns list of type strings."""
    config = types.GenerateContentConfig(temperature=0.0)
    # Plain text response — no JSON mime type
    try:
        config.thinking_config = types.ThinkingConfig(thinking_budget=0)
    except Exception:
        pass

    resp = None
    for attempt in range(3):
        try:
            resp = client.models.generate_content(
                model=MODEL,
                contents=[img_part(img_bytes), CLASSIFY_PROMPT],
                config=config,
            )
            break
        except Exception as e:
            if any(c in str(e) for c in ("503", "429", "UNAVAILABLE", "RESOURCE_EXHAUSTED")):
                wait = 15 * (2 ** attempt)
                print(f"      ⏳ {e.__class__.__name__} — waiting {wait}s")
                time.sleep(wait)
            else:
                raise

    if resp is None:
        return []

    raw = (resp.text or "").strip()
    raw = strip_fences(raw)

    # Track usage (Pass 0 is cheap but we still count it)
    um  = getattr(resp, "usage_metadata", None)
    pt  = getattr(um, "prompt_token_count", None) or 0
    ot  = getattr(um, "candidates_token_count", None) or 0
    cost = (pt / 1e6 * INPUT_PRICE_PER_M) + (ot / 1e6 * OUTPUT_PRICE_PER_M)
    _run_usage["calls"]  += 1
    _run_usage["prompt"] += pt
    _run_usage["output"] += ot
    _run_usage["cost"]   += cost

    try:
        types_found = json.loads(raw)
        if isinstance(types_found, list):
            return [str(t).strip() for t in types_found]
    except Exception:
        pass

    return []


def split_known_unknown(types_found: list[str]) -> tuple[list[str], list[str]]:
    known   = [t for t in types_found if t in KNOWN_TYPES]
    unknown = [t for t in types_found if t not in KNOWN_TYPES and t != "none"]
    return known, unknown


# ===========================================================================
# CONTRACT HELPERS (Pass 1 injection)
# ===========================================================================
def load_contracts() -> dict:
    if not CONTRACTS_PATH.exists():
        return {}
    with open(CONTRACTS_PATH) as f:
        raw = json.load(f)
    return {k: v["contract"] for k, v in raw.items() if v.get("status") == "approved"}


def build_contracts_block(known_types: list[str], contracts: dict) -> str:
    if not known_types:
        return ""
    blocks = [
        f"--- CONTRACT: {t.upper()} ---\n{contracts[t]}"
        for t in known_types if t in contracts
    ]
    if not blocks:
        return ""
    return (
        "\n\nCHART READING CONTRACTS FOR THIS SLIDE:\n"
        "Apply these when reasoning about each chart element.\n\n"
        + "\n\n".join(blocks)
    )


UNKNOWN_TYPE_TEMPLATE = """
--- CONTRACT: {unknown_type} (UNKNOWN — DERIVE YOUR OWN) ---
You identified "{unknown_type}" which has no predefined reading contract.

Before extracting, derive a contract for it:

1. DATA MODEL: what underlying data structure generated this visual?
   (what would the raw spreadsheet data look like?)

2. ARITHMETIC CONSTRAINT: is there a summation relationship between values?
   If none, state "no constraint".

3. HOW TO READ: what visual signals encode the values?
   (position, height, colour, label, size, angle)

4. SIGN RULE: can values be negative?
   If so, what visual signal indicates sign?

Write this contract explicitly in your description under the heading:
  DERIVED CONTRACT: {unknown_type}

This derived contract will be saved for future use and human review.
"""


def build_unknown_contracts_block(unknown_types: list[str]) -> str:
    if not unknown_types:
        return ""
    return "\n\n".join(
        UNKNOWN_TYPE_TEMPLATE.format(unknown_type=t) for t in unknown_types
    )


def save_derived_contract(description: str, unknown_types: list[str]) -> None:
    """Parse derived contracts from Pass 1 output and save as pending_review."""
    if not unknown_types:
        return

    existing: dict = {}
    if CONTRACTS_PATH.exists():
        with open(CONTRACTS_PATH) as f:
            existing = json.load(f)

    changed = False
    for t in unknown_types:
        pattern = rf"DERIVED CONTRACT:\s*{re.escape(t)}\s*\n(.*?)(?=\n---|\Z)"
        match = re.search(pattern, description, re.DOTALL | re.IGNORECASE)
        if match and t not in existing:
            existing[t] = {
                "status":   "pending_review",
                "contract": match.group(1).strip(),
            }
            print(f"  📝 New contract derived for '{t}' — saved as pending_review")
            changed = True

    if changed:
        with open(CONTRACTS_PATH, "w") as f:
            json.dump(existing, f, indent=2)


# ===========================================================================
# DATA MODEL
# ===========================================================================
class DataPoint(BaseModel):
    # Identity
    slide:         int
    element_idx:   int
    element_type:  str = "other"
    element_title: str = ""

    # Value
    series:    str = ""
    period:    str | None = None
    value:     str = ""
    value_num: float | None = None
    unit:      str = ""

    # Semantic metadata
    row_type: str = "data"
    level:    int = 1
    parent:   str | None = None
    group:    str | None = None
    sign:     str | None = None   # "+" or "-" for waterfall bridges
    order:    int = 0

    # Dynamic extra columns (qoq_pct, yoy_pct, etc.)
    extra_fields: dict[str, Any] = {}

    # Provenance
    source:     str   # "table" | "chart"
    bank:       str
    doc_title:  str
    doc_date:   str
    slide_title: str

    @field_validator("value_num", mode="before")
    @classmethod
    def parse_numeric(cls, v, info):
        if v is not None:
            return v
        raw = info.data.get("value", "")
        if not raw:
            return None
        s = str(raw).strip()
        neg = s.startswith("(") and s.endswith(")")
        try:
            core = s[1:-1] if neg else s
            num = float(core.replace(",", "").replace("+", "").rstrip("%"))
            return -num if neg else num
        except ValueError:
            return None


# ===========================================================================
# SINGLE-PASS PROMPT (visual slides — image present throughout)
# ===========================================================================
SINGLE_PASS_PROMPT = """Extract ALL financial data from this bank CFO presentation slide.

Return a JSON object with this exact structure:
{
  "slide_title": "the main heading of this slide verbatim",
  "elements": [
    {
      "element_idx": 0,
      "element_type": "text_table|waterfall|stacked_bar|stacked_bar_with_overlay|trend_line|kpi_grid|pie|donut_dual_ring|other",
      "element_title": "the title printed above this element verbatim",
      "source": "chart",
      "units": "e.g. S$m, S$b, %, bps",
      "self_check": "write the arithmetic check for this element e.g. '9755+658-1236=9177', or null",
      "data_points": [
        {
          "series":   "row / bar / segment label verbatim",
          "period":   "time period as printed on slide e.g. FY25, 4Q25, Dec-25, or null",
          "value":    "value VERBATIM as printed — keep commas, %, (), + signs",
          "row_type": "data|total|sub|start|end|bridge|note",
          "level":    1,
          "parent":   null,
          "group":    null,
          "sign":     null,
          "order":    0
        }
      ]
    }
  ]
}

EXTRACTION RULES:

value field — always verbatim:
  "5,948" not 5948.  "(244)" not -244.  "1.82%" not 0.0182.  "+14%" not 14.

row_type:
  start / end  → waterfall opening and closing bars
  bridge       → waterfall delta bars (must have sign="+" or "-")
  total        → bold rows, grand totals, subtotals
  sub          → indented sub-items (level=2, parent="nearest level-1 label")
  note         → footnotes, disclaimers
  data         → everything else

Waterfall charts:
  Every bridge bar needs sign="+" or "-". Read the colour legend on the slide.
  Verify: start_value + sum(signed_deltas) = end_value. Write this as self_check.

Stacked bars with overlay line:
  Bar segments → data_points with period = x-axis label.
  Overlay line → separate data_points, row_type="note", extra field overlay_series="<line name>".

Donut dual ring:
  Trace each period label's callout line to identify which ring it points to.
  Assign period from what you read on the slide. Do NOT assume inner=earlier or outer=later.
  One data_point per (segment, period) combination.

KPI grid:
  One data_point per KPI box. period=null if no time axis.
  Extra fields for change values: e.g. yoy_change="-2%", yoy_change_label="YoY".

Extra fields:
  Add freely for any additional data columns on the slide.
  Always pair with a _label field: e.g. yoy_pct="-3%", yoy_pct_label="YoY %".

order: sequential integer, 0-based, top-to-bottom / left-to-right reading order.

Return ONLY the JSON object. No markdown fences. No explanation.
"""


# ===========================================================================
# PASS 1 PROMPT (text slides only — multi-pass path)
# ===========================================================================
PASS1_PROMPT = """Examine this bank CFO presentation slide. It contains text tables.

═══════════════════════════════════════════════════
STEP 1 — INVENTORY
═══════════════════════════════════════════════════

List every printed number on the slide with:
  - the number exactly as printed (e.g. "2,296", "(1,236)", "1.91%")
  - its row label and column header
  - a "?" flag if it is partially obscured or ambiguous

Do not skip any number.

═══════════════════════════════════════════════════
STEP 2 — STRUCTURE
═══════════════════════════════════════════════════

For each table:
  1. TITLE: the label printed above it (verbatim)
  2. COLUMN HEADERS: left to right, exactly as printed
  3. ROW COUNT: how many data rows (excluding headers and footnotes)
  4. VISUAL CONVENTIONS on this specific slide:
       bold rows          → totals or subtotals?
       indented rows      → sub-items of a parent row?
       shaded/grey cells  → not applicable (blank, not zero)?
       parentheses        → negative values?
       dash cells         → zero or negligible?
       any footnotes that change interpretation of specific rows/cells?

═══════════════════════════════════════════════════
STEP 3 — VERIFY
═══════════════════════════════════════════════════

For each table, check whether sub-items sum to their parent total.
Write the check explicitly: e.g. "Fee income: 312 + 187 + 95 = 594 ✓"
Note any discrepancies without correcting the printed values.

If a table has NPA movement structure (opening + flows = closing),
check: opening + sum(flows) = closing for each column.

═══════════════════════════════════════════════════
STEP 4 — PRE-MAP TO SCHEMA FIELDS
═══════════════════════════════════════════════════

For every table, map each cell using this exact format:

  ELEMENT {idx} | text_table | "{title}"
  series="{row label verbatim}" period="{column header verbatim}" value="{cell value verbatim}" row_type="{type}" level={n}

row_type rules:
  total  → bold row (grand total, subtotal)
  sub    → indented row; add parent="{nearest non-indented label above it}"
  note   → footnote row
  data   → everything else

level rules:
  0 → grand total / section header
  1 → primary line item
  2 → indented sub-item

value rules:
  Copy verbatim. Keep commas, dashes, parentheses, % symbols.
  Shaded cell with no value → value=""
  Printed dash → value="-"

For any extra columns (YoY%, QoQ%, Change):
  add as key=value pairs on the same line with a matching _label key
  e.g. yoy_pct="-5%" yoy_pct_label="YoY %"

Complete this for EVERY table on the slide before finishing.
This pre-mapping is what Pass 2 will use — accuracy here is critical."""


# ===========================================================================
# PASS 2 PROMPT
# ===========================================================================
def build_pass2_prompt(description: str, bank: str) -> str:
    return f"""Transcribe the Step 4 pre-mapping below into structured JSON.
Do NOT re-interpret. Do NOT look at any image. Copy values exactly as written in Step 4.

<slide_description>
{description}
</slide_description>

OUTPUT FORMAT:
{{
  "slide_title": "main heading from the slide description — not 'Slide N'",
  "elements": [
    {{
      "element_idx": 0,
      "element_type": "text_table",
      "element_title": "...",
      "source": "table",
      "units": "...",
      "self_check": "arithmetic check from Step 3 if present, else null",
      "data_points": [
        {{
          "series": "...",
          "period": "...",
          "value": "...",
          "row_type": "...",
          "level": 1,
          "parent": null,
          "group": null,
          "sign": null,
          "order": 0
        }}
      ]
    }}
  ]
}}

RULES:
- One data_point per line in Step 4. Transcribe field values exactly.
- order: sequential integer per element, 0-based, following Step 4 line order.
- source: "table" for text_table and npa_movement_table.
- self_check: copy the arithmetic check string from Step 3, else null.
- Extra key=value pairs from Step 4 go into extra_fields dict.
  Always add a matching _label key: yoy_pct="-3%" → yoy_pct_label="YoY %"

Return ONLY the JSON object, no markdown fences."""


# ===========================================================================
# CORRECTION PROMPT
# ===========================================================================
def build_correction_prompt(errors: list[str], description: str) -> str:
    return f"""Your previous extraction had these validation errors:
{chr(10).join(f"  - {e}" for e in errors)}

Using the slide description as reference:
<slide_description>
{description}
</slide_description>

For any self_check arithmetic failure: re-examine the values or signs in the
affected element so the equation holds. Write the corrected self_check.

Re-extract ONLY the elements with errors, correcting these specific issues.
Return the same JSON format, only including the affected elements.
Return ONLY the JSON object, no markdown fences."""


# ===========================================================================
# HELPERS
# ===========================================================================
def render_page(pdf_path: str, page_1based: int, scale: float = IMAGE_SCALE) -> bytes:
    src = pdfium.PdfDocument(pdf_path)
    pil = src[page_1based - 1].render(scale=scale).to_pil()
    buf = io.BytesIO()
    pil.save(buf, format="PNG")
    return buf.getvalue()


def detect_bank(pdf_path: str) -> tuple[str | None, str | None]:
    try:
        pdf = pdfium.PdfDocument(pdf_path)
        txt = pdf[0].get_textpage().get_text_range()
        if len(pdf) > 1:
            txt += " " + pdf[1].get_textpage().get_text_range()
    except Exception:
        txt = ""
    key = None
    for k, info in BANKS.items():
        if re.search(info["match"], txt, re.I):
            key = k
            break
    m = re.search(r"\b(\d{1,2}\s+[A-Za-z]+\s+\d{4})\b", txt)
    return key, (m.group(1) if m else None)


def call_gemini(client, prompt_parts: list, *, text_only: bool = False) -> tuple[str, dict]:
    """Single Gemini call with retry. Returns (text, usage_rec)."""
    config_kwargs: dict = {"temperature": 0.0}
    if not text_only:
        config_kwargs["response_mime_type"] = "application/json"
    try:
        config_kwargs["thinking_config"] = types.ThinkingConfig(thinking_budget=0)
    except Exception:
        pass
    config = types.GenerateContentConfig(**config_kwargs)

    resp = None
    for attempt in range(3):
        try:
            resp = client.models.generate_content(
                model=MODEL,
                contents=prompt_parts,
                config=config,
            )
            break
        except Exception as e:
            if any(c in str(e) for c in ("503", "429", "UNAVAILABLE", "RESOURCE_EXHAUSTED")):
                wait = 15 * (2 ** attempt)
                print(f"      ⏳ {e.__class__.__name__} — waiting {wait}s")
                time.sleep(wait)
            else:
                raise

    if resp is None:
        raise RuntimeError("Gemini failed after 3 retries")

    text = (resp.text or "").strip()
    um   = getattr(resp, "usage_metadata", None)
    pt   = getattr(um, "prompt_token_count", None) or 0
    ot   = getattr(um, "candidates_token_count", None) or 0
    cost = (pt / 1e6 * INPUT_PRICE_PER_M) + (ot / 1e6 * OUTPUT_PRICE_PER_M)
    usage = {"prompt_tokens": pt, "output_tokens": ot, "est_cost_usd": round(cost, 6)}
    _run_usage["calls"]  += 1
    _run_usage["prompt"] += pt
    _run_usage["output"] += ot
    _run_usage["cost"]   += cost
    return text, usage


def img_part(img_bytes: bytes) -> types.Part:
    return types.Part.from_bytes(
        data=img_bytes, mime_type="image/png",
        media_resolution=types.PartMediaResolutionLevel.MEDIA_RESOLUTION_HIGH,
    )


def strip_fences(raw: str) -> str:
    s = raw.strip()
    if s.startswith("```"):
        s = s.split("```", 1)[1].lstrip("json").strip()
        s = s.rsplit("```", 1)[0].strip()
    return s


def _auto_label(key: str) -> str:
    """Generate a human-readable label from a snake_case key."""
    return key.replace("_", " ").title()


def parse_pass2(raw: str, bank: str, doc_title: str, doc_date: str,
                slide_num: int) -> tuple[list[DataPoint], str, list[str]]:
    """Returns (points, slide_title, self_checks)."""
    data = json.loads(strip_fences(raw))
    slide_title  = data.get("slide_title", "")
    points: list[DataPoint] = []
    self_checks: list[str] = []
    known = {"series", "period", "value", "row_type", "level",
             "parent", "group", "sign", "order"}

    for elem in data.get("elements", []):
        sc = elem.get("self_check")
        if sc:
            self_checks.append(f"[{elem.get('element_title', '')}] {sc}")

        for i, dp in enumerate(elem.get("data_points", [])):
            # Normalise row_type
            rt = str(dp.get("row_type") or "data")
            rt = ROW_TYPE_ALIASES.get(rt, rt)

            extra = {k: v for k, v in dp.items() if k not in known}

            # Auto-generate _label for any extra key missing one
            for k in list(extra):
                if not k.endswith("_label") and (k + "_label") not in extra:
                    extra[k + "_label"] = _auto_label(k)

            known_vals = {k: dp.get(k) for k in known}
            known_vals["row_type"] = rt
            # Coerce nulls to safe defaults so Pydantic doesn't reject them
            if known_vals.get("order") is None:
                known_vals["order"] = i
            if known_vals.get("level") is None:
                known_vals["level"] = 1
            if known_vals.get("series") is None:
                known_vals["series"] = ""
            if known_vals.get("value") is None:
                known_vals["value"] = ""

            points.append(DataPoint(
                slide=slide_num,
                slide_title=slide_title or "",
                element_idx=elem["element_idx"],
                element_type=elem.get("element_type") or "other",
                element_title=elem.get("element_title") or "",
                source=elem.get("source") or "table",
                unit=elem.get("units") or "",
                bank=bank,
                doc_title=doc_title,
                doc_date=doc_date,
                extra_fields=extra,
                **known_vals,
            ))
    return points, slide_title, self_checks


def merge_correction(original: list[DataPoint],
                     corrected: list[DataPoint]) -> list[DataPoint]:
    """Replace elements from original with corrected versions where idx matches."""
    corrected_idxs = {p.element_idx for p in corrected}
    merged = [p for p in original if p.element_idx not in corrected_idxs]
    merged.extend(corrected)
    merged.sort(key=lambda p: (p.element_idx, p.order))
    return merged


# ===========================================================================
# PDFPLUMBER CROSS-CHECK
# ===========================================================================
def _plumber_values(pdf_path: str, page_0based: int) -> set[str]:
    """Extract all numeric strings from native text layer via pdfplumber."""
    if not _HAS_PDFPLUMBER:
        return set()
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if page_0based >= len(pdf.pages):
                return set()
            text = pdf.pages[page_0based].extract_text() or ""
        # collect tokens that look like financial figures
        return {t for t in re.findall(r'[\d,]+(?:\.\d+)?', text) if len(t) > 1}
    except Exception:
        return set()


def _is_clean_table_page(pdf_path: str, page_0based: int) -> bool:
    """Return True if the page has a meaningful native text layer (not a pure chart slide)."""
    if not _HAS_PDFPLUMBER:
        return False
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if page_0based >= len(pdf.pages):
                return False
            tables = pdf.pages[page_0based].extract_tables()
            return bool(tables)
    except Exception:
        return False


def crosscheck_with_textlayer(points: list[DataPoint],
                              pdf_path: str, page_num: int) -> list[DataPoint]:
    """
    For text_table elements on pages that have a native table layer:
    mark any value NOT found in the text layer as source="unverified".
    unverified renders yellow the same as chart values.
    """
    if not _HAS_PDFPLUMBER:
        return points

    page_0 = page_num - 1
    if not _is_clean_table_page(pdf_path, page_0):
        return points

    native_vals = _plumber_values(pdf_path, page_0)
    if not native_vals:
        return points

    updated = []
    for p in points:
        if p.element_type == "text_table" and p.source == "table":
            # Strip commas/signs from the extracted value for comparison
            core = re.sub(r'[,\+\(\)]', '', p.value).strip().lstrip("-")
            if core and core not in native_vals:
                p = p.model_copy(update={"source": "unverified"})
        updated.append(p)
    return updated


# ===========================================================================
# VALIDATION
# ===========================================================================
def validate_self_check(self_check: str | None) -> str | None:
    if not self_check:
        return None
    try:
        # Strip leading "[Element title] " prefix if present
        equation = re.sub(r'^\[.*?\]\s*', '', self_check).strip()
        lhs, rhs = equation.split("=")
        expected = float(rhs.strip().replace(",", ""))
        actual   = eval(lhs.strip().replace(",", ""), {"__builtins__": {}})
        delta    = abs(actual - expected)
        if delta > 5:
            return (f"Self-check failed: {self_check.strip()} "
                    f"(off by {delta:.0f})")
        return None
    except Exception:
        return f"Self-check unparseable: {self_check}"


def validate(points: list[DataPoint], self_checks: list[str] | None = None) -> list[str]:
    errors: list[str] = []

    # Generic arithmetic self-checks written by Gemini during extraction
    for sc in (self_checks or []):
        err = validate_self_check(sc)
        if err:
            errors.append(err)

    # Structural checks (element-type agnostic)
    by_elem: dict[int, list[DataPoint]] = {}
    for p in points:
        by_elem.setdefault(p.element_idx, []).append(p)

    for pts in by_elem.values():
        title = pts[0].element_title
        etype = pts[0].element_type

        blank = [p for p in pts if not p.series.strip()]
        if blank:
            errors.append(f"Element '{title}': {len(blank)} rows with blank labels")

        if etype == "text_table":
            illegible = [p.series for p in pts if p.value == "?"]
            if illegible:
                errors.append(f"Table '{title}': illegible values in rows: {illegible}")

    return errors


# ===========================================================================
# PASS 3 — EXCEL RENDER
# ===========================================================================
def coerce(v: Any) -> Any:
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "–", "—", "n.m.", "nm", "NA", "N/A", "?"):
        return s
    if s.endswith("%") or s.endswith("bps") or s.endswith("x"):
        return s
    t = s.replace(",", "")
    neg = t.startswith("(") and t.endswith(")")
    core = t[1:-1] if neg else t
    try:
        num = float(core.replace("+", ""))
        if neg:
            num = -num
        return int(num) if num == int(num) else num
    except ValueError:
        return s


def tab_name(used: set, slide_num: int, slide_title: str) -> str:
    prefix = f"{slide_num:02d}"
    short  = re.sub(r'[\\/:*?\[\]]', '', slide_title).strip()[:20] or "Slide"
    base   = f"{prefix} - {short}"[:31]
    name, i = base, 2
    while name in used:
        suffix = f" ({i})"
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name


def _hdr(cell, dark: bool = False):
    cell.fill = PatternFill("solid", fgColor=DARK_GREY if dark else NAVY)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


# ===========================================================================
# PASS 3 — EXCEL RENDER (wide-format pivot)
# ===========================================================================

def group_by_elem(points: list[DataPoint]) -> dict[int, list[DataPoint]]:
    result: dict[int, list[DataPoint]] = {}
    for p in points:
        result.setdefault(p.element_idx, []).append(p)
    return result


def _ordered_periods(pts: list[DataPoint]) -> list[str | None]:
    seen: list[str | None] = []
    seen_set: set = set()
    for p in sorted(pts, key=lambda x: x.order):
        if p.period not in seen_set:
            seen.append(p.period)
            seen_set.add(p.period)
    return seen


def _ordered_series(pts: list[DataPoint]) -> list[str]:
    seen: list[str] = []
    seen_set: set[str] = set()
    for p in sorted(pts, key=lambda x: x.order):
        if p.series not in seen_set:
            seen.append(p.series)
            seen_set.add(p.series)
    return seen


def _extra_keys_for_period(pts: list[DataPoint], period: str | None) -> list[str]:
    keys: list[str] = []
    seen: set[str] = set()
    for p in (x for x in pts if x.period == period):
        for k in (p.extra_fields or {}):
            if not k.endswith("_label") and k not in seen:
                keys.append(k)
                seen.add(k)
    return keys


def _extra_col_header(key: str, pts: list[DataPoint]) -> str:
    label_key = key + "_label"
    for p in pts:
        if label_key in (p.extra_fields or {}):
            return p.extra_fields[label_key]
    return key.replace("_", " ").title()


def build_pivot(pts: list[DataPoint]) -> dict:
    periods = _ordered_periods(pts)
    series  = _ordered_series(pts)
    is_kpi  = (len(periods) == 1 and periods[0] is None)

    cells: dict[tuple, DataPoint] = {}
    for p in pts:
        key = (p.series, p.period)
        if key not in cells:
            cells[key] = p

    meta: dict[str, DataPoint] = {}
    for p in sorted(pts, key=lambda x: x.order):
        if p.series not in meta:
            meta[p.series] = p

    extra = {period: _extra_keys_for_period(pts, period) for period in periods}

    return {
        "is_kpi":  is_kpi,
        "periods": periods,
        "series":  series,
        "extra":   extra,
        "cells":   cells,
        "meta":    meta,
    }


def _build_col_layout(pivot: dict) -> list[dict]:
    cols: list[dict] = []
    if pivot["is_kpi"]:
        cols.append({"kind": "label", "header": "Metric"})
        for s in pivot["series"]:
            dp = pivot["meta"][s]
            cols.append({"kind": "kpi", "series": s, "header": s, "unit": dp.unit or ""})
    else:
        cols.append({"kind": "label", "header": ""})
        for period in pivot["periods"]:
            period_label = str(period) if period else "Value"
            cols.append({"kind": "value", "period": period, "header": period_label})
            for key in pivot["extra"].get(period, []):
                header = _extra_col_header(key, list(pivot["cells"].values()))
                cols.append({"kind": "extra", "period": period, "key": key, "header": header})
    return cols


def _compute_max_cols(points: list[DataPoint]) -> int:
    by_elem = group_by_elem(points)
    max_c = 1
    for pts in by_elem.values():
        pivot = build_pivot(pts)
        cols  = _build_col_layout(pivot)
        max_c = max(max_c, len(cols))
    return max_c


def render_element(ws, pts: list[DataPoint], cursor: int, brand: str) -> int:
    if not pts:
        return cursor

    elem_title = pts[0].element_title
    elem_type  = pts[0].element_type
    unit       = pts[0].unit or "S$m"

    ws.cell(cursor, 1, elem_title).font = Font(bold=True, color=brand, size=10)
    cursor += 1

    pivot  = build_pivot(pts)
    cols   = _build_col_layout(pivot)
    n_cols = len(cols)

    # ── Header row ────────────────────────────────────────────────────────────
    if pivot["is_kpi"]:
        for ci, col in enumerate(cols, 1):
            _hdr(ws.cell(cursor, ci, col["header"]), dark=(ci == 1))
        ws.row_dimensions[cursor].height = 20
        cursor += 1

        for ci, col in enumerate(cols, 1):
            if col["kind"] == "label":
                ws.cell(cursor, ci, "Value")
            elif col["kind"] == "kpi":
                dp  = pivot["cells"].get((col["series"], None))
                val = coerce(dp.value) if dp else ""
                cell = ws.cell(cursor, ci, val)
                if isinstance(val, (int, float)):
                    cell.number_format = NUM_FMT
                    cell.alignment = Alignment(horizontal="right")
                if dp and dp.source in ("chart", "unverified") and val not in (None, "", "-", "?"):
                    cell.fill = PatternFill("solid", fgColor=YELLOW)
        cursor += 1

    else:
        cols[0]["header"] = f"({unit})" if unit else ""
        for ci, col in enumerate(cols, 1):
            _hdr(ws.cell(cursor, ci, col["header"]), dark=(ci == 1))
        ws.row_dimensions[cursor].height = 20
        cursor += 1

        for series in pivot["series"]:
            dp_meta   = pivot["meta"][series]
            is_total  = dp_meta.row_type in ("total", "start", "end")
            is_sub    = dp_meta.level >= 2
            is_bridge = dp_meta.row_type == "bridge"
            is_note   = dp_meta.row_type == "note"
            indent    = "    " * max(0, dp_meta.level - 1)

            for ci, col in enumerate(cols, 1):
                kind = col["kind"]

                if kind == "label":
                    cell = ws.cell(cursor, ci, indent + series)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                elif kind == "value":
                    dp  = pivot["cells"].get((series, col["period"]))
                    val = coerce(dp.value) if dp else ""
                    cell = ws.cell(cursor, ci, val)
                    if isinstance(val, (int, float)):
                        cell.number_format = NUM_FMT
                        cell.alignment = Alignment(horizontal="right")
                    if dp and dp.source in ("chart", "unverified") and val not in (None, "", "-", "?"):
                        cell.fill = PatternFill("solid", fgColor=YELLOW)
                    if is_bridge and isinstance(val, (int, float)):
                        colour = "00B050" if val > 0 else "C00000"
                        cell.font = Font(color=colour, size=10, bold=is_total)
                        if is_total:
                            cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
                        cursor += 0  # font set; skip generic block
                        continue
                elif kind == "extra":
                    dp  = pivot["cells"].get((series, col["period"]))
                    raw = (dp.extra_fields or {}).get(col["key"], "") if dp else ""
                    val = coerce(raw) if raw != "" else ""
                    cell = ws.cell(cursor, ci, val)
                    cell.font = Font(color=MID_GREY, size=9)
                    cell.alignment = Alignment(horizontal="right")
                    continue
                else:
                    cell = ws.cell(cursor, ci, "")

                # Generic row styling
                if is_total:
                    cell.fill = PatternFill("solid", fgColor=TOTAL_BG)
                    cell.font = Font(bold=True, size=10)
                elif is_note:
                    cell.font = Font(italic=True, color=MID_GREY, size=8)
                elif is_sub:
                    cell.font = Font(italic=True, color=MID_GREY, size=9)
                else:
                    cell.font = Font(size=10)

            cursor += 1

        # Waterfall balance check row
        if elem_type == "waterfall":
            starts  = [p for p in pts if p.row_type == "start"]
            ends    = [p for p in pts if p.row_type == "end"]
            bridges = [p for p in pts if p.row_type == "bridge"]
            if starts and ends and bridges:
                opening = starts[0].value_num or 0
                closing = ends[0].value_num or 0
                total   = sum(p.value_num or 0 for p in bridges)
                delta   = abs(opening + total - closing)
                ok      = delta <= 5
                msg = (f"✓ Bridge balances: {opening:,.0f} + {total:+,.0f} = {closing:,.0f}"
                       if ok else
                       f"⚠ Bridge off by {delta:.0f} — check signs in source")
                c = ws.cell(cursor, 1, msg)
                c.font = Font(italic=True, size=8, color="00B050" if ok else "FF0000")
                ws.merge_cells(start_row=cursor, start_column=1,
                               end_row=cursor, end_column=n_cols)
                cursor += 1

    return cursor + 1  # 1-row spacer


def render_slide(ws, points: list[DataPoint], slide_num: int, slide_title: str,
                 bank: str, doc_title: str, doc_date: str, brand: str = None):
    brand   = brand or BRAND_COLOURS.get(bank, "404040")
    by_elem = group_by_elem(points)
    max_cols = _compute_max_cols(points)

    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=max(max_cols, 4))
    c = ws.cell(1, 1, f"{bank} Group  |  Slide {slide_num}: {slide_title}")
    c.fill = PatternFill("solid", fgColor=brand)
    c.font = Font(bold=True, color=WHITE, size=12)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2, end_column=max(max_cols, 4))
    c = ws.cell(2, 1,
        f"Source: {doc_title}"
        + (f", {doc_date}" if doc_date else "")
        + "  |  Units: S$m unless noted  |  ⚠ Yellow = chart-sourced, verify against slide")
    c.font = Font(italic=True, color=MID_GREY, size=9)

    cursor = 4
    for elem_idx in sorted(by_elem):
        cursor = render_element(ws, by_elem[elem_idx], cursor, brand)

    ws.column_dimensions[get_column_letter(1)].width = 44
    for i in range(2, (ws.max_column or 2) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.freeze_panes = "B4"


def build_contents(wb, index: list[dict], bank: str, doc_title: str,
                   doc_date: str, brand: str):
    if "Contents" in wb.sheetnames:
        wb.remove(wb["Contents"])
    ws = wb.create_sheet("Contents", 0)

    ws.merge_cells("A1:E1")
    c = ws.cell(1, 1, BANKS[bank]["institution"] if bank in BANKS else bank)
    c.fill = PatternFill("solid", fgColor=brand)
    c.font = Font(bold=True, color=WHITE, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:E2")
    c = ws.cell(2, 1, f"{doc_title}{' — ' + doc_date if doc_date else ''}")
    c.font = Font(italic=True, size=11)
    c.alignment = Alignment(horizontal="center")

    for ci, h in enumerate(["Slide", "Title", "Points", "Has Charts", "Sheet"], 1):
        _hdr(ws.cell(4, ci, h))

    for r, e in enumerate(sorted(index, key=lambda x: x["slide_num"]), start=5):
        ws.cell(r, 1, e["slide_num"]).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, e["slide_title"])
        ws.cell(r, 3, e["n_points"]).alignment = Alignment(horizontal="center")
        has_chart = "yes" if e.get("has_chart") else "no"
        c = ws.cell(r, 4, has_chart)
        c.alignment = Alignment(horizontal="center")
        if has_chart == "yes":
            c.font = Font(color="FF6600")
        link = ws.cell(r, 5, e["sheet"])
        link.hyperlink = f"#'{e['sheet']}'!A1"
        link.font = Font(color="0000CC", underline="single")

    for col, w in {1: 8, 2: 60, 3: 10, 4: 12, 5: 32}.items():
        ws.column_dimensions[get_column_letter(col)].width = w


# ===========================================================================
# SINGLE-PASS EXTRACTOR (visual slides)
# ===========================================================================
def extract_single_pass(client, img_bytes: bytes, bank: str,
                        doc_title: str, doc_date: str,
                        slide_num: int) -> tuple[list[DataPoint], float]:
    """One call: image + JSON schema prompt. No intermediate description."""
    raw, usage = call_gemini(client, [img_part(img_bytes), SINGLE_PASS_PROMPT])
    try:
        points, _, self_checks = parse_pass2(
            raw, bank, doc_title, doc_date, slide_num
        )
    except Exception as e:
        print(f"  slide {slide_num:02d}  ❌ single-pass parse error: {e}")
        return [], usage["est_cost_usd"]
    return points, usage["est_cost_usd"]


# ===========================================================================
# SLIDE PROCESSOR
# ===========================================================================
def process_slide(client, pdf_path: str, page_num: int,
                  bank: str, doc_title: str, doc_date: str,
                  audit_dir: str,
                  force: bool = False) -> tuple[list[DataPoint], float]:

    audit_path   = os.path.join(audit_dir, f"slide_{page_num:02d}")
    types_path   = os.path.join(audit_path, "element_types.json")
    img_path     = os.path.join(audit_path, f"slide_{page_num:02d}.png")
    desc_path    = os.path.join(audit_path, "description.txt")
    p1_path      = os.path.join(audit_path, "pass1_prompt.txt")
    p2_path      = os.path.join(audit_path, "pass2_prompt.txt")
    dp_path      = os.path.join(audit_path, "datapoints.json")
    meta_path    = os.path.join(audit_path, "meta.json")
    os.makedirs(audit_path, exist_ok=True)

    # Resume from audit
    if not force and os.path.exists(dp_path):
        raw = json.load(open(dp_path))
        points = [DataPoint(**p) for p in raw]
        print(f"  slide {page_num:02d}  resumed ({len(points)} points)")
        return points, 0.0

    img_bytes  = render_page(pdf_path, page_num)
    with open(img_path, "wb") as f:
        f.write(img_bytes)
    total_cost = 0.0
    usages: list[dict] = []

    # ── Pass 0: Classify ──
    if not force and os.path.exists(types_path):
        with open(types_path) as f:
            types_found = json.load(f)
        print(f"  slide {page_num:02d}  types resumed: {types_found}")
    else:
        types_found = classify_slide(client, img_bytes)
        with open(types_path, "w") as f:
            json.dump(types_found, f)

    known_types, unknown_types = split_known_unknown(types_found)

    # Skip slides with no data elements
    if types_found == ["none"] or not types_found:
        print(f"  slide {page_num:02d}  — no data elements, skipping")
        return [], 0.0

    # ── Route by element type ──────────────────────────────────────────────
    has_visual = any(t in VISUAL_TYPES for t in known_types)

    if has_visual:
        print(f"  slide {page_num:02d}  → single-pass (visual: {known_types})")
        raw1, u1 = call_gemini(client, [img_part(img_bytes), SINGLE_PASS_PROMPT])
        total_cost += u1["est_cost_usd"]
        usages.append({"pass": "1s", **u1})
        desc_text = "single-pass — no Pass 1 description"
        with open(p1_path, "w") as f:
            f.write(SINGLE_PASS_PROMPT)
        with open(desc_path, "w") as f:
            f.write(desc_text)
        try:
            points, parsed_title, self_checks = parse_pass2(
                raw1, bank, doc_title, doc_date, page_num
            )
        except Exception as e:
            print(f"  slide {page_num:02d}  ❌ single-pass parse error: {e}")
            return [], total_cost

    else:
        print(f"  slide {page_num:02d}  → multi-pass (text only: {known_types})")
        # ── Pass 1: Describe with text-table contracts ──
        contracts     = load_contracts()
        known_block   = build_contracts_block(known_types, contracts)
        unknown_block = build_unknown_contracts_block(unknown_types)
        pass1_prompt  = PASS1_PROMPT + known_block + unknown_block

        with open(p1_path, "w") as f:
            f.write(pass1_prompt)

        desc_text, u1 = call_gemini(
            client, [img_part(img_bytes), pass1_prompt], text_only=True,
        )
        total_cost += u1["est_cost_usd"]
        usages.append({"pass": 1, **u1})

        with open(desc_path, "w") as f:
            f.write(desc_text)

        if not desc_text.strip():
            print(f"  slide {page_num:02d}  — nothing described, skipped")
            return [], total_cost

        save_derived_contract(desc_text, unknown_types)

        # ── Pass 2: Text-only transcription ──
        p2_prompt = build_pass2_prompt(desc_text, bank)
        with open(p2_path, "w") as f:
            f.write(p2_prompt)
        raw2, u2 = call_gemini(client, [p2_prompt], text_only=False)
        total_cost += u2["est_cost_usd"]
        usages.append({"pass": 2, **u2})

        try:
            points, parsed_title, self_checks = parse_pass2(
                raw2, bank, doc_title, doc_date, page_num
            )
        except Exception as e:
            print(f"  slide {page_num:02d}  ❌ parse error: {e}")
            return [], total_cost

    # slide_title: prefer Pass 2 output, fall back to first TITLE: line in description
    if not parsed_title:
        m = re.search(r'(?:^|\n)(?:TITLE:|##\s+|#\s+)(.+)', desc_text)
        parsed_title = m.group(1).strip() if m else ""
    if parsed_title:
        points = [p.model_copy(update={"slide_title": parsed_title}) for p in points]

    # ── pdfplumber cross-check ──
    points = crosscheck_with_textlayer(points, pdf_path, page_num)

    # ── Validate ──
    errors = validate(points, self_checks)

    # ── Correction pass (once) ──
    if errors:
        print(f"  slide {page_num:02d}  ⚠ validation: {errors}")
        corr_prompt = build_correction_prompt(errors, desc_text)
        # Single-pass: re-attach image since there is no text description to reason from
        corr_parts: list = [corr_prompt] if not has_visual else [img_part(img_bytes), corr_prompt]
        try:
            raw3, u3 = call_gemini(client, corr_parts, text_only=False)
            total_cost += u3["est_cost_usd"]
            usages.append({"pass": "2c", **u3})
            corrected, _, self_checks_c = parse_pass2(raw3, bank, doc_title, doc_date, page_num)
            corrected = crosscheck_with_textlayer(corrected, pdf_path, page_num)
            errors_after = validate(corrected, self_checks_c)
            if len(errors_after) < len(errors):
                points = merge_correction(points, corrected)
                errors = errors_after
                self_checks = self_checks_c
                print(f"  slide {page_num:02d}  ✓ correction improved output")
            else:
                print(f"  slide {page_num:02d}  ⚠ correction did not help, using original")
        except Exception as e:
            print(f"  slide {page_num:02d}  ⚠ correction failed: {e}")

    # ── Save audit ──
    with open(dp_path, "w") as f:
        json.dump([p.model_dump() for p in points], f, indent=2)
    with open(meta_path, "w") as f:
        json.dump({"slide": page_num, "cost_usd": total_cost,
                   "n_points": len(points), "validation_errors": errors,
                   "self_checks": self_checks, "usages": usages}, f, indent=2)

    status     = "✓" if not errors else f"⚠ {len(errors)} err"
    unverified = sum(1 for p in points if p.source == "unverified")
    flags      = (" 📊" if any(p.source == "chart" for p in points) else "")
    flags     += (f" ⚡{unverified}unverified" if unverified else "")
    in_t  = sum(u.get("prompt_tokens", 0) for u in usages)
    out_t = sum(u.get("output_tokens", 0) for u in usages)
    print(f"  slide {page_num:02d}  {len(points)} pts  "
          f"[{in_t}in/{out_t}out]  ${total_cost:.4f}  {status}{flags}")

    return points, total_cost


# ===========================================================================
# MAIN
# ===========================================================================
def main():
    ap = argparse.ArgumentParser(description="CFO slide deck PDF → Excel (4-pass)")
    ap.add_argument("pdf")
    ap.add_argument("--out",         default=None)
    ap.add_argument("--slide",       type=int, help="extract this slide only (1-based)")
    ap.add_argument("--start-slide", type=int, help="resume from this slide number")
    ap.add_argument("--bank",        choices=list(BANKS))
    ap.add_argument("--doc-date",    default="")
    ap.add_argument("--force",       action="store_true",
                    help="re-extract even if audit exists")
    ap.add_argument("--dry-run",     action="store_true",
                    help="list slides without calling API")
    args = ap.parse_args()
    doc_start_time = time.time()

    if not os.path.exists(args.pdf):
        sys.exit(f"PDF not found: {args.pdf}")

    detected, det_date = detect_bank(args.pdf)
    bank      = args.bank or detected or "DBS"
    doc_date  = args.doc_date or det_date or ""
    doc_title = os.path.basename(args.pdf).replace(".pdf", "")
    bank_slug = bank.lower()
    brand     = BANKS[bank]["brand"]

    out_path  = args.out or f"outputs/{bank_slug}_slides.xlsx"
    audit_dir = f"outputs/audit/slides/{bank_slug}_{doc_title}"
    os.makedirs(os.path.dirname(out_path), exist_ok=True)

    pdf     = pdfium.PdfDocument(args.pdf)
    n_pages = len(pdf)
    if args.slide:
        pages = [args.slide]
    elif args.start_slide:
        pages = list(range(args.start_slide, n_pages + 1))
    else:
        pages = list(range(1, n_pages + 1))

    print(f"🎞  {args.pdf}  ({n_pages} slides)  bank={bank}  model={MODEL}  [hybrid: single-pass visual / multi-pass text]")
    print(f"    Output → {out_path}  |  slides: {len(pages)}")

    if args.dry_run:
        print(f"\nDRY RUN — {len(pages)} slide(s) would be processed:")
        for p in pages:
            print(f"  slide {p:02d}")
        return

    client = genai.Client()

    # Load existing workbook when running a subset so other tabs are preserved.
    # Full runs (all pages) always start fresh to avoid stale tabs.
    is_subset = bool(args.slide or args.start_slide)
    if is_subset and os.path.exists(out_path):
        wb = openpyxl.load_workbook(out_path)
        # Rebuild index from existing sheets (skip Contents tab)
        index: list[dict] = []
        for sname in wb.sheetnames:
            if sname == "Contents":
                continue
            # Parse slide number from tab name prefix "NN - ..."
            try:
                slide_num = int(sname.split(" - ")[0])
            except ValueError:
                continue
            ws_existing = wb[sname]
            index.append({
                "slide_num":   slide_num,
                "slide_title": ws_existing.cell(1, 1).value or sname,
                "sheet":       sname,
                "n_points":    0,   # not critical for Contents rebuild
                "has_chart":   False,
            })
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        index: list[dict] = []

    used_names: set[str] = set(wb.sheetnames)
    total_cost = 0.0

    for pg in pages:
        try:
            points, cost = process_slide(
                client, args.pdf, pg,
                bank, doc_title, doc_date,
                audit_dir,
                force=args.force,
            )
        except Exception as e:
            print(f"  slide {pg:02d}  ❌ {e}")
            continue

        total_cost += cost

        if not points:
            continue

        slide_title = points[0].slide_title or f"Slide {pg}"
        sname = tab_name(used_names, pg, slide_title)
        if sname in wb.sheetnames:
            del wb[sname]
        ws = wb.create_sheet(title=sname)

        render_slide(ws, points, pg, slide_title,
                     bank, doc_title, doc_date, brand)

        entry = {
            "slide_num":   pg,
            "slide_title": slide_title,
            "sheet":       sname,
            "n_points":    len(points),
            "has_chart":   any(p.source in ("chart", "unverified") for p in points),
        }
        # Replace existing index entry for this slide if present, else append
        existing = next((i for i, e in enumerate(index) if e["slide_num"] == pg), None)
        if existing is not None:
            index[existing] = entry
        else:
            index.append(entry)

        build_contents(wb, index, bank, doc_title, doc_date, brand)
        wb.save(out_path)
        print(f"    💾 saved  |  running cost ≈ ${total_cost:.4f}")

    doc_elapsed = time.time() - doc_start_time
    u = _run_usage

    # ── Per-document summary ──────────────────────────────────────
    summary = {
        "document":          os.path.basename(args.pdf),
        "bank":              bank,
        "slides_processed":  len(pages),
        "api_calls":         u["calls"],
        "input_tokens":      u["prompt"],
        "output_tokens":     u["output"],
        "est_cost_usd":      round(u["cost"], 5),
        "elapsed_seconds":   round(doc_elapsed, 1),
        "elapsed_human":     f"{int(doc_elapsed // 60)}m {int(doc_elapsed % 60)}s",
        "generated_at":      datetime.datetime.now().isoformat(timespec="seconds"),
        "output_file":       out_path,
    }

    summary_path = os.path.splitext(out_path)[0] + "_run_summary.json"
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)

    print(f"\n✅ Done → {out_path}")
    print(f"   {u['calls']} API calls  "
          f"input={u['prompt']:,}  output={u['output']:,}  "
          f"≈ ${u['cost']:.4f}  |  ⏱ {summary['elapsed_human']}")
    print(f"   Run summary → {summary_path}")


if __name__ == "__main__":
    main()
