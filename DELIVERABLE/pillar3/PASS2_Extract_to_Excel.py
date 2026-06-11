"""
extract_to_excel.py — simplified, auditable PDF -> Excel extraction.

No docling, no heavy intermediate JSON. The model returns a COMPACT grid
(3 hierarchy columns + label + values) which a deterministic writer drops
straight into one Excel tab per section, matching the chat-extracted workbook.

Pipeline:
  1. Read the deterministic TOC produced by build_toc.py (out/step1_toc.json) —
     a flat list of sections with physical page ranges. ZERO API calls there.
  2. Walk sections in document order. For each section, go PAGE BY PAGE:
       - cut the PDF to that single page (smallest, most accurate task),
       - skip obvious narrative pages (cheap numeric-density check, no call),
       - send the 1-page PDF NATIVELY (Gemini reads text + layout + visuals),
       - get back every table on the page as a compact grid.
     Tables flagged continued_from_previous are stitched onto the prior page's
     table (handles spanning tables). A rendered PNG is attached only with
     --image or when the first response looks unreasonable.
  3. Write the stitched tables into the section's tab (one tab per section).
  4. Pause after each section for review (unless --no-pause).

AUDIT: for every call we save, under out/audit/<unit_id>/:
   prompt.txt    — the exact prompt text sent
   pages.pdf     — the exact page-cut PDF sent to Gemini
   response.txt  — the raw model response
   parsed.json   — the parsed grid
   meta.json     — pages, mode, image_used, token usage
Token usage for every call is also appended to out/api_usage_log.jsonl.

Usage:
  export GEMINI_API_KEY=...
  python extract_to_excel.py DBS_4Q25_Pillar3.pdf                 # all sections, pause each
  python extract_to_excel.py DBS_4Q25_Pillar3.pdf --section 5.1   # one section only
  python extract_to_excel.py DBS_4Q25_Pillar3.pdf --no-pause      # run straight through
  python extract_to_excel.py DBS_4Q25_Pillar3.pdf --image         # force image alongside PDF
  python extract_to_excel.py DBS_4Q25_Pillar3.pdf --list          # list sections and exit
"""
from __future__ import annotations
import os, sys, json, io, re, argparse, datetime, time, threading, random, hashlib
from typing import Literal
from collections import defaultdict, Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import pypdfium2 as pdfium
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill, Alignment
from pydantic import BaseModel, Field, model_validator
import pdfplumber
from google import genai
from google.genai import types

# ---------------------------------------------------------------------------
MODEL          = "gemini-3.5-flash"
TOC_PATH       = "out/step1_toc.json"          # produced by build_toc.py (zero API)
OUT_XLSX       = "out/sections.xlsx"
AUDIT_DIR      = "out/audit"
USAGE_LOG_PATH = "out/api_usage_log.jsonl"
COST_LOG_PATH  = "out/api_cost_summary.json"
INDEX_PATH     = "out/sections_index.json"   # section -> sheet map, for the Contents sheet
IMAGE_SCALE    = 2.0           # PDF points -> pixels for the fallback render

# --- Per-bank identity + brand colour (auto-detected; override with --bank) ---
# brand colours per the skill: DBS & OCBC are red, UOB is blue.
BANKS = {
    "DBS":  {"institution": "DBS Group Holdings Ltd",
             "brand": "CC0000", "match": r"\bDBS\b"},
    "OCBC": {"institution": "Oversea-Chinese Banking Corporation Limited",
             "brand": "CC0000", "match": r"OCBC|Oversea[- ]?Chinese"},
    "UOB":  {"institution": "United Overseas Bank Limited",
             "brand": "1B6EC2", "match": r"\bUOB\b|United Overseas"},
}

# --- Document metadata + brand styling (set per-bank at runtime in main) -------
INSTITUTION  = "DBS Group Holdings Ltd"
DOC_TITLE    = "Pillar 3 Disclosures"
DOC_DATE     = "31 December 2025"
BRAND_COLOUR = "CC0000"   # overwritten per detected bank
HEADER_FILL  = "1F3864"   # column-header navy (never the brand colour)
DARK_GREY    = "404040"   # section_header row shading
MID_GREY     = "595959"   # metadata-column font / source line
WHITE        = "FFFFFF"
LIGHT_GREY   = "D9D9D9"   # grey cell fill (cell_state="grey")
NUM_FMT      = '#,##0;(#,##0);"-"'   # integer dollar amounts
# Metadata columns written before data columns in every row.
# N_META is derived from this list — add columns here, not by bumping N_META.
META_HEADERS = ["unique_row_id", "hierarchy_level", "parent_row_id", "Label"]
N_META       = len(META_HEADERS)

# Pricing: gemini-2.5-flash. Last verified 2026-06-05.
# Update when Gemini pricing changes: https://ai.google.dev/pricing
INPUT_PRICE_PER_M   = 0.30
OUTPUT_PRICE_PER_M  = 2.50
THINK_PRICE_PER_M   = 3.50

_run_usage = {"calls": 0, "prompt": 0, "output": 0, "thinking": 0, "cost": 0.0}
_call_log: list[dict] = []   # one record per API call, written to Cost sheet + summary JSON
_run_usage_lock = threading.Lock()
_call_log_lock  = threading.Lock()
_pdfium_lock    = threading.Lock()  # pypdfium2 is not thread-safe for concurrent opens

# ===========================================================================
# COMPACT OUTPUT SCHEMA  (the only thing Gemini returns — light, auditable)
# ===========================================================================
class GColumn(BaseModel):
    group: str | None = Field(default=None, description="2nd-level group header spanning sub-columns; null if single-level")
    leaf:  str = Field(description="the column header text — a full descriptive phrase; NEVER a bare letter like '(a)' or '(b)' which are reference indices, not headers")

# Cell state enum — 5 states capturing what is visually observable in the PDF.
# reported   : a numeric or text value is printed (numbers, %, text, #, <0.5)
# nil        : any dash variant printed ("-", "–", "—") — zero or negligible
# empty      : cell is truly blank with no mark
# grey       : cell is visually shaded/greyed — not applicable to this row/column
# zero       : printed "0" — explicitly zero (distinct from nil/empty)
CELL_STATES = {"reported", "nil", "empty", "grey", "zero"}

# Migration map for old cached parsed.json files that used legacy cell_state names
_LEGACY_CELL_STATES = {
    "suppressed":     "grey",       # old name for grey
    "rounds_to_zero": "reported",   # treat as reported verbatim value
}

class GCell(BaseModel):
    value:      str  = Field(description="cell value verbatim as printed; use '-' for any dash (-, –, —); '' ONLY for cells that are truly blank with absolutely no mark; '0' for printed zero")
    cell_state: Literal["reported", "nil", "empty", "grey", "zero"] = Field(default="reported",
                             description="reported | nil | empty | grey | zero")

    @model_validator(mode="before")
    @classmethod
    def _migrate_legacy(cls, obj):
        if isinstance(obj, str):
            return cls.from_str(obj).model_dump()
        if isinstance(obj, dict):
            cs = obj.get("cell_state")
            if cs in _LEGACY_CELL_STATES:
                obj = {**obj, "cell_state": _LEGACY_CELL_STATES[cs]}
        return obj

    @classmethod
    def from_str(cls, v: str) -> "GCell":
        """Upgrade a plain string (legacy parsed.json) to a GCell."""
        s = str(v).strip()
        if s in ("-", "–", "—"):
            return cls(value="-", cell_state="nil")
        if s == "0":
            return cls(value="0", cell_state="zero")
        if s == "":
            return cls(value="", cell_state="empty")
        return cls(value=s, cell_state="reported")

class GRow(BaseModel):
    row_id:   str | None = Field(default=None, description="printed line number EXACTLY as shown ('1','4a','14a'); null for rows with no printed number (section headers, sub-headers, footnotes)")
    row_type: Literal["section_header", "data", "total", "sub_header", "note"] = Field(default="data", description="section_header | data | total | sub_header | note")
    level:    int = Field(description="0=section header or grand total; 1=primary line item; 2=sub-item (indented / 'of which' / named breakdown); 3=rare")
    parent:   str | None = Field(default=None, description="null for level-0 and level-1 rows; for level-2+ the row_id of the nearest row one level above")
    label:    str = Field(description="row label text, verbatim, including footnote markers")
    values:   list[GCell] = Field(default_factory=list, description="cells left-to-right, one GCell per column; [] for section_header/sub_header/note rows")

    @model_validator(mode="before")
    @classmethod
    def _upgrade_string_values(cls, obj):
        if isinstance(obj, dict):
            vals = obj.get("values")
            if vals and isinstance(vals[0], str):
                obj = {**obj, "values": [GCell.from_str(v).model_dump() for v in vals]}
        return obj

class GTable(BaseModel):
    title:        str = Field(description="printed table title, verbatim, including the reporting date if shown")
    label_header: str = Field(default="", description="header of the row-label column, e.g. 'Metric'; '' if none")
    continued_from_previous: bool = Field(default=False, description="true if this table is the continuation of a table that started on the previous page (rows continue under the same columns, header NOT repeated)")
    section_id:   str = Field(default="", description="for multiple-section pages only: the section number this table belongs to (e.g. '12.2'); leave '' for single-section pages")
    columns:      list[GColumn]
    rows:         list[GRow]

class Extraction(BaseModel):
    tables: list[GTable]

# ===========================================================================
# GEMINI CONFIG  (structured output, temp 0)
# gemini-2.5-flash: thinking_budget=0 disables thinking (cheap, sufficient for structured tables).
# Pass --thinking to enable thinking_budget=8192 for hard sections.
# ===========================================================================
def build_config(with_thinking: bool) -> types.GenerateContentConfig:
    kwargs = dict(
        response_mime_type="application/json",
        response_schema=Extraction,
        temperature=0.0,
        max_output_tokens=65536,
    )
    budget = 8192 if with_thinking else 0
    try:
        kwargs["thinking_config"] = types.ThinkingConfig(thinking_budget=budget)
    except Exception:
        print("  ⚠️  ThinkingConfig not available in this SDK build — thinking left at default")
    try:
        return types.GenerateContentConfig(**kwargs)
    except TypeError:
        kwargs.pop("thinking_config", None)
        return types.GenerateContentConfig(**kwargs)

# ===========================================================================
# PROMPTS
# ===========================================================================
_PROMPT = """
═══════════════════════════════════════════════════
TABLE STRUCTURE
═══════════════════════════════════════════════════
For each table return:

title
  The printed table title verbatim. Include the reporting date if shown (e.g. "31 Dec 2025").

label_header
  The header of the row-label column (e.g. "$m", "ASF Item"). "" if none is printed.

continued_from_previous
  true ONLY when ALL of these hold:
  1. The columns are identical to the previous table (same count, same leaf labels).
  2. NO new bold heading, date-period label, or section header appears at the top of this chunk.
  3. The first substantive row is a data or total row — NOT a section_header or sub_header.
  If a new bold title, date header, or section_header exists before the first data row,
  set continued_from_previous=false and give this table its own title.

  DIFFERENT DATE PERIODS = DIFFERENT TABLES — always. Two blocks belong to different tables if:
  - A date/period label appears as a header between them, OR
  - Each block ends with a dated total row (e.g. "At 31 December 2025" then "At 31 December 2024")
    followed by a visual break before the next block with the same structure, OR
  - The same column structure repeats for a different reporting date.
  Never merge two date-period blocks into one table. Hard rule, not a judgment call.

columns  (left to right — DATA columns only)
  EXCLUDE the row-label column and the row-number column.
  Two-level headers: set group (spanning label) and leaf (sub-column label).
  Single-level headers: group=null, put the text in leaf.
  Scope/currency lines above headers (e.g. "Group – All Currencies") are NOT columns.
  SUB-LABEL ROW: if a row of descriptive labels sits between the column headers and the
  first data row, those ARE the column leaf values — NOT data rows or sub_header rows.
  EXAMPLE: if the PDF shows:
      Group header:  "Gross carrying amount of¹/"  spanning columns (a) and (b)
      Letter row:    (a)          (b)          (c)      ...
      Label row:     Defaulted    Non-defaulted Allowances ...
      First data row: 3,229       337,891      (3,615)  ...
  Then the correct columns are:
      {"group": "Gross carrying amount of¹/", "leaf": "Defaulted exposures"}
      {"group": "Gross carrying amount of¹/", "leaf": "Non-defaulted exposures"}
      {"group": null, "leaf": "Allowances and impairments"}
  The label row ("Defaulted exposures", "Non-defaulted exposures"...) becomes the leaf.
  The letter row ("(a)", "(b)"...) is discarded — it is just a reference label, not a header.
  NEVER emit the label row as sub_header or data rows.

  LETTER-ROW RULE (hard): If a row of single letters or bracketed letters — "(a)", "(b)",
  "(c)" etc. — appears anywhere in the header band, that entire row is a reference index.
  Discard it unconditionally. The descriptive text row immediately below it provides the
  leaf labels. This applies even when the letter row is the only row between the group
  header and the first data row. NEVER emit "(a)", "(b)" etc. as a leaf value.

rows  (EVERY row, top to bottom)
  row_id    Printed line number exactly as shown ("1", "4a"). null for rows with no number.
  row_type  "section_header" — category title, date/period header, shaded block header
                               (e.g. "31 Dec 2025", "CASH OUTFLOWS", "Loans"). No values.
            "data"           — normal line item.
            "total"          — bold total, grand total, or subtotal.
            "sub_header"     — bold divider introducing a sub-group, no values.
            "note"           — footnote or disclaimer line.
            Use "section_header" for date/period headers — NOT "sub_header".
  level     0 = section_header or grand total
            1 = primary line item
            2 = sub-item (indented, "of which", named breakdown)
            3 = sub-sub-item (rare)
  parent    null for level 0 and 1. For level 2+: row_id of nearest ancestor one level up.
            If that ancestor has no printed number, assign it a synthetic id ("h1","h2",…)
            and use the SAME id in both rows.
  label     Row label verbatim, including footnote markers. Do not re-indent or trim.
  values    One GCell per column (see CELL STATE below).
            section_header / sub_header / note rows use an empty list [].

═══════════════════════════════════════════════════
CATEGORY LABELS — NEVER DROP THEM
═══════════════════════════════════════════════════
Every category / portfolio / asset-class block MUST be captured as a row with
row_type="section_header" (level 0). These labels often appear on the same line as the
column headers — in that case the leading words are the CATEGORY LABEL, not column names.
Emit them as a section_header row. NEVER fold a category label into a column name.

═══════════════════════════════════════════════════
VALUE FIDELITY
═══════════════════════════════════════════════════
- Copy each value EXACTLY as printed, including thousands separators and signs:
  "62,195"  "(1,505)"  "17.0%"  "NM"  ">100"  "unchanged"
- If a number is split by a stray render space ("2 64,680"), join it ("264,680").
- Never invent, merge, split, reorder, or omit any row, column, or value.

═══════════════════════════════════════════════════
CELL STATE  (every GCell must carry one of these 5 states)
═══════════════════════════════════════════════════
  "reported" — any printed value: number, %, text, #, <0.5, NM, etc.
  "nil"      — a dash printed: -, –, —  (zero or negligible).  value = "-"
  "zero"     — printed "0" (explicitly zero).                  value = "0"
  "empty"    — cell is truly blank with no visual mark.         value = ""
  "grey"     — cell is visually shaded / greyed out.            value = ""

Rules — apply in this order:
  1. Any dash (-, –, —)        → cell_state="nil",   value="-"
  2. Printed "0"               → cell_state="zero",  value="0"
  3. Visually grey/shaded blank → cell_state="grey",  value=""
  4. Truly blank, no mark      → cell_state="empty", value=""
  5. Anything else printed     → cell_state="reported", value=verbatim

CRITICAL — DASH PRESERVATION (most common error):
  A printed dash "-" must ALWAYS be captured. NEVER replace a printed dash with "".
  When in doubt between "empty" and "nil": if ANYTHING is visually printed in the cell
  (even a faint or small dash), it is "nil" not "empty".
  Use "empty" ONLY for cells with NO mark whatsoever — completely blank white space.
  WRONG: {"value": "",  "cell_state": "empty"}   ← for a printed dash
  RIGHT: {"value": "-", "cell_state": "nil"}     ← for a printed dash

═══════════════════════════════════════════════════
COLUMN ALIGNMENT  (critical — never shift values)
═══════════════════════════════════════════════════
values must contain EXACTLY one GCell per column, in left-to-right order.
Even if a row is sparse, emit a GCell for every column slot — never skip or shift.

Example — 3 columns [A, B, C], only B has a value:
  [{"value":"","cell_state":"empty"},
   {"value":"42","cell_state":"reported"},
   {"value":"","cell_state":"empty"}]

Trailing empty columns must still be emitted — never truncate the list early."""

def _anchor(sect_num: str, sect: str) -> str:
    """Section boundary rule injected into every prompt type."""
    top_num = sect_num.split(".")[0]
    return (
        f"BOUNDARY RULE (mandatory):\n"
        f"1. Scan top-to-bottom until you find the heading '{sect_num}' or '{sect}'.\n"
        f"2. START extracting tables only AFTER that heading — ignore everything before it.\n"
        f"3. STOP immediately when you see any heading for a DIFFERENT section number "
        f"(e.g. any section that is not {sect_num}) — "
        f"tables after that point do not belong here.\n"
        f"4. If the heading is not found, return {{\"tables\": []}}.\n"
        f"5. If only narrative text exists under the heading (no grid), return {{\"tables\": []}}."
    )


_PROMPT_HASH = hashlib.sha1(_PROMPT.encode()).hexdigest()[:8]


def build_prompt(unit: dict) -> str:
    """Build the extraction prompt for a unit. Unit type determines the lead context:
      single   — one section, one page
      spanning — one section, multiple pages (chunk 1)
      multiple — multiple sections share one page (now unused but kept for legacy)
    """
    pages    = unit["pages"]
    pr       = ", ".join(map(str, pages))
    sect     = unit["leaves"][0]["title"]      if unit.get("leaves") else ""
    sect_num = unit["leaves"][0]["number"]     if unit.get("leaves") else ""

    sep_note = (
        "TABLE SPLITTING RULES:\n"
        "- Every distinct table is a SEPARATE entry. Never merge two tables into one.\n"
        "- A new bold heading followed by its own column headers = NEW table, "
        "even if the column names are identical to the previous table.\n"
        "- Different date periods = different tables. If two blocks of rows are separated "
        "by a visual break and each ends with a dated total (e.g. 'At 31 December 2025' "
        "then 'At 31 December 2024'), they are TWO tables. Name each table's `title` "
        "with the date it belongs to (e.g. 'Credit Quality of Restructured Exposures — "
        "31 December 2025' and 'Credit Quality of Restructured Exposures — 31 December 2024')."
    )

    if unit["type"] == "multiple":
        # Multiple sections share a page — tell Gemini the section order and route by heading
        sections_desc = "; ".join(
            f'{lf["number"]} "{lf["title"]}"'
            for lf in unit["leaves"]
        )
        lead = (
            f"You are given PDF page {pages[0]} from a bank's regulatory disclosure.\n"
            f"This page contains tables from MULTIPLE sections (in reading order): {sections_desc}.\n\n"
            f"Rules:\n"
            f"- Read top-to-bottom. When you encounter a section heading, all tables that follow "
            f"belong to THAT section — until the next section heading appears.\n"
            f"- Do NOT extract tables that appear before the first listed section heading.\n"
            f"- Set `section_id` on each table to the section NUMBER it belongs to.\n"
            f"- Do NOT merge tables that belong to different sections.\n"
            f"- {sep_note}"
        )

    elif unit["type"] == "spanning":
        lead = (
            f"You are given PDF pages {pr} — section {sect_num} '{sect}' "
            f"of a bank's regulatory disclosure.\n\n"
            f"A table that spans a page break (same columns resume, header not repeated) → "
            f"merge into ONE table, set continued_from_previous=true on the continuation.\n"
            f"Genuinely different tables (different title or columns) → SEPARATE entries.\n"
            f"{sep_note}\n\n"
            f"{_anchor(sect_num, sect)}"
        )

    else:  # single
        lead = (
            f"You are given PDF page {pages[0]} — section {sect_num} '{sect}' "
            f"of a bank's regulatory disclosure.\n\n"
            f"{sep_note}\n\n"
            f"{_anchor(sect_num, sect)}"
        )

    return lead + "\n\n" + _PROMPT


def build_continuation_prompt(unit: dict, chunk_pages: list[int],
                               prev_tables: list) -> str:
    """Prompt for chunk 2+ of a spanning section.
    Injects column context from previous chunk so Gemini can stitch continuation rows."""
    pr       = ", ".join(map(str, chunk_pages))
    sect     = unit["leaves"][0]["title"]      if unit.get("leaves") else ""
    sect_num = unit["leaves"][0]["number"]     if unit.get("leaves") else ""

    open_tables_desc = "\n".join(
        f'  - "{t.title or "(untitled)"}": columns [{" | ".join(c.leaf for c in t.columns)}]'
        for t in prev_tables
    )
    context_block = (
        f"═══════════════════════════════════════════════════\n"
        f"CONTEXT FROM PREVIOUS CHUNK\n"
        f"═══════════════════════════════════════════════════\n"
        f"These tables were partially extracted from earlier pages of section {sect_num}.\n"
        f"If this chunk continues any of them (same columns, no new section heading), "
        f"set continued_from_previous=true and emit only the new rows — do NOT repeat headers.\n"
        f"If a genuinely new table starts (different title or columns), create a fresh entry "
        f"with continued_from_previous=false.\n\n"
        f"Open tables:\n{open_tables_desc}"
    )

    lead = (
        f"You are given PDF pages {pr} — continuation of section {sect_num} '{sect}'.\n\n"
        f"{_anchor(sect_num, sect)}\n\n"
        f"{context_block}"
    )
    return lead + "\n\n" + _PROMPT

# ===========================================================================
# PDF HELPERS
# ===========================================================================
def parse_pages(pages_field: str) -> list[int]:
    return [int(p) for p in str(pages_field).replace(" ", "").split("+") if p]

def cut_pdf(pdf_path: str, pages_1based: list[int]) -> bytes:
    """Return a new PDF containing only the given (1-based) pages."""
    with _pdfium_lock:
        src = pdfium.PdfDocument(pdf_path)
        dest = pdfium.PdfDocument.new()
        dest.import_pages(src, [p - 1 for p in pages_1based])
        buf = io.BytesIO()
        dest.save(buf)
    return buf.getvalue()

def detect_bank(pdf_path: str) -> tuple[str | None, str | None]:
    """Scan the first two pages for a bank fingerprint. Returns (key, detected_date)."""
    txt = ""
    try:
        with _pdfium_lock:
            pdf = pdfium.PdfDocument(pdf_path)
            txt = pdf[0].get_textpage().get_text_range()
            if len(pdf) > 1:
                txt += " " + pdf[1].get_textpage().get_text_range()
    except Exception:
        pass
    key = None
    for k, info in BANKS.items():
        if re.search(info["match"], txt, re.I):
            key = k
            break
    m = re.search(r"\b(\d{1,2}\s+[A-Za-z]+\s+\d{4})\b", txt)   # e.g. "31 December 2025"
    return key, (m.group(1) if m else None)

def page_is_narrative(pdf_path: str, page_1based: int, min_numbers: int = 10) -> bool:
    """Cheap deterministic pre-filter: a page with very few numbers is almost
    certainly narrative text (intro / scope / policy) — skip it to avoid an
    empty, billed extraction call. Conservative: only skips clearly text pages."""
    try:
        with _pdfium_lock:
            pdf = pdfium.PdfDocument(pdf_path)
            txt = pdf[page_1based - 1].get_textpage().get_text_range()
    except Exception:
        return False
    return len(re.findall(r"\d[\d,\.]*", txt)) < min_numbers


def page_has_table_structure(pdf_path: str, page_1based: int,
                              min_h_edges: int = 5) -> bool:
    """Return True if pdfplumber detects meaningful horizontal ruling lines on
    the page — the structural signature of a real table. A page with only 2
    h-edges (top/bottom page border) is narrative; a data table has many row
    separators. min_h_edges=5 is conservative: even a 3-row table has 4 lines."""
    if pdfplumber is None:
        return True   # can't tell — don't suppress the retry
    try:
        with _pdfium_lock:
            with pdfplumber.open(pdf_path) as pdf:
                page = pdf.pages[page_1based - 1]
                page_w = page.width
                real_h = [e for e in page.edges
                          if e.get("orientation") == "h"
                          and (e.get("x1", 0) - e.get("x0", 0)) > page_w * 0.10]
                return len(real_h) >= min_h_edges
    except Exception:
        return True   # can't tell — don't suppress the retry

def render_images(pdf_path: str, pages_1based: list[int], scale: float = IMAGE_SCALE) -> list[bytes]:
    """Render the given pages to PNG bytes (used only as a fallback)."""
    with _pdfium_lock:
        src = pdfium.PdfDocument(pdf_path)
        out = []
        for p in pages_1based:
            pil = src[p - 1].render(scale=scale).to_pil()
            buf = io.BytesIO()
            pil.save(buf, format="PNG")
            out.append(buf.getvalue())
    return out

# ===========================================================================
# USAGE LOGGING
# ===========================================================================
def log_usage(resp, label: str, image_used: bool) -> dict:
    try:
        um        = getattr(resp, "usage_metadata", None)
        prompt_t  = getattr(um, "prompt_token_count", None) or 0
        output_t  = getattr(um, "candidates_token_count", None) or 0
        thought_t = getattr(um, "thoughts_token_count", None) or 0
        total_t   = getattr(um, "total_token_count", None) or 0
        cost = (prompt_t / 1e6 * INPUT_PRICE_PER_M) + (output_t / 1e6 * OUTPUT_PRICE_PER_M) + (thought_t / 1e6 * THINK_PRICE_PER_M)
        rec = {
            "ts": datetime.datetime.now().isoformat(timespec="seconds"),
            "script": "extract_to_excel", "label": label, "model": MODEL,
            "image_used": image_used,
            "prompt_tokens": prompt_t, "output_tokens": output_t,
            "thinking_tokens": thought_t, "total_tokens": total_t,
            "est_cost_usd": round(cost, 5),
        }
    except Exception as e:
        prompt_t = output_t = thought_t = 0
        cost = 0.0
        rec = {"ts": datetime.datetime.now().isoformat(timespec="seconds"),
               "script": "extract_to_excel", "label": label, "error": f"usage_capture_failed: {e}"}
    with _run_usage_lock:
        _run_usage["calls"]    += 1
        _run_usage["prompt"]   += prompt_t
        _run_usage["output"]   += output_t
        _run_usage["thinking"] += thought_t
        _run_usage["cost"]     += cost
    with _call_log_lock:
        _call_log.append(rec)
    try:
        os.makedirs(os.path.dirname(USAGE_LOG_PATH) or ".", exist_ok=True)
        with open(USAGE_LOG_PATH, "a") as f:
            f.write(json.dumps(rec) + "\n")
    except Exception:
        pass
    return rec

# ===========================================================================
# PDFPLUMBER COLUMN BOUNDARY EXTRACTION
# ===========================================================================
def get_column_boundaries(pdf_path: str, pages: list[int],
                           min_count: int = 10, tolerance: float = 3.0) -> list[float]:
    """Extract dominant vertical column boundaries from the PDF data region.
    Returns sorted x-coordinates of column dividers (including left and right edges),
    so len(result)-1 == number of data columns. Returns [] on failure."""
    try:
        all_x: dict[float, int] = {}
        with pdfplumber.open(pdf_path) as pdf:
            for pg in pages:
                page = pdf.pages[pg - 1]
                page_w = page.width
                for e in page.edges:
                    if e.get("orientation") != "v":
                        continue
                    x = e["x0"]
                    # exclude far-left margin and right padding
                    if x < page_w * 0.20 or x > page_w * 0.98:
                        continue
                    bucket = round(x / tolerance) * tolerance
                    all_x[bucket] = all_x.get(bucket, 0) + 1

        # keep dominant lines; merge adjacent buckets keeping the heavier one
        candidates = sorted((cnt, x) for x, cnt in all_x.items() if cnt >= min_count)
        candidates = sorted(x for _, x in candidates)
        # merge buckets within 2*tolerance
        merged = []
        for x in candidates:
            if not merged or x - merged[-1] > tolerance * 2:
                merged.append(x)
        return merged
    except Exception:
        return []

def col_boundaries_hint(pdf_path: str, pages: list[int]) -> str:
    """Reserved for future deterministic span pre-computation (see DEVLOG E-07). Not called."""
    return ""

# ===========================================================================
# EXTRACTION
# ===========================================================================
_DASH_CHARS = {"-", "–", "—", "‐", "‑", "‒", "−"}

def _normalise_cell_states(ext: Extraction) -> Extraction:
    """Post-extraction normalisation: fix mis-classified cell states.
    A printed dash must always be nil — schema pressure sometimes causes Gemini
    to emit cell_state='empty' for dashes. Correct it here as defence in depth."""
    for t in ext.tables:
        for row in t.rows:
            for cell in row.values:
                if isinstance(cell, GCell):
                    if cell.value.strip() in _DASH_CHARS:
                        cell.cell_state = "nil"
                        cell.value = "-"
                    elif cell.value.strip() == "0" and cell.cell_state != "zero":
                        cell.cell_state = "zero"
    return ext

_DATE_HEADER_RE = re.compile(
    r'(?:\b\d{1,2}\s+)?(?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*\s+\d{4}\b'
)

def split_date_blocks(t: GTable) -> list[GTable]:
    """Split a table that has >= 2 date-period section_header rows into one
    GTable per period.  If the condition is not met, returns [t] unchanged.
    Pure function — does not mutate the input GTable."""
    # Find all row indices that are date-header rows
    date_header_indices = [
        i for i, r in enumerate(t.rows)
        if r.row_type == "section_header" and _DATE_HEADER_RE.search(r.label)
    ]
    if len(date_header_indices) < 2:
        return [t]

    # Each date-header must be followed by >= 1 row with non-empty values
    # before the next date-header (or end of table).
    boundaries = date_header_indices + [len(t.rows)]
    for k, idx in enumerate(date_header_indices):
        block_start = idx + 1
        block_end   = boundaries[k + 1]
        has_data = any(t.rows[j].values for j in range(block_start, block_end))
        if not has_data:
            return [t]

    # Build one GTable per block.
    # Rows before the first date-header go with the first block.
    result: list[GTable] = []
    pre_rows = list(t.rows[: date_header_indices[0]])

    for k, idx in enumerate(date_header_indices):
        header_row  = t.rows[idx]
        block_start = idx + 1
        block_end   = boundaries[k + 1]
        block_rows  = pre_rows + list(t.rows[block_start:block_end])
        pre_rows    = []   # only prepend to first block

        date_text = _DATE_HEADER_RE.search(header_row.label).group(0)
        orig_title = t.title.strip()
        if orig_title and orig_title != date_text:
            new_title = f"{orig_title} — {date_text}"
        else:
            new_title = date_text

        result.append(t.model_copy(update={"title": new_title, "rows": block_rows}))

    return result


def fill_parents(t: GTable) -> GTable:
    """Overwrite GRow.parent for every row in t using deterministic level-walk.

    Rules:
    - Level 0 and 1 rows: parent = None.
    - Level N (N >= 2) rows: parent = row_id of the nearest preceding level N-1 row.
    - When a level-L row is processed, clear the ancestor stack for all levels > L.
    - Note rows are skipped (parent unchanged; they are not anchors either).
    - If the needed ancestor row has no printed row_id, assign it a synthetic id
      ("h1", "h2", ...) and use that as the parent.  Synthetic ids are only
      assigned to rows that are actually referenced as parents, not to every
      unnumbered row.

    Pure function — returns a new GTable with updated rows; does not mutate input."""
    rows        = [r.model_copy() for r in t.rows]
    # ancestor_stack[level] = (row_index, row_id_or_None)
    ancestor_stack: dict[int, tuple[int, str | None]] = {}
    synthetic_counter = 0

    for i, row in enumerate(rows):
        if row.row_type == "note":
            continue

        level = row.level

        # Clear deeper ancestors — a shallower row invalidates them
        for lvl in list(ancestor_stack.keys()):
            if lvl > level:
                del ancestor_stack[lvl]

        if level <= 1:
            row.parent = None
        else:
            parent_level = level - 1
            if parent_level in ancestor_stack:
                parent_idx, parent_row_id = ancestor_stack[parent_level]
                if parent_row_id is None:
                    # Assign synthetic id to the ancestor row now that it's needed
                    synthetic_counter += 1
                    parent_row_id = f"h{synthetic_counter}"
                    rows[parent_idx] = rows[parent_idx].model_copy(update={"row_id": parent_row_id})
                    ancestor_stack[parent_level] = (parent_idx, parent_row_id)
                row.parent = parent_row_id
            else:
                row.parent = None  # no ancestor at the required level

        # Register this row as the current anchor for its level
        ancestor_stack[level] = (i, row.row_id)

    return t.model_copy(update={"rows": rows})


def _apply_transforms(tables: list[GTable]) -> list[GTable]:
    """Apply split_date_blocks then fill_parents to every table."""
    result: list[GTable] = []
    for t in tables:
        for s in split_date_blocks(t):
            result.append(fill_parents(s))
    return result


def _to_extraction(resp) -> Extraction:
    """Prefer the SDK's parsed pydantic object; fall back to parsing text."""
    parsed = getattr(resp, "parsed", None)
    if isinstance(parsed, Extraction):
        return _normalise_cell_states(parsed)
    raw = (resp.text or "").strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.lstrip().startswith("json"):
            raw = raw.lstrip()[4:]
    data = json.loads(raw)
    return _normalise_cell_states(Extraction(**data))

def _reasonable(ext: Extraction) -> bool:
    """Sanity check on returned tables. Returns False only when tables were returned
    but are structurally malformed (missing columns, rows, or values). Returns True
    for both well-formed tables AND empty results (0 tables = narrative section,
    which is correct and should NOT trigger an image retry)."""
    if not ext.tables:
        return True   # 0 tables is a valid answer for narrative sections
    for t in ext.tables:
        if not t.columns or not t.rows:
            return False
        if not any(r.values for r in t.rows):
            return False
    return True

# ===========================================================================
# VALIDATORS  (zero API cost — pure Python on extracted JSON + PDF text layer)
# ===========================================================================
def validate_spans(ext: Extraction) -> list[str]:
    """Check len(row.values) == len(columns) for every data row.
    Returns list of violation strings; empty list = all good."""
    issues = []
    for t in ext.tables:
        n = len(t.columns)
        for row in t.rows:
            if not row.values:
                continue
            total = len(row.values)
            if total != n:
                issues.append(
                    f"  column count mismatch [{t.title[:40]}] row {row.row_id or repr(row.label[:30])}: "
                    f"got {total} values != ncols={n}"
                )
    return issues

def validate_labels(ext: Extraction) -> list[str]:
    """Detect row-shift corruption: duplicate stripped labels among data/total rows.
    Returns list of issue strings; empty list = all good."""
    issues = []
    for t in ext.tables:
        counts: dict[str, int] = {}
        for row in t.rows:
            if row.row_type not in ("data", "total"):
                continue
            lbl = row.label.strip()
            if lbl:
                counts[lbl] = counts.get(lbl, 0) + 1
        for lbl, n in counts.items():
            if n > 1:
                issues.append(
                    f"  duplicate row label x{n}: '{lbl[:40]}' [{t.title[:30]}]"
                )
    return issues

def _page_raw_text(pdf_path: str, pages: list[int]) -> str:
    """Return concatenated raw text from the given pages via pypdfium2."""
    with _pdfium_lock:
        pdf = pdfium.PdfDocument(pdf_path)
        parts = []
        for pg in pages:
            parts.append(pdf[pg - 1].get_textpage().get_text_range())
    return "\n".join(parts)

def _page_numbers(pdf_path: str, pages: list[int]) -> Counter:
    """Extract numeric tokens from the PDF text layer. Returns Counter of canonical strings."""
    counts: Counter = Counter()
    raw = _page_raw_text(pdf_path, pages)
    for tok in re.findall(r'\(?\d[\d,]*(?:\.\d+)?\)?%?', raw):
        cleaned = tok.strip("()% \n").replace(",", "")
        if cleaned and any(c.isdigit() for c in cleaned):
            counts[cleaned] += 1
    return counts

# Year pattern — suppress year tokens as known noise in deficit checks
_YEAR_RE = re.compile(r'^(19|20)\d{2}$')

def validate_numbers(ext: Extraction, pdf_path: str, pages: list[int],
                     section_ids: tuple = ()) -> list[str]:
    """Calibrated number-recall validator (v2).

    Class A fix — JSON-side: only count pure numeric tokens from GCell values.
      Strips concatenated text (ISINCode:..., Page57to58...) that pdfplumber
      never sees, eliminating phantom issues from text table columns.

    Class B fix — phantom check: before flagging a JSON value as phantom,
      check if it appears anywhere in the page text without spaces/commas.
      Catches kerning-split numbers (4909 → PDF has '4 909') and similar.

    Noise suppression in deficit check:
      - Tokens ≤ 2 chars (row ids, short ints)
      - 4-digit years (appear in headers/footers, not in tables)
      - Section id tokens (e.g. '9.4', '15.1') — correctly not extracted

    Issues sorted by severity: longer numbers and bigger gaps first.
    """
    raw_text     = _page_raw_text(pdf_path, pages)
    text_nospace = re.sub(r'[\s,]', '', raw_text)   # for Class B phantom check
    pdf_counts   = _page_numbers(pdf_path, pages)

    # Class A: only count pure numeric tokens from JSON (no concatenated strings)
    json_counts: Counter = Counter()
    for t in ext.tables:
        for row in t.rows:
            for gcell in row.values:
                raw = gcell.value if isinstance(gcell, GCell) else str(gcell)
                cleaned = re.sub(r'[,()\s%]', '', raw)
                if re.fullmatch(r'\d+(?:\.\d+)?', cleaned):
                    json_counts[cleaned] += 1

    # Known noise: section id tokens passed in by caller.
    # Also add bare numeric suffix of each id (e.g. 'A.5.3' → '5.3', 'A.12.1' → '12.1')
    # so section-number tokens that appear as cell values are suppressed on the deficit side.
    noise = set(section_ids)
    for sid in section_ids:
        # strip leading letter prefix: 'A.14.2.3' → '14.2.3'
        m = re.match(r'^[A-Za-z]\.(.+)$', sid)
        numeric_suffix = m.group(1) if m else sid
        # add all dot-prefixes of the numeric suffix:
        # '14.2.3' → '14.2.3', '14.2', '14'
        parts = numeric_suffix.split('.')
        for i in range(len(parts), 0, -1):
            noise.add('.'.join(parts[:i]))

    # Suppress deficits where json is empty (text table) and token is very long (≥7 digits):
    # these come from currency amounts / ISINs in running text that pdfplumber tokenizes
    # as bare integers but the JSON stored as formatted strings (e.g. 'US$30,000,000,000').
    json_is_empty = not json_counts

    # Build a set of tokens that appear ONLY embedded in larger alphanumeric words in raw_text
    # (never standalone). These are running-text fragments, not table values.
    # A token is "text-only" if every occurrence in raw_text is adjacent to a letter/digit.
    def _is_text_only(token: str) -> bool:
        escaped = re.escape(token)
        standalone = re.search(r'(?<![A-Za-z0-9])' + escaped + r'(?![A-Za-z0-9])', raw_text)
        return standalone is None and token in text_nospace

    issues = []
    # Phantom: in JSON more than in PDF
    for num, cnt in json_counts.items():
        if pdf_counts[num] >= cnt:
            continue
        # Class B: present somewhere in raw text without spaces → kerning/format artefact
        if num in text_nospace:
            continue
        issues.append(("phantom", num, cnt, pdf_counts[num]))

    # Deficit: in PDF more than in JSON
    for num, cnt in pdf_counts.items():
        if len(num) <= 2:
            continue
        if _YEAR_RE.match(num):
            continue
        if num in noise:
            continue
        # Text-table guard: if the unit produced no numeric JSON tokens at all,
        # long integers (≥7 digits) are currency/ISIN fragments from running text.
        if json_is_empty and len(num) >= 7:
            continue
        if json_counts[num] < cnt:
            issues.append(("deficit", num, cnt, json_counts[num]))

    # Sort: longer numbers and bigger gaps first (higher severity)
    issues.sort(key=lambda x: (len(x[1]), abs(x[2] - x[3])), reverse=True)

    return [
        f"  phantom: '{n}' json={j}x pdf={p}x"  if kind == "phantom"
        else f"  deficit: '{n}' pdf={p}x json={j}x"
        for kind, n, j, p in issues
    ]

def extract_unit(client, pdf_path: str, unit: dict, force_image: bool, with_thinking: bool,
                 save_audit: bool = True):
    """Run one Gemini call for a unit, with audit + optional image fallback.
    Returns (Extraction, meta dict)."""
    pages   = unit["pages"]
    prompt  = build_prompt(unit)
    pdf_bytes = cut_pdf(pdf_path, pages)

    if save_audit:
        udir = os.path.join(AUDIT_DIR, unit["unit_id"])
        os.makedirs(udir, exist_ok=True)
        with open(os.path.join(udir, "prompt.txt"), "w") as f:
            f.write(prompt)
        with open(os.path.join(udir, "pages.pdf"), "wb") as f:
            f.write(pdf_bytes)

    pdf_part = types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf")
    config   = build_config(with_thinking)

    def _call(attach_image: bool):
        parts = [pdf_part]
        if attach_image:
            for img in render_images(pdf_path, pages):
                parts.append(types.Part.from_bytes(data=img, mime_type="image/png"))
        parts.append(prompt)
        # retry up to 3 times on transient 503/429 errors with exponential backoff
        last_err = None
        for attempt in range(3):
            try:
                resp = client.models.generate_content(model=MODEL, contents=parts, config=config)
                break
            except Exception as e:
                last_err = e
                msg = str(e)
                if any(code in msg for code in ("503", "429", "UNAVAILABLE", "RESOURCE_EXHAUSTED")):
                    wait = 15 * (2 ** attempt) + random.uniform(0, 5)
                    print(f"      ⏳ {e.__class__.__name__} — waiting {wait}s before retry {attempt+1}/3")
                    time.sleep(wait)
                else:
                    raise
        else:
            raise last_err
        usage = log_usage(resp, unit["unit_id"], attach_image)
        if save_audit:
            with open(os.path.join(udir, "response.txt"), "w") as f:
                f.write(resp.text or "")
        return _to_extraction(resp), usage

    # image only when explicitly forced; otherwise try text-only first and
    # retry with image if the response looks unreasonable — but only when the
    # page actually has tables (enough numbers to be a data page). Narrative
    # pages returning 0 tables are correct; adding an image won't help and
    # just wastes tokens.
    image_first = force_image
    try:
        ext, usage = _call(attach_image=image_first)
    except (json.JSONDecodeError, Exception) as e:
        # truncated / malformed response — retry once with image which often yields a tighter output
        if page_has_table_structure(pdf_path, pages[0]):
            print(f"      ↻ parse error ({e.__class__.__name__}) — retrying {unit['unit_id']} with image")
            ext, usage = _call(attach_image=True)
            image_first = True
        else:
            raise
    image_used = image_first

    if not image_used and not _reasonable(ext) and page_has_table_structure(pdf_path, pages[0]):
        print(f"      ↻ first response looked thin — retrying {unit['unit_id']} with image")
        ext, usage = _call(attach_image=True)
        image_used = True

    # --- validators (zero API cost) ---
    sids = tuple(lf["section_id"] for lf in unit.get("leaves", []))
    span_issues   = validate_spans(ext)
    number_issues = validate_numbers(ext, pdf_path, pages, section_ids=sids)
    label_issues  = validate_labels(ext)

    if span_issues:
        print(f"  ⚠  span violations in {unit['unit_id']}:")
        for s in span_issues:
            print(s)
    if number_issues:
        print(f"  ⚠  number recall issues in {unit['unit_id']} "
              f"({len(number_issues)} discrepancies):")
        for s in number_issues[:5]:   # cap at 5 lines to avoid log spam
            print(s)
        if len(number_issues) > 5:
            print(f"     … and {len(number_issues)-5} more (see meta.json)")
    if label_issues:
        print(f"  ⚠  row-shift / duplicate labels in {unit['unit_id']}:")
        for s in label_issues:
            print(s)

    meta = {
        # Document provenance — used to invalidate stale cache from a different document
        "document":    os.path.basename(pdf_path),
        "bank":        INSTITUTION,
        "doc_date":    DOC_DATE,
        "model":       MODEL,
        "prompt_hash": _PROMPT_HASH,
        # Unit identity
        "unit_id":     unit["unit_id"],
        "section_ids": [lf["section_id"] for lf in unit.get("leaves", [])],
        "section_titles": [lf.get("title", "") for lf in unit.get("leaves", [])],
        "pages":       pages,
        "type":        unit.get("type", "single"),
        # Extraction quality
        "image_used":  image_used,
        "n_tables":    len(ext.tables),
        "n_rows":      sum(len(t.rows) for t in ext.tables),
        "usage":       usage,
        "validation":  {
            "span_issues":   span_issues,
            "number_issues": number_issues,
            "label_issues":  label_issues,
        },
    }
    if save_audit:
        with open(os.path.join(udir, "parsed.json"), "w") as f:
            f.write(ext.model_dump_json(indent=2))
        with open(os.path.join(udir, "meta.json"), "w") as f:
            json.dump(meta, f, indent=2)
    return ext, meta

def extract_unit_chunked(client, pdf_path: str, unit: dict, force_image: bool,
                          with_thinking: bool, save_audit: bool,
                          chunk_size: int = 2) -> tuple:
    """For spanning units longer than chunk_size pages, split into chunks and
    pass column context forward so Gemini never loses track of open tables.
    For everything else (or spanning <= chunk_size), falls through to extract_unit."""
    pages = unit["pages"]
    if unit["type"] != "spanning":
        return extract_unit(client, pdf_path, unit, force_image, with_thinking, save_audit)
    # For sections ≥ 5 pages, chunking hurts: each page is typically a fresh
    # sub-table with no cross-page continuity. Send as one call so Gemini sees
    # the full section context and splits tables correctly.
    effective_chunk = chunk_size if len(pages) < 5 else len(pages)
    if len(pages) <= effective_chunk:
        return extract_unit(client, pdf_path, unit, force_image, with_thinking, save_audit)

    # Split pages into chunks
    chunks = [pages[i:i + effective_chunk] for i in range(0, len(pages), effective_chunk)]
    print(f"     ↷ chunking {len(pages)} pages into {len(chunks)} chunks of ≤{chunk_size}")

    all_tables: list = []
    combined_usage: dict = {}

    for ci, chunk_pages in enumerate(chunks):
        chunk_unit = dict(unit, pages=chunk_pages,
                          unit_id=f"{unit['unit_id']}_c{ci+1}")

        if ci == 0:
            # First chunk: standard spanning prompt
            ext, meta = extract_unit(client, pdf_path, chunk_unit,
                                     force_image, with_thinking, save_audit)
        else:
            # Subsequent chunks: inject column context from tables seen so far
            prompt = build_continuation_prompt(unit, chunk_pages, all_tables)
            pdf_bytes = cut_pdf(pdf_path, chunk_pages)
            pdf_part = types.Part.from_bytes(data=pdf_bytes, mime_type="application/pdf")
            config = build_config(with_thinking)
            parts = [pdf_part]
            parts.append(prompt)
            resp = client.models.generate_content(model=MODEL, contents=parts, config=config)
            usage = log_usage(resp, chunk_unit["unit_id"], image_used=False)
            ext = _to_extraction(resp)
            meta = {"unit_id": chunk_unit["unit_id"], "pages": chunk_pages,
                    "type": "spanning", "image_used": True, "usage": usage}

        # Accumulate usage
        for k, v in meta.get("usage", {}).items():
            combined_usage[k] = combined_usage.get(k, 0) + (v if isinstance(v, (int, float)) else 0)

        pr = "+".join(map(str, chunk_pages))
        ut = meta["usage"]
        print(f"        chunk {ci+1}/{len(chunks)} p{pr}: {len(ext.tables)} table(s)  "
              f"[{ut.get('prompt_tokens','?')}in/{ut.get('output_tokens','?')}out tok]")

        # Merge: continued tables stitch onto the last open table only when ALL hold:
        # same columns, no title, AND first substantive row is not a new section/date header.
        for t in ext.tables:
            first_sub = next((r for r in t.rows if r.row_type not in ("note",)), None)
            is_true_continuation = (
                t.continued_from_previous
                and all_tables
                and len(all_tables[-1].columns) == len(t.columns)
                and not t.title.strip()
                and first_sub is not None
                and first_sub.row_type not in ("section_header", "sub_header")
            )
            if is_true_continuation:
                all_tables[-1].rows.extend(t.rows)
            else:
                all_tables.append(t)

        # Save partial progress after each chunk to parsed.partial.json (NOT
        # parsed.json) so a crash leaves no file that the cache-load path would
        # treat as a complete result.
        if save_audit and all_tables:
            udir = os.path.join(AUDIT_DIR, unit["unit_id"])
            os.makedirs(udir, exist_ok=True)
            partial_ext = Extraction(tables=all_tables)
            with open(os.path.join(udir, "parsed.partial.json"), "w") as f:
                f.write(partial_ext.model_dump_json(indent=2))
            with open(os.path.join(udir, "meta.json"), "w") as f:
                json.dump({"unit_id": unit["unit_id"], "pages": pages,
                           "partial": True, "chunks_completed": ci + 1}, f)

    # Return a combined Extraction object and merged meta
    combined_ext = Extraction(tables=all_tables)

    # Validate combined result across the full page range — this is the path
    # where dropped pages show up (e.g. page-40 table missing from a 7-page section)
    sids = tuple(lf["section_id"] for lf in unit.get("leaves", []))
    span_issues   = validate_spans(combined_ext)
    number_issues = validate_numbers(combined_ext, pdf_path, pages, section_ids=sids)
    label_issues  = validate_labels(combined_ext)
    if span_issues:
        print(f"  ⚠  span violations in {unit['unit_id']} (combined):")
        for s in span_issues:
            print(s)
    if number_issues:
        print(f"  ⚠  number recall in {unit['unit_id']} (combined, {len(number_issues)} issues):")
        for s in number_issues[:5]:
            print(s)
        if len(number_issues) > 5:
            print(f"     … and {len(number_issues)-5} more")
    if label_issues:
        print(f"  ⚠  row-shift / duplicate labels in {unit['unit_id']} (combined):")
        for s in label_issues:
            print(s)

    combined_meta = {
        "unit_id":  unit["unit_id"], "pages": pages, "type": "spanning",
        "image_used": True, "usage": combined_usage,
        "chunks":   len(chunks),
        "document": os.path.basename(pdf_path),
        "bank":     INSTITUTION,
        "doc_date": DOC_DATE,
        "model":    MODEL,
        "prompt_hash": _PROMPT_HASH,
        "section_ids":    [lf["section_id"] for lf in unit.get("leaves", [])],
        "section_titles": [lf.get("title", "") for lf in unit.get("leaves", [])],
        "n_tables": len(all_tables),
        "n_rows":   sum(len(t.rows) for t in all_tables),
        "partial":  False,   # complete — safe to use as resume cache
        "validation": {
            "span_issues":   span_issues,
            "number_issues": number_issues,
            "label_issues":  label_issues,
        },
    }
    # Atomically promote parsed.partial.json → parsed.json only after all chunks
    # succeed. Write to a .tmp first so a kill during the write also leaves no
    # partial parsed.json for the cache-load path to pick up.
    if save_audit:
        udir = os.path.join(AUDIT_DIR, unit["unit_id"])
        os.makedirs(udir, exist_ok=True)
        tmp_path = os.path.join(udir, "parsed.json.tmp")
        with open(tmp_path, "w") as f:
            f.write(combined_ext.model_dump_json(indent=2))
        os.replace(tmp_path, os.path.join(udir, "parsed.json"))
        partial_path = os.path.join(udir, "parsed.partial.json")
        if os.path.exists(partial_path):
            os.remove(partial_path)
        with open(os.path.join(udir, "meta.json"), "w") as f:
            json.dump(combined_meta, f, indent=2)
    return combined_ext, combined_meta

# ===========================================================================
# EXCEL WRITER  (deterministic — matches the chat workbook layout)
# ===========================================================================
def coerce(v):
    """PDF text -> typed cell value. Numbers become numbers (commas stripped);
    dashes, percentages, parentheses-negatives and other text stay as-is."""
    if v is None:
        return None
    s = str(v).strip()
    if s in ("", "-", "–", "—", "n.m.", "nm", "NA", "N/A"):
        return s
    if s.endswith("%"):
        return s
    t = s.replace(",", "")
    neg = t.startswith("(") and t.endswith(")")
    core = t[1:-1] if neg else t
    try:
        num = float(core)
    except ValueError:
        return s
    if neg:
        num = -num
    return int(num) if num == int(num) else num

def _hdr_style(cell, meta=False):
    cell.fill = PatternFill("solid", fgColor=DARK_GREY if meta else HEADER_FILL)
    cell.font = Font(bold=True, color=WHITE, size=10)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

def write_table(ws, start_row: int, t: GTable) -> int:
    """Write one table (headers + data) starting at start_row.
    Returns the next free row. Each table has its own tab so no in-tab title banner."""
    nbase = N_META
    cols = t.columns
    ncols = nbase + len(cols)
    r = start_row

    base_headers = META_HEADERS[:-1] + [t.label_header or META_HEADERS[-1]]
    has_group = any(c.group for c in cols)

    if has_group:
        group_row, leaf_row = r, r + 1
        j, col = 0, nbase + 1
        while j < len(cols):
            g = cols[j].group or ""
            k = j
            while k + 1 < len(cols) and (cols[k + 1].group or "") == g:
                k += 1
            if g:
                gc = ws.cell(group_row, col, g)
                _hdr_style(gc)
                if k > j:
                    ws.merge_cells(start_row=group_row, start_column=col,
                                   end_row=group_row, end_column=col + (k - j))
            col += (k - j + 1)
            j = k + 1
        for ci, h in enumerate(base_headers):
            _hdr_style(ws.cell(leaf_row, 1 + ci, h), meta=(ci < 3))
        for ci, c2 in enumerate(cols):
            _hdr_style(ws.cell(leaf_row, nbase + 1 + ci, c2.leaf))
        r = leaf_row + 1
    else:
        for ci, h in enumerate(base_headers):
            _hdr_style(ws.cell(r, 1 + ci, h), meta=(ci < 3))
        for ci, c2 in enumerate(cols):
            _hdr_style(ws.cell(r, nbase + 1 + ci, c2.leaf))
        r += 1

    meta_font = Font(color=MID_GREY, size=9)
    for row in t.rows:
        is_header = row.row_type == "section_header"
        is_total  = row.row_type == "total"
        is_note   = row.row_type == "note"

        ws.cell(r, 1, row.row_id)
        ws.cell(r, 2, row.level)
        ws.cell(r, 3, row.parent or "")
        indent = "    " * max(0, row.level - 1) if not is_header else ""
        ws.cell(r, 4, indent + row.label)
        col_cursor = nbase + 1
        for gcell in row.values:
            # Support legacy plain-string values from old cached parsed.json
            if isinstance(gcell, str):
                gcell = GCell.from_str(gcell)
            cell = ws.cell(r, col_cursor, coerce(gcell.value))
            state = gcell.cell_state
            if state == "grey":
                cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
            elif state == "nil":
                cell.value = "-"
                cell.alignment = Alignment(horizontal="center")
            elif state == "zero":
                cell.value = 0
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")
            elif state == "reported" and isinstance(cell.value, (int, float)):
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")
            # empty / reported-string: default left alignment, no fill
            col_cursor += 1

        # row-type styling
        for ci in range(1, 4):
            ws.cell(r, ci).font = meta_font
            ws.cell(r, ci).alignment = Alignment(horizontal="center")
        if is_header:
            for ci in range(1, ncols + 1):
                cell = ws.cell(r, ci)
                cell.fill = PatternFill("solid", fgColor=DARK_GREY)
                cell.font = Font(bold=True, color=WHITE, size=10)
        elif is_total:
            for ci in range(1, ncols + 1):
                cur = ws.cell(r, ci).font
                ws.cell(r, ci).font = Font(bold=True, color=cur.color, size=cur.size or 10)
        elif is_note:
            for ci in range(1, ncols + 1):
                ws.cell(r, ci).font = Font(italic=True, color=MID_GREY, size=8)
        r += 1
    return r + 1  # one blank spacer row between stacked tables

def write_cost_sheet(wb, call_log: list[dict], run_usage: dict, out_path: str):
    """Write (or replace) a 'Cost' tab summarising every API call made this run."""
    if "Cost" in wb.sheetnames:
        wb.remove(wb["Cost"])
    ws = wb.create_sheet("Cost")

    # Title banner
    ws.merge_cells("A1:I1")
    c = ws.cell(1, 1, f"{INSTITUTION}  |  API Cost Log  |  {DOC_DATE}")
    c.fill = PatternFill("solid", fgColor=BRAND_COLOUR)
    c.font = Font(bold=True, color=WHITE, size=12)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 22

    headers = ["#", "Timestamp", "Label", "Model", "Image?",
               "Input tok", "Output tok", "Think tok", "Est. cost (USD)"]
    for ci, h in enumerate(headers, 1):
        _hdr_style(ws.cell(2, ci, h))

    total_cost = 0.0
    for ri, rec in enumerate(call_log, start=1):
        cost = rec.get("est_cost_usd", 0) or 0
        total_cost += cost
        ws.cell(ri + 2, 1, ri).alignment = Alignment(horizontal="center")
        ws.cell(ri + 2, 2, rec.get("ts", ""))
        ws.cell(ri + 2, 3, rec.get("label", ""))
        ws.cell(ri + 2, 4, rec.get("model", MODEL))
        ws.cell(ri + 2, 5, "yes" if rec.get("image_used") else "no").alignment = Alignment(horizontal="center")
        for ci, key in enumerate(["prompt_tokens", "output_tokens", "thinking_tokens"], start=6):
            cell = ws.cell(ri + 2, ci, rec.get(key, 0))
            cell.number_format = "#,##0"
            cell.alignment = Alignment(horizontal="right")
        cell = ws.cell(ri + 2, 9, round(cost, 5))
        cell.number_format = "$#,##0.00000"
        cell.alignment = Alignment(horizontal="right")

    # Summary row
    sr = len(call_log) + 4
    ws.merge_cells(start_row=sr, start_column=1, end_row=sr, end_column=5)
    c = ws.cell(sr, 1, f"TOTAL  ({run_usage['calls']} calls)")
    c.font = Font(bold=True, size=10)
    for ci, key in enumerate(["prompt", "output", "thinking"], start=6):
        cell = ws.cell(sr, ci, run_usage.get(key, 0))
        cell.number_format = "#,##0"
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="right")
    cell = ws.cell(sr, 9, round(run_usage.get("cost", 0), 5))
    cell.number_format = "$#,##0.00000"
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="right")

    # Note row: pricing basis
    ws.merge_cells(start_row=sr + 1, start_column=1, end_row=sr + 1, end_column=9)
    note = ws.cell(sr + 1, 1,
        f"Pricing: ${INPUT_PRICE_PER_M}/M input, ${OUTPUT_PRICE_PER_M}/M output, ${THINK_PRICE_PER_M}/M thinking  "
        f"|  Model: {MODEL}  |  Log: {os.path.basename(USAGE_LOG_PATH)}")
    note.font = Font(italic=True, color=MID_GREY, size=8)

    for col, w in {1:4, 2:20, 3:28, 4:18, 5:8, 6:12, 7:12, 8:12, 9:16}.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def save_cost_summary(call_log: list[dict], run_usage: dict, out_path: str):
    """Write a JSON cost summary next to the Excel file."""
    summary_path = os.path.splitext(out_path)[0] + "_cost_summary.json"
    summary = {
        "generated_at": datetime.datetime.now().isoformat(timespec="seconds"),
        "output_file": os.path.basename(out_path),
        "model": MODEL,
        "pricing": {"input_per_million": INPUT_PRICE_PER_M,
                    "output_per_million": OUTPUT_PRICE_PER_M},
        "totals": {
            "calls": run_usage["calls"],
            "input_tokens": run_usage["prompt"],
            "output_tokens": run_usage["output"],
            "thinking_tokens": run_usage["thinking"],
            "total_tokens": run_usage["prompt"] + run_usage["output"] + run_usage["thinking"],
            "est_cost_usd": round(run_usage["cost"], 5),
        },
        "calls": call_log,
    }
    with open(summary_path, "w") as f:
        json.dump(summary, f, indent=2)
    print(f"   💰 cost summary → {summary_path}")


def _describe_call(label: str, image_used: bool, bank: str) -> str:
    """One-line human description of what a call was doing."""
    parts = label.split("_")
    # label formats: "15_3_p47-48", "p7_multi", "15_3_p47-48_c2"
    is_chunk = label.endswith(("_c1", "_c2", "_c3", "_c4", "_c5"))
    is_multi = "multi" in label
    page_tok = next((p for p in parts if p.startswith("p") and any(c.isdigit() for c in p)), "")
    section_parts = [p for p in parts if p not in (page_tok.lstrip("p"),) and not p.startswith("p")]
    section_id = ".".join(section_parts[:-1]) if is_chunk else ".".join(section_parts)
    section_id = section_id.strip("._")
    desc = f"{bank} §{section_id} {page_tok}"
    if is_multi:
        desc += " (shared page — multiple sections)"
    if is_chunk:
        chunk_n = label.rsplit("_c", 1)[-1]
        desc += f" chunk {chunk_n}"
    if image_used:
        desc += " +image"
    return desc.strip()


_API_LOG_HEADERS = [
    "#", "Run date", "Bank", "Section label", "Description",
    "Model", "Pages", "Image?",
    "Input tok", "Output tok", "Think tok", "Total tok",
    "Est. cost (USD)", "Cumulative cost (USD)",
]

def append_to_api_log(call_log: list[dict], bank: str, out_dir: str):
    """Append this run's calls to the shared API_Log.xlsx in the outputs folder."""
    log_path = os.path.join(out_dir, "API_Log.xlsx")
    if os.path.exists(log_path):
        wb = openpyxl.load_workbook(log_path)
        ws = wb.active
        # find the last row with data and the current cumulative cost
        last_row = ws.max_row
        cum_cost = 0.0
        for r in range(2, last_row + 1):
            v = ws.cell(r, 14).value
            if isinstance(v, (int, float)):
                cum_cost = v
        next_row = last_row + 1
        next_num = last_row   # row 1 = header, so call# = row-1
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "API Calls"
        # Header row
        for ci, h in enumerate(_API_LOG_HEADERS, 1):
            c = ws.cell(1, ci, h)
            c.fill = PatternFill("solid", fgColor=HEADER_FILL)
            c.font = Font(bold=True, color=WHITE, size=10)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28
        next_row, next_num, cum_cost = 2, 1, 0.0

    run_date = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")
    for rec in call_log:
        label     = rec.get("label", "")
        image_used = rec.get("image_used", False)
        prompt_t  = rec.get("prompt_tokens", 0) or 0
        output_t  = rec.get("output_tokens", 0) or 0
        think_t   = rec.get("thinking_tokens", 0) or 0
        cost      = rec.get("est_cost_usd", 0) or 0
        cum_cost += cost
        desc = _describe_call(label, image_used, bank)
        # extract page hint from label
        page_hint = next((p for p in label.split("_") if p.startswith("p") and any(c.isdigit() for c in p)), "")

        row_vals = [
            next_num, run_date, bank, label, desc,
            rec.get("model", MODEL), page_hint,
            "yes" if image_used else "no",
            prompt_t, output_t, think_t, prompt_t + output_t + think_t,
            round(cost, 5), round(cum_cost, 5),
        ]
        for ci, v in enumerate(row_vals, 1):
            cell = ws.cell(next_row, ci, v)
            if ci in (9, 10, 11, 12):
                cell.number_format = "#,##0"
                cell.alignment = Alignment(horizontal="right")
            elif ci in (13, 14):
                cell.number_format = "$#,##0.00000"
                cell.alignment = Alignment(horizontal="right")
            elif ci == 8:
                cell.alignment = Alignment(horizontal="center")
        next_row += 1
        next_num += 1

    # column widths
    for col, w in {1:5, 2:18, 3:8, 4:26, 5:48, 6:20, 7:12, 8:8,
                   9:13, 10:13, 11:13, 12:13, 13:16, 14:18}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    wb.save(log_path)
    print(f"   📋 API log updated → {log_path}  (total log rows: {next_row - 2})")


def write_section_header(ws, section_id: str, title: str, last_col: int,
                         table_label: str = "", table_n: int = 0, total_tables: int = 0):
    """Row 1 = brand banner with section title; row 2 = table identifier; row 3 = source line."""
    last_col = max(last_col, N_META)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    c = ws.cell(1, 1, f"{INSTITUTION}  |  Section {section_id}: {title}")
    c.fill = PatternFill("solid", fgColor=BRAND_COLOUR)
    c.font = Font(bold=True, color=WHITE, size=12)
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=last_col)
    ref = f"[{section_id} Table {table_n}]" if table_n else f"[{section_id}]"
    count_str = f"  ({table_n} of {total_tables})" if total_tables > 1 else ""
    label_str = f"  —  {table_label}" if table_label else ""
    c = ws.cell(2, 1, f"{ref}{count_str}{label_str}")
    c.font = Font(bold=True, color=BRAND_COLOUR, size=10)
    ws.row_dimensions[2].height = 16

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=last_col)
    c = ws.cell(3, 1, f"Source: {DOC_TITLE}, {DOC_DATE}  |  Units: S$ millions unless noted")
    c.font = Font(italic=True, color=MID_GREY, size=9)
    ws.row_dimensions[3].height = 14

def style_sheet_columns(ws):
    widths = {1: 13, 2: 14, 3: 14, 4: 58}
    for i in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(i, 15)

def sheet_name(used: set, section_id: str, title: str) -> str:
    base = f"{section_id} - {title}".strip(" -")
    for ch in '[]:*?/\\':
        base = base.replace(ch, " ")
    base = " ".join(base.split())[:31] or section_id[:31] or "Sheet"
    name, i = base, 2
    while name in used:
        suffix = f" ({i})"
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name

def table_sheet_name(used: set, section_id: str, table_n: int) -> str:
    """One tab per table: '18.4 Table 1', '18.4 Table 2', etc."""
    base = f"{section_id} Table {table_n}"
    for ch in '[]:*?/\\':
        base = base.replace(ch, " ")
    base = " ".join(base.split())[:31] or f"t{table_n}"
    name, i = base, 2
    while name in used:
        suffix = f" ({i})"
        name = base[:31 - len(suffix)] + suffix
        i += 1
    used.add(name)
    return name

# ===========================================================================
# CONTENTS (TABLE OF CONTENTS) SHEET
# ===========================================================================
def load_index() -> list[dict]:
    if os.path.exists(INDEX_PATH):
        try:
            idx = json.load(open(INDEX_PATH))
            # Normalise legacy entries that used "first_tab" instead of "sheet"
            for e in idx:
                if "sheet" not in e and "first_tab" in e:
                    e["sheet"] = e["first_tab"]
            return idx
        except Exception:
            return []
    return []

def save_index(idx: list[dict]):
    os.makedirs(os.path.dirname(INDEX_PATH) or ".", exist_ok=True)
    json.dump(idx, open(INDEX_PATH, "w"), indent=2)

def update_index(idx: list[dict], entry: dict) -> list[dict]:
    idx = [e for e in idx if e["section_id"] != entry["section_id"]]
    idx.append(entry)
    return idx

def rebuild_contents(wb, idx: list[dict]):
    """(Re)build the Contents sheet as the first tab, hyperlinked to each section."""
    if "Contents" in wb.sheetnames:
        wb.remove(wb["Contents"])
    ws = wb.create_sheet("Contents", 0)   # index 0 = first sheet

    ws.merge_cells("A1:E1")
    c = ws.cell(1, 1, INSTITUTION)
    c.fill = PatternFill("solid", fgColor=BRAND_COLOUR)
    c.font = Font(bold=True, color=WHITE, size=14)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:E2")
    c = ws.cell(2, 1, f"{DOC_TITLE} — {DOC_DATE}")
    c.font = Font(italic=True, size=11)
    c.alignment = Alignment(horizontal="center")

    headers = ["Section", "Title", "Pages", "Tables", "Sheet"]
    for ci, h in enumerate(headers, start=1):
        _hdr_style(ws.cell(4, ci, h))

    r = 5
    for e in sorted(idx, key=lambda x: x.get("first_page", 0)):
        ws.cell(r, 1, e["section_id"]).alignment = Alignment(horizontal="center")
        ws.cell(r, 2, e["title"])
        ws.cell(r, 3, e.get("pages", "")).alignment = Alignment(horizontal="center")
        ws.cell(r, 4, e.get("n_tables", "")).alignment = Alignment(horizontal="center")
        sname = e.get("sheet") or e.get("first_tab", "")
        link = ws.cell(r, 5, sname)
        if sname and sname in wb.sheetnames:
            link.hyperlink = f"#'{sname}'!A1"
        link.font = Font(color="0000CC", underline="single")
        r += 1

    for col, w in {1: 12, 2: 60, 3: 16, 4: 10, 5: 32}.items():
        ws.column_dimensions[get_column_letter(col)].width = w
    return ws

# ===========================================================================
# UNIT GROUPING  (table shape derived from leaf<->page mapping; no API)
# ===========================================================================
def _contig(pages: list[int]) -> list[list[int]]:
    """Split a sorted page list into contiguous runs: [3,4,5,8] -> [[3,4,5],[8]]."""
    runs, cur = [], []
    for p in sorted(pages):
        if cur and p == cur[-1] + 1:
            cur.append(p)
        else:
            if cur:
                runs.append(cur)
            cur = [p]
    if cur:
        runs.append(cur)
    return runs

def group_key(section: dict) -> str:
    """Top-level grouping key for a leaf — part-aware so DBS 'A.12.x' and 'B.1.x'
    don't collide. e.g. A.12.2.5 -> 'A.12'; B.1.1 -> 'B.1'; OCBC 18.4 -> '18'."""
    num0 = str(section["number"]).split(".")[0]
    part = section.get("part")
    return f"{part}.{num0}" if part else num0

def build_units(leaves: list[dict]) -> list[dict]:
    """One unit per leaf section. Pages = full page range of that section.
    If two sections share a page, that page is sent in both calls — simpler
    and more correct than shared-page routing which caused cross-group bugs.
    """
    units: list[dict] = []
    for i, s in enumerate(leaves):
        pages = list(range(int(s["start_page"]), int(s["end_page"]) + 1))
        typ = "single" if len(pages) == 1 else "spanning"
        sid_slug = s["section_id"].replace(".", "_")
        p_str = str(pages[0]) + (f"-{pages[-1]}" if len(pages) > 1 else "")
        uid = f"{sid_slug}_p{p_str}"
        units.append({
            "type":     typ,
            "pages":    pages,
            "leaves":   [s],
            "unit_id":  uid,
            "group":    group_key(s),
            "next_leaf": leaves[i + 1] if i + 1 < len(leaves) else None,
        })
    units.sort(key=lambda u: u["pages"][0])
    return units

def load_sections() -> tuple[dict, list[dict]]:
    """Load the deterministic TOC produced by build_toc.py.
    Returns (document_meta, [leaf-section, ...]) in document order."""
    if not os.path.exists(TOC_PATH):
        sys.exit(f"{TOC_PATH} not found — run:  python build_toc.py <pdf>")
    toc = json.load(open(TOC_PATH))
    # Build part ordering from actual parts in the document, alphabetically.
    # None (no part) sorts first alongside "A".
    all_parts = sorted({s.get("part") for s in toc.get("sections", []) if s.get("part")})
    _PART_ORD = {None: 0}
    for i, p in enumerate(all_parts):
        _PART_ORD[p] = i
    secs = sorted(toc.get("sections", []),
                  key=lambda s: (int(s["start_page"]),
                                 _PART_ORD.get(s.get("part"), 0),
                                 [int(x) for x in s["number"].split(".")]))
    return toc.get("document", {}), secs

def _norm_words(s: str) -> set:
    return set(re.findall(r"[a-z0-9]+", (s or "").lower()))

def _title_score(table_title: str, leaf_title: str) -> float:
    """Token-overlap fraction of the leaf title covered by the table title."""
    tt, lt = _norm_words(table_title), _norm_words(leaf_title)
    return len(tt & lt) / (len(lt) or 1)

def drop_next_section_tables(tables: list, unit: dict) -> list:
    nl = unit.get("next_leaf")
    if not nl or int(nl["start_page"]) != unit["pages"][-1]:
        return tables
    own = unit["leaves"][0]["title"]
    kept = []
    for t in tables:
        if _title_score(t.title, nl["title"]) > _title_score(t.title, own):
            print(f"   ✂ dropped '{t.title[:40]}' — belongs to next "
                  f"section {nl['section_id']}")
        else:
            kept.append(t)
    return kept

def route_tables(tables: list, leaves: list[dict]) -> list[tuple]:
    """Assign each table on a shared page to one subsection leaf.

    Strategy (in priority order):
      1. section_id tag — Gemini explicitly tagged the table with a section number;
         use it directly. This is the primary mechanism for multiple-section pages.
      2. Title match fallback — for tables without a tag, match by title similarity.
      3. Overflow fallback — still-unmatched tables go to the last matched leaf in
         reading order (handles multiple tables under one section heading).

    Returns [(table, leaf, method, score, flagged)] in the input table order.
    """
    leaves_ord = sorted(leaves, key=lambda lf: [int(x) for x in lf["number"].split(".")])
    leaf_by_num = {lf["number"]: lf for lf in leaves_ord}

    chosen: dict[int, tuple] = {}

    # 1) section_id tag — Gemini's explicit top-to-bottom section assignment
    for ti, t in enumerate(tables):
        sid = (t.section_id or "").strip()
        if sid and sid in leaf_by_num:
            chosen[ti] = (leaf_by_num[sid], "section_id", 1.0)

    # 2) title match for any remaining untagged tables
    untagged = [ti for ti in range(len(tables)) if ti not in chosen]
    if untagged:
        taken_leaves = {chosen[ti][0]["section_id"] for ti in chosen}
        pairs = sorted(
            ((_title_score(tables[ti].title, lf["title"]), ti, lf["section_id"], lf)
             for ti in untagged for lf in leaves_ord),
            reverse=True, key=lambda x: x[0],
        )
        taken_title = set()
        for sc, ti, lid, lf in pairs:
            if sc <= 0 or ti in taken_title or (lid in taken_leaves and sc < 0.5):
                continue
            chosen[ti] = (lf, "title", sc)
            taken_title.add(ti)
            taken_leaves.add(lid)

    # 3) overflow: still-unmatched go to last matched leaf in reading order
    last_leaf = None
    free = [lf for lf in leaves_ord
            if lf["section_id"] not in {chosen[ti][0]["section_id"] for ti in chosen}]
    fi = 0
    for ti in range(len(tables)):
        if ti in chosen:
            last_leaf = chosen[ti][0]
        else:
            if last_leaf is not None:
                chosen[ti] = (last_leaf, "overflow", 0.0)
            else:
                lf = free[fi] if fi < len(free) else leaves_ord[-1]
                chosen[ti] = (lf, "order", 0.0)
                fi += 1

    out = []
    count_mismatch = len(tables) != len(leaves_ord)
    for ti, t in enumerate(tables):
        lf, method, sc = chosen[ti]
        flagged = method in ("order", "overflow") or sc < 0.34 or count_mismatch
        out.append((t, lf, method, sc, flagged))
    return out

# ===========================================================================
# MAIN
# ===========================================================================
def main():
    ap = argparse.ArgumentParser(description="Simplified PDF -> Excel extraction (one tab per section)")
    ap.add_argument("pdf")
    ap.add_argument("--toc", default=None, help="path to TOC JSON (default: out/<bank>_toc.json or out/step1_toc.json)")
    ap.add_argument("--section", help="only this section_id")
    ap.add_argument("--start-section", help="begin at this section_id (document order)")
    ap.add_argument("--out", default=None)
    ap.add_argument("--no-pause", action="store_true", help="do not pause after each section")
    ap.add_argument("--image", action="store_true", help="always attach a rendered image alongside the PDF")
    ap.add_argument("--thinking", action="store_true", help="enable model thinking (higher cost)")
    ap.add_argument("--force", action="store_true", help="re-extract sections whose tab already exists")
    ap.add_argument("--list", action="store_true", help="list sections and exit")
    ap.add_argument("--no-audit", action="store_true",
                    help="skip writing audit files (prompt.txt, pages.pdf, response.txt, parsed.json, meta.json)")
    ap.add_argument("--dry-run", action="store_true",
                    help="print the per-section call plan (units/prompts/tabs) and exit — no API, no key")
    ap.add_argument("--bank", choices=list(BANKS), help="force the institution/brand (else auto-detected)")
    ap.add_argument("--institution", help="override the banner institution name")
    ap.add_argument("--brand", help="override the brand colour hex (e.g. 1B6EC2)")
    ap.add_argument("--doc-date", help="override the source-line date (e.g. '31 December 2025')")
    ap.add_argument("--chunk-pages", type=int, default=2, metavar="N",
                    help="max pages per Gemini call for spanning sections (default 2; use 0 to disable chunking)")
    ap.add_argument("--workers", type=int, default=5, metavar="N",
                    help="max concurrent section groups (default 5; use 1 for sequential)")
    args = ap.parse_args()

    if not os.path.exists(args.pdf):
        sys.exit(f"PDF not found: {args.pdf}")

    # set per-bank identity + brand (auto-detect, then apply any explicit overrides)
    global INSTITUTION, BRAND_COLOUR, DOC_DATE
    detected, det_date = detect_bank(args.pdf)
    bank = args.bank or detected

    # auto-derive TOC path and output path from bank name if not specified
    global TOC_PATH
    bank_slug = (bank or "dbs").lower()
    _p3_out = Path(__file__).parent.parent / "outputs" / "pillar3"
    if args.toc:
        TOC_PATH = args.toc
    elif (_p3_out / f"{bank_slug}_toc.json").exists():
        TOC_PATH = str(_p3_out / f"{bank_slug}_toc.json")
    # else keep default out/step1_toc.json

    out_path = args.out or str(_p3_out / f"{bank_slug}_pillar3.xlsx")

    # one Contents index per output workbook (so different banks don't mix)
    global INDEX_PATH, USAGE_LOG_PATH, COST_LOG_PATH, AUDIT_DIR
    INDEX_PATH     = os.path.splitext(out_path)[0] + ".index.json"
    out_dir        = os.path.dirname(os.path.abspath(out_path))
    USAGE_LOG_PATH = os.path.join(out_dir, f"{bank_slug}_api_usage.jsonl")
    COST_LOG_PATH  = os.path.join(out_dir, f"{bank_slug}_cost_summary.json")
    # Folder hierarchy: audit/{bank}/{doc_stem}/  e.g. audit/dbs/DBS_1Q26_Pillar3/
    doc_stem  = Path(args.pdf).stem   # "DBS_1Q26_Pillar3" from "DBS_1Q26_Pillar3.pdf"
    AUDIT_DIR = os.path.join(out_dir, "audit", bank_slug, doc_stem)
    if bank:
        INSTITUTION  = BANKS[bank]["institution"]
        BRAND_COLOUR = BANKS[bank]["brand"]
    if det_date:
        DOC_DATE = det_date
    if args.institution:
        INSTITUTION = args.institution
    if args.brand:
        BRAND_COLOUR = args.brand
    if args.doc_date:
        DOC_DATE = args.doc_date
    print(f"🏦 Institution: {INSTITUTION}  |  brand #{BRAND_COLOUR}  |  date: {DOC_DATE}"
          f"  ({'detected ' + bank if bank and not args.bank else ('--bank ' + bank if args.bank else 'default')})")

    document, sections = load_sections()
    if args.list:
        print(f"{document.get('title','')[:60]} — {len(sections)} sections (document order):")
        for s in sections:
            print(f"  {s['section_id']:<8} p{s['start_page']}-{s['end_page']:<3} "
                  f"[{s.get('page_label') or '—'}]  {s['title'][:50]}")
        return

    # resume: load an existing workbook so completed tabs are kept
    if os.path.exists(out_path):
        wb = openpyxl.load_workbook(out_path)
    else:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
    used_names = set(wb.sheetnames)
    idx = load_index()

    # Build typed units for the whole document, then group them by top-level
    # section number (so we can review + pause one section at a time).
    units = build_units(sections)
    groups: dict[str, list[dict]] = {}
    for u in units:
        groups.setdefault(u["group"], []).append(u)

    # A 'multiple' unit whose leaves span more than one top-level group must
    # appear in every group it contributes to — otherwise the group that owns
    # the unit runs it, but sibling groups never see the tables routed to them.
    # e.g. A_3+A_4_p5 is assigned group=A.3 but also has a leaf in A.4.

    # --dry-run: print the call plan and exit (no client, no API, no key needed)
    if args.dry_run:
        sel = [(g, us) for g, us in groups.items()
               if not args.section or args.section == g
               or any(args.section == lf["section_id"] for u in us for lf in u["leaves"])]
        total = sum(len(us) for _, us in sel)
        print(f"DRY RUN — {len(sel)} section group(s), {total} Gemini call(s) planned (no API used):")
        for gnum, gunits in sel:
            print(f"\n##### Section {gnum}")
            for u in gunits:
                pr = "+".join(map(str, u["pages"]))
                tabs = ", ".join(lf["section_id"] for lf in u["leaves"])
                print(f"   [{u['type']:8}] pages {pr:<8} -> PROMPT_{u['type'].upper():8} -> tab(s): {tabs}")
        return

    client = genai.Client()

    def _tab_exists(sid: str) -> bool:
        # per-table tabs are named "{sid} Table N" — any such tab counts as done
        return any(nm == sid or nm.startswith(f"{sid} -") or nm.startswith(f"{sid} Table ") for nm in wb.sheetnames)

    # ── Per-group extraction function (runs in thread pool) ──────────────────
    # Returns {section_id: [GTable, ...]} for one group.
    # Units within a group run sequentially (chunked sections have ordering deps).
    def _extract_group(gnum: str, gunits: list, group_leaves: list,
                       leaf_target: str | None) -> dict[str, list]:
        grp_tables: dict[str, list] = defaultdict(list)
        for u in gunits:
            if leaf_target and not any(lf["section_id"] == leaf_target for lf in u["leaves"]):
                continue
            pr = "+".join(map(str, u["pages"]))
            leaf_ids = ", ".join(lf["section_id"] for lf in u["leaves"])
            if u["type"] in ("single", "multiple") and (
                    page_is_narrative(args.pdf, u["pages"][0]) or
                    not page_has_table_structure(args.pdf, u["pages"][0])):
                print(f"   • [{u['type']:8}] p{pr:<7} {leaf_ids}  — narrative, skipped (no call)")
                continue
            audit_exists = os.path.exists(os.path.join(AUDIT_DIR, u["unit_id"], "parsed.json"))
            if not args.force and not args.no_audit and audit_exists:
                try:
                    saved_meta_path = os.path.join(AUDIT_DIR, u["unit_id"], "meta.json")
                    if os.path.exists(saved_meta_path):
                        saved_meta = json.load(open(saved_meta_path))
                        # Reject partial cache (crashed mid-chunk run)
                        if saved_meta.get("partial"):
                            print(f"   • [{u['type']:8}] p{pr:<7} {leaf_ids}  — partial cache, re-extracting")
                            audit_exists = False
                            raise StopIteration
                        cached_doc = saved_meta.get("document", "")
                        current_doc = os.path.basename(args.pdf)
                        if cached_doc and cached_doc != current_doc:
                            print(f"   • [{u['type']:8}] p{pr:<7} {leaf_ids}  — cache mismatch "
                                  f"(cached={cached_doc} current={current_doc}), re-extracting")
                            audit_exists = False
                            raise StopIteration  # skip to live call
                        # Invalidate if the cached page list doesn't match the unit's pages —
                        # catches partial-chunk caches that survived a crash before the fix.
                        cached_pages_meta = saved_meta.get("pages")
                        if cached_pages_meta is None or cached_pages_meta != u["pages"]:
                            print(f"   • [{u['type']:8}] p{pr:<7} {leaf_ids}  — cache mismatch "
                                  f"(cached pages={cached_pages_meta} current={u['pages']}), re-extracting")
                            audit_exists = False
                            raise StopIteration
                        # Invalidate if the prompt has changed since the cache was written,
                        # or if the cache predates prompt_hash tracking (legacy cache).
                        cached_hash = saved_meta.get("prompt_hash")
                        if cached_hash is None or cached_hash != _PROMPT_HASH:
                            reason = "legacy cache (no prompt_hash)" if cached_hash is None \
                                     else f"prompt changed ({cached_hash} → {_PROMPT_HASH})"
                            print(f"   • [{u['type']:8}] p{pr:<7} {leaf_ids}  — {reason}, re-extracting")
                            audit_exists = False
                            raise StopIteration
                    saved = json.load(open(os.path.join(AUDIT_DIR, u["unit_id"], "parsed.json")))
                    ext = _normalise_cell_states(Extraction(**saved))
                    # Validate cache on load — surfaces stale/poisoned results immediately
                    cached_pages = saved_meta.get("pages", u["pages"]) if os.path.exists(saved_meta_path) else u["pages"]
                    num_issues = validate_numbers(ext, args.pdf, cached_pages)
                    if num_issues:
                        print(f"   • [{u['type']:8}] p{pr:<7} {leaf_ids}  — resumed from audit "
                              f"(⚠ {len(num_issues)} number issues in cache)")
                    else:
                        print(f"   • [{u['type']:8}] p{pr:<7} {leaf_ids}  — resumed from audit (no call)")
                except StopIteration:
                    pass
                except Exception:
                    pass
                else:
                    tables = drop_next_section_tables(_apply_transforms(ext.tables), u)
                    if u["type"] == "multiple":
                        for t, lf, method, sc, flagged in route_tables(tables, u["leaves"]):
                            grp_tables[lf["section_id"]].append(t)
                    else:
                        bucket = grp_tables[u["leaves"][0]["section_id"]]
                        for t in tables:
                            if (t.continued_from_previous and bucket
                                    and len(bucket[-1].columns) == len(t.columns)
                                    and not t.title.strip()):
                                bucket[-1].rows.extend(t.rows)
                            else:
                                bucket.append(t)
                    continue
            print(f"   • [{u['type']:8}] p{pr:<7} prompt=PROMPT_{u['type'].upper():8} -> tab(s): {leaf_ids}")
            try:
                chunk_size = args.chunk_pages if args.chunk_pages > 0 else 9999
                ext, meta = extract_unit_chunked(client, args.pdf, u,
                                                  force_image=args.image,
                                                  with_thinking=args.thinking,
                                                  save_audit=not args.no_audit,
                                                  chunk_size=chunk_size)
            except Exception as e:
                print(f"     ❌ FAILED: {e}")
                continue
            tables = drop_next_section_tables(_apply_transforms(ext.tables), u)
            if u["type"] == "multiple":
                for t, lf, method, sc, flagged in route_tables(tables, u["leaves"]):
                    grp_tables[lf["section_id"]].append(t)
                    mark = "⚠ " if flagged else "  "
                    print(f"        {mark}→ [{method:5} {sc:.2f}] '{t.title[:30]}'  →  tab {lf['section_id']}")
                if len(tables) != len({lf["section_id"] for lf in u["leaves"]}):
                    print(f"        ⚠ {len(tables)} table(s) vs {len(u['leaves'])} subsection(s) on this page")
            else:
                bucket = grp_tables[u["leaves"][0]["section_id"]]
                for t in tables:
                    first_sub = next((r for r in t.rows if r.row_type not in ("note",)), None)
                    is_true_continuation = (
                        t.continued_from_previous and bucket
                        and len(bucket[-1].columns) == len(t.columns)
                        and not t.title.strip()
                        and first_sub is not None
                        and first_sub.row_type not in ("section_header", "sub_header")
                    )
                    if is_true_continuation:
                        bucket[-1].rows.extend(t.rows)
                    else:
                        bucket.append(t)
                if u["type"] == "spanning":
                    print(f"        ({len(tables)} table(s) kept across pages {'+'.join(map(str,u['pages']))})")
            ut, tag = meta["usage"], (" +img" if meta["image_used"] else "")
            print(f"     ✓ {len(tables)} table(s){tag}  "
                  f"[{ut.get('prompt_tokens','?')}in/{ut.get('output_tokens','?')}out/"
                  f"{ut.get('thinking_tokens','?')}think tok]")
        return dict(grp_tables)

    # ── Build the ordered list of groups to process ───────────────────────────
    section_tables: dict[str, list[GTable]] = defaultdict(list)
    started = args.start_section is None

    # Resolve which groups to process and in what order
    # Build ordered list of (gnum, gunits, group_leaves, leaf_target) to process
    ordered_groups: list[tuple] = []
    for gnum, gunits in groups.items():
        group_leaves = [s for s in sections if group_key(s) == gnum]
        unit_leaf_ids = {lf["section_id"] for u in gunits for lf in u["leaves"]}
        extra_leaves = [s for s in sections if s["section_id"] in unit_leaf_ids
                        and s not in group_leaves]
        group_leaves = group_leaves + extra_leaves
        leaf_ids_in_group = {s["section_id"] for s in group_leaves}

        if args.section and not (args.section == gnum or args.section in leaf_ids_in_group):
            continue
        if not started:
            if gnum == args.start_section:
                started = True
            else:
                continue
        if not args.force and not args.section and group_leaves and all(_tab_exists(s["section_id"]) for s in group_leaves):
            print(f"⏭️  group {gnum} already present — skip (use --force to redo)")
            continue

        leaf_target = args.section if (args.section and args.section in leaf_ids_in_group) else None
        ordered_groups.append((gnum, gunits, group_leaves, leaf_target))

    # ── --force: delete audit cache + Excel tabs for targeted sections ──────────
    # --section 9.4 --force: deletes only 9.4's audit folder(s) and tabs.
    # Full --force (no --section): clears all targeted groups.
    if args.force:
        if args.section:
            # --section X --force: clear only X's audit and tab.
            # Sibling sections in the same group reload from their existing cache.
            sids_to_clear = {args.section}
        else:
            # --force (no --section): clear all sections in every targeted group.
            sids_to_clear = {
                lf["section_id"]
                for _, _, group_leaves, _ in ordered_groups
                for lf in group_leaves
            }
        # Delete audit folders whose unit_id starts with any targeted section_id slug
        for sid in sids_to_clear:
            sid_slug = sid.replace(".", "_")
            audit_bank_dir = Path(AUDIT_DIR)
            if audit_bank_dir.exists():
                for unit_dir in list(audit_bank_dir.iterdir()):
                    if unit_dir.is_dir() and unit_dir.name.startswith(sid_slug):
                        import shutil
                        shutil.rmtree(unit_dir)
                        print(f"   🗑  --force deleted audit '{unit_dir.name}'")
        # Delete Excel tabs for targeted sections
        for sname in list(wb.sheetnames):
            if sname in ("Contents", "Cost"):
                continue
            for sid in sids_to_clear:
                if (sname == sid or
                        sname.startswith(f"{sid} Table ") or
                        sname.startswith(f"{sid} -")):
                    wb.remove(wb[sname])
                    used_names.discard(sname)
                    print(f"   🗑  --force cleared tab '{sname}'")
                    break

    # ── Parallel extraction: groups run concurrently, rendering is serial ─────
    # Secondary units (shared pages owned by another group) are excluded from
    # the pool — they read from audit after the pool exits, guaranteeing the
    # primary group has already written its parsed.json.
    n_workers = 1 if args.section else args.workers
    print(f"\n⚡ extracting {len(ordered_groups)} group(s) with {n_workers} worker(s)")

    group_results: dict[str, dict] = {}  # gnum -> {section_id: [GTable]}

    with ThreadPoolExecutor(max_workers=n_workers) as pool:
        futures = {
            pool.submit(_extract_group, gnum, gunits, group_leaves, leaf_target): gnum
            for gnum, gunits, group_leaves, leaf_target in ordered_groups
        }
        for fut in as_completed(futures):
            gnum = futures[fut]
            try:
                group_results[gnum] = fut.result()
            except Exception as e:
                print(f"  ❌ group {gnum} failed: {e}")
                group_results[gnum] = {}

    # ── Serial rendering: write tabs in document order ────────────────────────
    for gnum, gunits, group_leaves, leaf_target in ordered_groups:
        grp_tables = group_results.get(gnum, {})
        for sid, tables in grp_tables.items():
            section_tables[sid].extend(tables)

        print(f"\n##### Section {gnum}  ({len(gunits)} unit(s) -> {len(group_leaves)} subsection tab(s))")

        # 2) write one tab per table (not per subsection)
        for lf in group_leaves:
            sid, title = lf["section_id"], lf["title"]
            tables = section_tables.get(sid, [])

            # Sibling sections skipped by leaf_target: reload from their audit cache
            # so their existing tabs are preserved and re-rendered correctly.
            if not tables and _tab_exists(sid):
                sid_slug = sid.replace(".", "_")
                for unit_dir in Path(AUDIT_DIR).glob(f"{sid_slug}_*"):
                    pj = unit_dir / "parsed.json"
                    if pj.exists():
                        try:
                            ext = Extraction(**json.load(open(pj)))
                            tables = _apply_transforms(ext.tables)
                            print(f"   • {sid} — sibling, reloaded from audit cache")
                        except Exception:
                            pass
                        break

            if not tables:
                print(f"   · {sid} '{title[:34]}' — no tables, no tab")
                continue
            pages_str = (f"{lf['start_page']}" if lf["start_page"] == lf["end_page"]
                         else f"{lf['start_page']}-{lf['end_page']}")
            # stitch footnote-only tables onto the previous table first
            written_tables: list = []
            for t in tables:
                if not t.columns and all(r.row_type == "note" for r in t.rows):
                    if written_tables:
                        written_tables[-1].rows.extend(t.rows)
                        print(f"   · {sid} — footnote-only table stitched onto '{written_tables[-1].title[:30]}'")
                    continue
                written_tables.append(t)
            total = len(written_tables)
            # Remove ALL existing tabs for this section before writing new ones
            # so stale tabs from a previous run don't cause (2)-suffix collisions
            for existing_sname in list(wb.sheetnames):
                if (existing_sname.startswith(f"{sid} Table ") or
                        existing_sname.startswith(f"{sid} -") or
                        existing_sname == sid):
                    wb.remove(wb[existing_sname])
                    used_names.discard(existing_sname)
            for ti, t in enumerate(written_tables, start=1):
                table_label = t.title or f"Table {ti}"
                sname = table_sheet_name(used_names, sid, ti)
                ws = wb.create_sheet(title=sname)
                cursor = 4   # rows 1-3 are header banner; data starts at row 4
                cursor = write_table(ws, cursor, t)
                write_section_header(ws, sid, title, ws.max_column,
                                     table_label=table_label, table_n=ti, total_tables=total)
                style_sheet_columns(ws)
                idx = update_index(idx, {"section_id": sid, "title": title, "sheet": sname,
                                         "pages": pages_str, "first_page": int(lf["start_page"]),
                                         "table_n": ti, "n_tables": total})
                print(f"   📄 tab '{sname}'  — {t.title[:40]}")

        # remove stale duplicate tabs: (N)-suffix sheets and empty/footnote-only sheets
        stale = []
        for sname in list(wb.sheetnames):
            if sname in ("Contents", "Cost"):
                continue
            if re.search(r'\(\d+\)$', sname):
                stale.append(sname)
            else:
                ws = wb[sname]
                has_data = any(
                    ws.cell(r, c).value
                    for r in range(5, min((ws.max_row or 0) + 1, 20))
                    for c in range(5, (ws.max_column or 0) + 1)
                )
                if not has_data and (ws.max_row or 0) <= 6:
                    stale.append(sname)
        for sname in stale:
            wb.remove(wb[sname])
            used_names.discard(sname)
            print(f"   🗑  removed stale tab '{sname}'")

        save_index(idx)
        # ── Reorder all tabs into document order after each group ────────────
        _PART_SORT = {None: 0, "A": 1, "B": 2, "C": 3, "D": 4, "E": 5}
        def _sort_key(name):
            if name == "Contents": return (-1, 0, [], 0)
            if name == "Cost":     return (9999, 0, [], 0)
            m = re.match(r'^([A-Z])\.(\d+(?:\.\d+)*)(?:\s+Table\s+(\d+))?', name)
            if m:  # DBS part-structured: A.5.2 Table 1
                part  = _PART_SORT.get(m.group(1), 9)
                parts = [int(x) for x in m.group(2).split(".")]
                tnum  = int(m.group(3)) if m.group(3) else 0
                return (0, part, parts, tnum)
            m = re.match(r'^(\d+(?:\.\d+)*)(?:\s+Table\s+(\d+))?', name)
            if m:  # OCBC/UOB plain numbered: 9.4 Table 1
                parts = [int(x) for x in m.group(1).split(".")]
                tnum  = int(m.group(2)) if m.group(2) else 0
                return (0, 0, parts, tnum)
            return (1, 0, [], 0)
        ordered_names = sorted(wb.sheetnames, key=_sort_key)
        for i, sname in enumerate(ordered_names):
            current_pos = wb.sheetnames.index(sname)
            if current_pos != i:
                wb.move_sheet(sname, offset=i - current_pos)

        rebuild_contents(wb, idx)
        wb.save(out_path)
        print(f"   💾 saved {out_path}  |  run so far: {_run_usage['calls']} calls, "
              f"think={_run_usage['thinking']:,} tok, ≈ ${_run_usage['cost']:.2f}")

        if not args.no_pause and not args.section:
            ans = input("   ⏸  Review the tabs. Enter to continue, 'q' to stop: ").strip().lower()
            if ans == "q":
                print("Stopped by user.")
                break

    # Write Cost sheet into the main workbook and save the JSON summary
    if _call_log:
        write_cost_sheet(wb, _call_log, _run_usage, out_path)
        rebuild_contents(wb, idx)   # refresh Contents so Cost tab doesn't appear there
        wb.save(out_path)
        save_cost_summary(_call_log, _run_usage, out_path)
        append_to_api_log(_call_log, bank or bank_slug.upper(), out_dir)

    u = _run_usage
    print(f"\n🎉 Done. Workbook: {out_path}")
    print(f"   Total: {u['calls']} calls, input={u['prompt']:,} output={u['output']:,} "
          f"thinking={u['thinking']:,} tok, ≈ ${u['cost']:.4f}")

if __name__ == "__main__":
    main()
